"""
Academic Services
Business logic for GPA/CGPA calculation, result processing, and academic operations
Fat services pattern - all business logic here
"""
import hashlib
import json
import re
from typing import Dict, List, Tuple, Optional, Any
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.db.models import Q
from django.contrib.auth import get_user_model
from django.utils import timezone
from .models import Result, GPA, Course, ResultUploadBatch, ResultRow, Department, SemesterSummary, DepartmentBorrowedCourse
from apps.accounts.models import UserRole

User = get_user_model()


def _course_code_variants(normalized_code: str) -> List[str]:
    """Return [CSC401, CSC 401] so we match both admin-registered 'CSC 401' and file 'CSC401'."""
    if not normalized_code or len(normalized_code) < 4:
        return [normalized_code] if normalized_code else []
    m = re.match(r'^([A-Z]+)(\d{2,4})$', normalized_code)
    if m:
        return [normalized_code, f"{m.group(1)} {m.group(2)}"]
    return [normalized_code]


def _normalize_semester_for_upload(semester: str) -> str:
    """Normalize semester label to FIRST or SECOND."""
    sem_val = (semester or '').strip().upper()
    if sem_val in ('FIRST', 'SECOND'):
        return sem_val
    if '1' in sem_val or 'FIRST' in sem_val:
        return 'FIRST'
    if '2' in sem_val or 'SECOND' in sem_val:
        return 'SECOND'
    return sem_val


def _pick_course_for_upload(
    candidates: List[Course],
    department_id: Optional[int],
    borrowed_ids: set,
) -> Optional[Course]:
    if not candidates:
        return None
    if department_id is None:
        return candidates[0]
    in_dept = [c for c in candidates if getattr(c, 'department_id', None) == department_id]
    if in_dept:
        return in_dept[0]
    unassigned = [c for c in candidates if getattr(c, 'department_id', None) is None]
    if unassigned:
        return unassigned[0]
    borrowed = [c for c in candidates if c.id in borrowed_ids]
    if borrowed:
        return borrowed[0]
    return candidates[0]


def _build_course_upload_cache(department_id: Optional[int]) -> Dict[str, Optional[Course]]:
    """Load catalogue once; map every code variant to the resolved upload course."""
    courses = list(Course.objects.filter(is_active=True).select_related('department'))
    variant_candidates: Dict[str, List[Course]] = {}
    for course in courses:
        seen_ids: set = set()
        for variant in _course_code_variants((course.code or '').strip().replace(' ', '').upper()):
            if not variant:
                continue
            bucket = variant_candidates.setdefault(variant, [])
            if course.id not in seen_ids:
                bucket.append(course)
                seen_ids.add(course.id)

    borrowed_ids: set = set()
    if department_id is not None:
        borrowed_ids = set(
            DepartmentBorrowedCourse.objects.filter(department_id=department_id).values_list(
                'course_id', flat=True
            )
        )

    cache: Dict[str, Optional[Course]] = {}
    for variant, candidates in variant_candidates.items():
        cache[variant] = _pick_course_for_upload(candidates, department_id, borrowed_ids)
    return cache


def _resolve_course_from_cache(course_code: str, cache: Dict[str, Optional[Course]]) -> Optional[Course]:
    normalized = (course_code or '').strip().replace(' ', '').upper()
    if not normalized:
        return None
    for variant in _course_code_variants(normalized):
        course = cache.get(variant)
        if course is not None:
            return course
    return None


def _excel_cell_str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'nat', 'none', ''):
        return ''
    return s


def _friendly_report_error(exc: Exception) -> str:
    """Turn an exception into a short, clear message for the error report (non-technical)."""
    msg = str(exc).strip()
    if not msg:
        return 'This row could not be saved. Check that the student and course exist and the score is between 0 and 100.'
    if 'not found' in msg.lower() and 'student' in msg.lower():
        return 'Student is not in the system. Add the student in User management first.'
    if 'not in your department' in msg.lower():
        return 'Course is not in your department. Add the course in Courses and assign it to your department.'
    if 'not found' in msg.lower() and 'course' in msg.lower():
        return 'Course is not in the catalogue. Add the course under your department in Courses first.'
    if 'already exists' in msg.lower() or 'duplicate' in msg.lower():
        return 'A result for this student and course (same session and semester) already exists.'
    if 'not a student' in msg.lower() or 'is not a student' in msg.lower():
        return 'This ID is not registered as a student. Correct in User management.'
    if 'required' in msg.lower() and 'course' in msg.lower():
        return 'Course code is required. Check the format.'
    if 'unexpected keyword' in msg.lower() or 'got an unexpected' in msg.lower():
        return 'This row could not be saved. If it continues, contact support.'
    if len(msg) > 200:
        return msg[:197] + '...'
    return msg


def register_borrowed_course(department_id: Optional[int], course: Course) -> None:
    """Allow a department to use a course owned elsewhere (borrowed/service course)."""
    if not department_id or not course:
        return
    if getattr(course, 'department_id', None) == department_id:
        return
    DepartmentBorrowedCourse.objects.get_or_create(
        department_id=department_id,
        course=course,
    )


def get_course_for_upload(
    normalized_code: str,
    department_id: Optional[int] = None,
) -> Optional[Course]:
    """
    Find the course to use for result upload. Uses only the title from the course
    as registered in admin (no file title). Tries both 'CSC401' and 'CSC 401' so
    admin-registered 'CSC 401' is matched.

    When department_id is set (HOD scope): returns a course in that department,
    an unassigned catalogue match, an explicitly borrowed course, or any active
    catalogue match (auto-registered as borrowed when used for results).
    """
    variants = _course_code_variants(normalized_code)
    if not variants:
        return None
    q = Q(code__iexact=variants[0])
    if len(variants) > 1:
        q = q | Q(code__iexact=variants[1])
    qs = Course.objects.filter(is_active=True).filter(q)
    candidates = list(qs.distinct())
    if not candidates:
        return None
    if department_id is not None:
        in_dept = [c for c in candidates if getattr(c, 'department_id', None) == department_id]
        if in_dept:
            return in_dept[0]
        # Catalogue course with no department yet — HOD upload may claim it for their department.
        unassigned = [c for c in candidates if getattr(c, 'department_id', None) is None]
        if unassigned:
            return unassigned[0]
        borrowed_ids = set(
            DepartmentBorrowedCourse.objects.filter(
                department_id=department_id,
                course_id__in=[c.id for c in candidates],
            ).values_list('course_id', flat=True)
        )
        borrowed = [c for c in candidates if c.id in borrowed_ids]
        if borrowed:
            return borrowed[0]
        # Any active catalogue course may be used as a borrowed/service course
        return candidates[0]
    return candidates[0]


def resolve_course_for_upload(
    course_code: str,
    department_id: Optional[int] = None,
    *,
    claim_unassigned: bool = True,
) -> Course:
    """
    Resolve an active catalogue course for result entry.
    HOD departments may use owned, unassigned, or borrowed (cross-department) courses.
    """
    normalized_code = (course_code or '').strip().replace(' ', '').upper()
    if not normalized_code:
        raise ValueError('Course code is required.')

    course = get_course_for_upload(normalized_code, department_id=department_id)
    if not course:
        if department_id is not None:
            raise ValueError(
                f'Course "{course_code}" is not in the catalogue. '
                'Add the course in Courses first.'
            )
        raise ValueError(
            f'Course "{course_code}" not found. Add the course in the course catalogue first.'
        )

    if department_id is not None:
        course_dept_id = getattr(course, 'department_id', None)
        if course_dept_id is None and claim_unassigned:
            course.department_id = department_id
            course.save(update_fields=['department'])
        elif course_dept_id is not None and course_dept_id != department_id:
            register_borrowed_course(department_id, course)

    return course


class GPACalculationService:
    """Service for calculating GPA and CGPA using Nigerian university grading system"""
    
    # Nigerian University Grading System
    GRADE_POINTS = {
        'A': Decimal('5.0'),  # 70-100 (Excellent)
        'B': Decimal('4.0'),  # 60-69 (Very Good)
        'C': Decimal('3.0'),  # 50-59 (Good)
        'D': Decimal('2.0'),  # 45-49 (Fair)
        'E': Decimal('1.0'),  # 40-44 (Pass)
        'F': Decimal('0.0'),  # 0-39 (Fail)
    }
    
    @staticmethod
    def calculate_semester_gpa(student, session: str, semester: str) -> Dict:
        """
        Calculate semester GPA for a student
        
        Args:
            student: User instance (student)
            session: Academic session (e.g., "2023/2024")
            semester: "FIRST" or "SECOND"
            
        Returns:
            Dict with GPA, total credits, and other metrics
        """
        # Get all approved results for this semester
        results = Result.objects.filter(
            student=student,
            session=session,
            semester=semester,
            status='APPROVED'
        )
        
        if not results.exists():
            return {
                'gpa': Decimal('0.00'),
                'total_credit_units': 0,
                'total_credit_points': Decimal('0.00'),
                'courses_count': 0
            }
        
        total_credit_points = Decimal('0.00')
        total_credit_units = 0
        
        for result in results:
            credit_units = result.course.credit_units
            grade_point = result.grade_point or Decimal('0.00')
            credit_points = credit_units * grade_point
            
            total_credit_points += credit_points
            total_credit_units += credit_units
        
        # Calculate GPA: Total Credit Points / Total Credit Units
        if total_credit_units > 0:
            gpa = total_credit_points / Decimal(str(total_credit_units))
        else:
            gpa = Decimal('0.00')
        
        return {
            'gpa': gpa.quantize(Decimal('0.01')),  # Round to 2 decimal places
            'total_credit_units': total_credit_units,
            'total_credit_points': total_credit_points.quantize(Decimal('0.01')),
            'courses_count': results.count()
        }
    
    @staticmethod
    def calculate_cgpa(student) -> Dict:
        """
        Calculate Cumulative GPA (CGPA) across all semesters
        
        Args:
            student: User instance (student)
            
        Returns:
            Dict with CGPA and total metrics
        """
        # Get all approved results
        all_results = Result.objects.filter(
            student=student,
            status='APPROVED'
        )
        
        if not all_results.exists():
            return {
                'cgpa': Decimal('0.00'),
                'total_credit_units': 0,
                'total_credit_points': Decimal('0.00'),
                'total_courses': 0
            }
        
        total_credit_points = Decimal('0.00')
        total_credit_units = 0
        
        for result in all_results:
            credit_units = result.course.credit_units
            grade_point = result.grade_point or Decimal('0.00')
            credit_points = credit_units * grade_point
            
            total_credit_points += credit_points
            total_credit_units += credit_units
        
        # Calculate CGPA
        if total_credit_units > 0:
            cgpa = total_credit_points / Decimal(str(total_credit_units))
        else:
            cgpa = Decimal('0.00')
        
        return {
            'cgpa': cgpa.quantize(Decimal('0.01')),
            'total_credit_units': total_credit_units,
            'total_credit_points': total_credit_points.quantize(Decimal('0.01')),
            'total_courses': all_results.count()
        }
    
    @staticmethod
    def get_academic_standing(cgpa: Decimal) -> str:
        """
        Determine academic standing based on CGPA
        
        Args:
            cgpa: Cumulative GPA
            
        Returns:
            Academic standing string
        """
        cgpa_float = float(cgpa)
        
        if cgpa_float >= 4.50:
            return 'Excellent'
        elif cgpa_float >= 3.50:
            return 'Very Good'
        elif cgpa_float >= 2.50:
            return 'Good'
        elif cgpa_float >= 1.50:
            return 'Fair'
        elif cgpa_float >= 1.00:
            return 'Pass'
        else:
            return 'Probation'
    
    @staticmethod
    @transaction.atomic
    def update_semester_gpa(student, session: str, semester: str) -> GPA:
        """
        Calculate and save/update semester GPA record
        
        Args:
            student: User instance
            session: Academic session
            semester: "FIRST" or "SECOND"
            
        Returns:
            GPA instance
        """
        semester_data = GPACalculationService.calculate_semester_gpa(
            student, session, semester
        )
        cgpa_data = GPACalculationService.calculate_cgpa(student)
        
        # Create or update GPA record
        gpa_record, created = GPA.objects.update_or_create(
            student=student,
            session=session,
            semester=semester,
            defaults={
                'gpa': semester_data['gpa'],
                'cgpa': cgpa_data['cgpa'],
                'total_credits': semester_data['total_credit_units']
            }
        )
        
        return gpa_record


class ResultUploadService:
    """Service for handling result uploads (CSV and manual)"""
    
    @staticmethod
    def get_or_create_student(student_id: str) -> User:
        """
        Get existing student or create placeholder user for unregistered students.
        This allows uploading results before students sign up.
        
        Args:
            student_id: Student ID (matric number)
            
        Returns:
            User instance (existing or newly created placeholder)
        """
        from apps.accounts.models import UserRole
        from common.validators.student_id_validator import validate_student_id_format
        
        # Validate student ID format
        validate_student_id_format(student_id)
        
        # Try to get existing user
        try:
            student = User.objects.get(student_id=student_id)
            # Ensure it's a student role
            if student.role != UserRole.STUDENT:
                raise ValueError(f'User {student_id} exists but is not a student (role: {student.role})')
            return student
        except User.DoesNotExist:
            # Create placeholder user (inactive, no password, will be activated on signup)
            # Use a temporary email based on student_id
            placeholder_email = f"{student_id.replace('/', '_')}@placeholder.ibbul.edu.ng"
            
            # Check if placeholder email already exists (shouldn't happen, but safety check)
            if User.objects.filter(email=placeholder_email).exists():
                # Try to get by email instead
                student = User.objects.get(email=placeholder_email)
                if student.student_id != student_id:
                    raise ValueError(f'Email conflict: {placeholder_email} exists with different student_id')
                return student
            
            # Create placeholder user
            student = User.objects.create_user(
                email=placeholder_email,
                student_id=student_id,
                password=None,  # No password - must sign up to activate
                role=UserRole.STUDENT,
                is_active=False,  # Inactive until they sign up
                first_name='',
                last_name=''
            )
            return student
    
    @staticmethod
    def resolve_student_for_results(student_id: str) -> User:
        """
        Resolve a student for result upload/entry.
        Active and pending-activation (invited, not yet logged in) students are allowed.
        """
        from apps.accounts.models import UserRole
        from common.validators.student_id_validator import validate_student_id_format

        validate_student_id_format(student_id)
        sid = (student_id or '').strip().upper()
        student = User.objects.filter(student_id=sid, role=UserRole.STUDENT).first()
        if student:
            return student
        raise ValueError(
            f'Student {student_id} is not in the system. '
            'Add them via Invite or bulk CSV first — results can be saved before they activate.'
        )

    @staticmethod
    def get_student(student_id: str) -> User:
        """Alias for resolve_student_for_results (inactive/pending students allowed)."""
        return ResultUploadService.resolve_student_for_results(student_id)
    
    @staticmethod
    def validate_student_registration(student_id: str) -> User:
        """
        Validate that student exists and is registered (not admin/HOD/Examiner)
        DEPRECATED: Use get_or_create_student instead for uploads
        
        Args:
            student_id: Student ID (matric number)
            
        Returns:
            User instance
            
        Raises:
            ValueError: If student not found or is not a student
        """
        try:
            student = User.objects.get(student_id=student_id)
        except User.DoesNotExist:
            raise ValueError(f'Student with ID {student_id} is not registered')
        
        if student.role != UserRole.STUDENT:
            raise ValueError(f'User {student_id} is not a student (role: {student.role})')
        
        return student
    
    @staticmethod
    def get_or_create_course(
        course_code: str,
        course_title: Optional[str] = None,
        credit_units: Optional[int] = None,
        semester: Optional[str] = None,
        level: Optional[str] = None
    ) -> Course:
        """
        Get existing course or create new one with provided details.
        ALWAYS updates course with uploaded data to ensure student view shows exact uploaded format.
        Normalizes course codes (removes spaces) to handle variations like "CSC 300" vs "CSC300".
        
        Args:
            course_code: Course code (e.g., "CSC301" or "CSC 301")
            course_title: Course title (ALWAYS updates if provided)
            credit_units: Credit units (ALWAYS updates if provided)
            semester: Semester (updates if provided)
            level: Level (updates if provided)
            
        Returns:
            Course instance (existing or newly created)
        """
        # Normalize course code (remove spaces, uppercase)
        normalized_code = course_code.replace(' ', '').upper()
        
        try:
            # Try to find by normalized code first
            course = Course.objects.filter(code__iexact=normalized_code).first()
            if not course:
                # Try exact match
                course = Course.objects.get(code=course_code)
            
            # ALWAYS update course with uploaded data to match exact upload format
            if course_title:
                course.title = course_title  # Always use uploaded title
            if credit_units is not None:
                course.credit_units = credit_units  # Always use uploaded credit units
            if semester:
                course.semester = semester
            if level:
                course.level = level
            # Update code to normalized version if different
            if course.code != normalized_code:
                course.code = normalized_code
            course.is_active = True  # Ensure course is active
            course.save()
            return course
        except Course.DoesNotExist:
            # Create new course. Official IBBUL format often has only course codes (no title/CU).
            # Use code as placeholder title and default credit units so uploads succeed; admin can update later.
            title = (course_title or normalized_code).strip() or normalized_code
            units = credit_units if credit_units is not None else 2
            if units < 1 or units > 6:
                units = 2
            
            course = Course.objects.create(
                code=normalized_code,
                title=title,
                credit_units=units,
                semester=semester or 'FIRST',
                level=level or '300',
                is_active=True
            )
            return course
        except Course.MultipleObjectsReturned:
            # If multiple courses found, get the first one and update it
            course = Course.objects.filter(code__iexact=normalized_code).first()
            if course_title:
                course.title = course_title
            if credit_units is not None:
                course.credit_units = credit_units
            if semester:
                course.semester = semester
            if level:
                course.level = level
            if course.code != normalized_code:
                course.code = normalized_code
            course.is_active = True
            course.save()
            return course
    
    @staticmethod
    def validate_course(course_code: str, department_id: Optional[int] = None) -> Course:
        """
        Get existing course only (active). Does NOT create courses.
        Matches both "CSC303" and "CSC 303" so admin-registered code is found.
        When department_id is set (HOD), only returns a course assigned to that department.

        Args:
            course_code: Course code (e.g., "CSC301" or "CSC 303")
            department_id: Optional department scope (for HOD — only courses in this department are accepted)

        Returns:
            Course instance

        Raises:
            ValueError: If course not found or not in scope
        """
        return resolve_course_for_upload(
            course_code,
            department_id=department_id,
            claim_unassigned=True,
        )
    
    @staticmethod
    def check_duplicate_result(
        student: User,
        course: Course,
        session: str,
        semester: str
    ) -> bool:
        """
        Check if an active (non-deleted) result already exists for this student/course/session/semester.
        """
        return Result.objects.filter(
            student=student,
            course=course,
            session=session,
            semester=semester,
            is_deleted=False,
        ).exists()

    @staticmethod
    def _find_result_for_key(
        student: User,
        course: Course,
        session: str,
        semester: str,
    ) -> Optional[Result]:
        return Result.objects.filter(
            student=student,
            course=course,
            session=session,
            semester=semester,
        ).first()

    @staticmethod
    def _save_result(
        *,
        student: User,
        course: Course,
        score: Decimal,
        session: str,
        semester: str,
        uploaded_by: User,
        remark: str = '',
        department: Optional[Department] = None,
        upload_batch: Optional[ResultUploadBatch] = None,
        status: str = 'PENDING',
        grade: Optional[str] = None,
        grade_point: Optional[Decimal] = None,
    ) -> Result:
        """
        Create a result or restore a previously soft-deleted row with new values.
        """
        if department is None and uploaded_by is not None:
            from apps.accounts.scope import get_hod_department_id
            dept_id = get_hod_department_id(uploaded_by)
            if dept_id:
                department = Department.objects.filter(pk=dept_id).first()

        existing = ResultUploadService._find_result_for_key(student, course, session, semester)
        if existing:
            if not existing.is_deleted:
                raise ValueError(
                    f'Result already exists for {student.student_id} - {course.code} '
                    f'({session} {semester})'
                )
            existing.is_deleted = False
            existing.deleted_at = None
            existing.deleted_by = None
            existing.score = score
            existing.remark = remark or ''
            existing.uploaded_by = uploaded_by
            existing.status = status
            existing.department = department
            existing.upload_batch = upload_batch
            existing.approved_by = None
            existing.approved_at = None
            if grade:
                existing.grade = grade
            if grade_point is not None:
                existing.grade_point = grade_point
            existing.save()
            return existing

        create_kw = dict(
            student=student,
            course=course,
            score=score,
            session=session,
            semester=semester,
            uploaded_by=uploaded_by,
            status=status,
            remark=remark or '',
            department=department,
        )
        if upload_batch is not None:
            create_kw['upload_batch'] = upload_batch
        if grade:
            create_kw['grade'] = grade
        if grade_point is not None:
            create_kw['grade_point'] = grade_point
        return Result.objects.create(**create_kw)
    
    @staticmethod
    @transaction.atomic
    def create_result(
        student_id: str,
        course_code: str,
        score: Decimal,
        session: str,
        semester: str,
        uploaded_by: User,
        course_title: Optional[str] = None,
        credit_unit: Optional[int] = None,
        remark: Optional[str] = None,
        department: Optional[Department] = None,
    ) -> Result:
        """
        Create a single result entry

        Args:
            student_id: Student ID
            course_code: Course code
            score: Score (0-100)
            session: Academic session
            semester: "FIRST" or "SECOND"
            uploaded_by: User who uploaded (Examiner/HOD)
            course_title: Optional course title (for validation)
            credit_unit: Optional credit unit (for validation)
            remark: Optional remark (e.g., "Excellent", "Very Good")
            department: Optional department scope (for HOD)

        Returns:
            Result instance

        Raises:
            ValueError: If validation fails or duplicate exists
        """
        # Get existing student only (no placeholder creation)
        student = ResultUploadService.get_student(student_id)

        if department is not None:
            student_dept_id = getattr(student, 'department_fk_id', None)
            if student_dept_id is not None and student_dept_id != department.pk:
                raise ValueError(
                    f'Student {student_id} is not in your department. '
                    'Assign them to your department in User management first.'
                )

        course = ResultUploadService.validate_course(
            course_code,
            department_id=department.pk if department else None,
        )

        return ResultUploadService._save_result(
            student=student,
            course=course,
            score=score,
            session=session,
            semester=semester,
            uploaded_by=uploaded_by,
            remark=remark or '',
            department=department,
        )
    
    @staticmethod
    @transaction.atomic
    def bulk_create_results(
        results_data: List[Dict],
        uploaded_by: User
    ) -> Tuple[List[Result], List[str]]:
        """
        Bulk create results from CSV or manual entry
        
        Args:
            results_data: List of dicts with keys:
                - student_id (or matric_number)
                - course_code
                - score
                - session
                - semester
                - level (optional)
                - course_title (optional)
                - credit_unit (optional)
            uploaded_by: User who uploaded
            
        Returns:
            Tuple of (created_results, error_messages)
        """
        created_results = []
        errors = []

        from apps.accounts.scope import get_hod_department_id
        department_id = get_hod_department_id(uploaded_by)
        department = Department.objects.filter(pk=department_id).first() if department_id else None
        
        for idx, data in enumerate(results_data, start=1):
            try:
                # Normalize field names
                student_id = data.get('student_id') or data.get('matric_number')
                course_code = data.get('course_code')
                score = Decimal(str(data.get('score', 0)))
                session = data.get('session')
                semester = data.get('semester')
                course_title = data.get('course_title')
                credit_unit = data.get('credit_unit') or data.get('credit_units')
                
                # Validate required fields
                if not all([student_id, course_code, session, semester]):
                    errors.append(
                        f'Row {idx}: Missing required fields (student_id, course_code, session, semester)'
                    )
                    continue
                
                # Normalize semester
                semester_upper = semester.upper()
                if semester_upper not in ['FIRST', 'SECOND']:
                    # Try to map common variations
                    if '1' in semester_upper or 'FIRST' in semester_upper:
                        semester = 'FIRST'
                    elif '2' in semester_upper or 'SECOND' in semester_upper:
                        semester = 'SECOND'
                    else:
                        errors.append(f'Row {idx}: Invalid semester "{semester}"')
                        continue
                
                # Get existing student only (no creation)
                student_id_clean = (student_id or '').strip().upper()
                student = ResultUploadService.get_student(student_id_clean)

                if department is not None:
                    student_dept_id = getattr(student, 'department_fk_id', None)
                    if student_dept_id is not None and student_dept_id != department.pk:
                        errors.append(
                            f'Row {idx}: Student {student_id_clean} is not in your department.'
                        )
                        continue

                course = ResultUploadService.validate_course(
                    course_code,
                    department_id=department.pk if department else None,
                )

                result = ResultUploadService._save_result(
                    student=student,
                    course=course,
                    score=score,
                    session=session,
                    semester=semester,
                    uploaded_by=uploaded_by,
                    remark=data.get('remark', ''),
                    department=department,
                )

                created_results.append(result)
                
            except ValueError as e:
                errors.append(f'Row {idx}: {str(e)}')
            except Exception as e:
                errors.append(f'Row {idx}: Unexpected error - {str(e)}')
        
        return created_results, errors

    @staticmethod
    @transaction.atomic
    def process_upload_batch(
        filename: str,
        rows_data: List[Dict],
        uploaded_by: User,
        session: str,
        semester: str,
        department_id: Optional[int] = None,
        faculty_id: Optional[int] = None,
    ) -> Tuple[ResultUploadBatch, List[Dict]]:
        """
        Process a result upload: create batch, process each row, create ResultRow (ATTACHED or ERROR) and Result.
        Returns (batch, report_failed) where report_failed is a list of dicts with line_no, reg_number, course_code, error_message.
        """
        batch = ResultUploadBatch.objects.create(
            filename=filename,
            uploaded_by=uploaded_by,
            department_id=department_id,
            faculty_id=faculty_id,
            status=ResultUploadBatch.Status.PROCESSING,
            session=session,
            semester=semester,
        )
        report_failed = ResultUploadService._apply_rows_to_batch(
            batch, rows_data, uploaded_by, department_id, update_progress=False
        )
        return batch, report_failed

    @staticmethod
    @transaction.atomic
    def process_upload_batch_from_file(batch_id: int) -> ResultUploadBatch:
        """
        Process an existing ResultUploadBatch from its saved upload_file_path.
        Used by /api/admin/upload-results/ background upload and Celery task.
        """
        import os
        batch = ResultUploadBatch.objects.select_related('uploaded_by', 'department').get(pk=batch_id)
        file_path = batch.upload_file_path
        if not file_path or not os.path.isfile(file_path):
            batch.status = ResultUploadBatch.Status.FAILED
            batch.completed_at = timezone.now()
            batch.save(update_fields=['status', 'completed_at'])
            return batch

        try:
            rows_data = ResultUploadService._parse_upload_file_rows(
                file_path, batch.session, batch.semester
            )
        except Exception as exc:
            err_msg = f'Failed to parse upload file: {exc}'
            batch.status = ResultUploadBatch.Status.FAILED
            batch.error_count = 1
            batch.completed_at = timezone.now()
            batch.save(update_fields=['status', 'error_count', 'completed_at'])
            ResultUploadService._write_error_report_csv(batch, [{
                'line_no': 0,
                'reg_number': '',
                'course_code': '',
                'error_message': err_msg,
            }])
            return batch

        if not rows_data:
            batch.status = ResultUploadBatch.Status.FAILED
            batch.error_count = 1
            batch.completed_at = timezone.now()
            batch.save(update_fields=['status', 'error_count', 'completed_at'])
            return batch

        batch.status = ResultUploadBatch.Status.PROCESSING
        batch.progress = 0
        batch.success_count = 0
        batch.error_count = 0
        batch.completed_at = None
        batch.save(update_fields=['status', 'progress', 'success_count', 'error_count', 'completed_at'])

        report_failed = ResultUploadService._apply_rows_to_batch(
            batch,
            rows_data,
            batch.uploaded_by,
            batch.department_id,
            update_progress=True,
        )
        if report_failed:
            ResultUploadService._write_error_report_csv(batch, report_failed)
        return batch

    @staticmethod
    def _parse_upload_file_rows(file_path: str, session: str, semester: str) -> List[Dict]:
        """Parse CSV or Excel upload file into canonical row dicts."""
        import os
        import csv
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.csv':
            from .ibbul_format import normalize_column_name, map_to_canonical_columns
            rows: List[Dict] = []
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    row_normalized = {
                        normalize_column_name(k): (v.strip() if v else '')
                        for k, v in row.items()
                    }
                    data = map_to_canonical_columns(row_normalized)
                    student_id = (data.get('student_id') or '').strip().upper()
                    course_code = (data.get('course_code') or '').strip().replace(' ', '').upper()
                    if not student_id or not course_code:
                        continue
                    score_str = (data.get('score') or '').strip()
                    if not score_str:
                        continue
                    try:
                        data['score'] = float(score_str)
                    except (TypeError, ValueError):
                        continue
                    sem_val = data.get('semester') or semester
                    if sem_val:
                        sem_upper = str(sem_val).upper()
                        if '1' in sem_upper or 'FIRST' in sem_upper:
                            sem_val = 'FIRST'
                        elif '2' in sem_upper or 'SECOND' in sem_upper:
                            sem_val = 'SECOND'
                    data['student_id'] = student_id
                    data['course_code'] = course_code
                    data['session'] = data.get('session') or session
                    data['semester'] = sem_val
                    rows.append(data)
            return rows

        if ext in ('.xlsx', '.xls'):
            from .parsers.ibbul_wide import parse_ibbul_university_excel, parse_ibbul_wide_excel
            raw_rows = ResultUploadService._read_excel_raw_rows(file_path, ext)
            if not raw_rows:
                return []
            uni_results, _ = parse_ibbul_university_excel(raw_rows, session=session, semester=semester)
            if uni_results:
                return uni_results
            wide_results = parse_ibbul_wide_excel(raw_rows, session=session, semester=semester)
            return wide_results or []

        raise ValueError(f'Unsupported file format: {ext}')

    @staticmethod
    def detect_upload_session_semester(uploaded_file, raw_rows: Optional[List[List]] = None) -> Tuple[Optional[str], Optional[str]]:
        """Read session/semester printed on an official IBBUL Excel sheet header."""
        import os
        import tempfile
        from .parsers.ibbul_wide import detect_session_semester_from_sheet

        if raw_rows is not None:
            return detect_session_semester_from_sheet(raw_rows)

        ext = os.path.splitext(getattr(uploaded_file, 'name', '') or '')[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return None, None
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                for chunk in uploaded_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            raw_rows = ResultUploadService._read_excel_raw_rows(tmp_path, ext)
            return detect_session_semester_from_sheet(raw_rows)
        finally:
            if tmp_path and os.path.isfile(tmp_path):
                os.unlink(tmp_path)
            try:
                uploaded_file.seek(0)
            except Exception:
                pass

    @staticmethod
    def parse_upload_path(
        file_path: str,
        session: str,
        semester: str,
        timer=None,
    ) -> Tuple[List[Dict], List[Dict], Optional[str], Optional[str]]:
        """Parse a saved upload file once; return rows, summaries, detected session/semester."""
        import os

        ext = os.path.splitext(file_path)[1].lower()
        if ext in ('.xlsx', '.xls'):
            from .parsers.ibbul_wide import (
                parse_ibbul_university_excel,
                parse_ibbul_wide_excel,
                detect_session_semester_from_sheet,
            )
            raw_rows = ResultUploadService._read_excel_raw_rows(file_path, ext, timer=timer)
            if timer is not None:
                timer.stage('header_validation')
            detected_session, detected_semester = detect_session_semester_from_sheet(raw_rows)
            if not raw_rows:
                return [], [], detected_session, detected_semester
            if timer is not None:
                timer.stage('parse_rows')
            uni_results, summaries = parse_ibbul_university_excel(
                raw_rows, session=session, semester=semester
            )
            if uni_results:
                return uni_results, summaries, detected_session, detected_semester
            wide_results = parse_ibbul_wide_excel(raw_rows, session=session, semester=semester)
            return wide_results or [], [], detected_session, detected_semester

        rows = ResultUploadService._parse_upload_file_rows(file_path, session, semester)
        return rows, [], None, None

    @staticmethod
    def _parse_upload_file_rows_with_summaries(
        file_path: str, session: str, semester: str, timer=None
    ) -> Tuple[List[Dict], List[Dict]]:
        """Parse upload file; return (result_rows, semester_summaries from sheet)."""
        rows, summaries, _, _ = ResultUploadService.parse_upload_path(
            file_path, session, semester, timer=timer
        )
        return rows, summaries

    @staticmethod
    def parse_upload_from_uploaded_file(uploaded_file, session: str, semester: str) -> List[Dict]:
        """Parse a Django uploaded file using the same IBBUL parsers as batch upload."""
        rows, _ = ResultUploadService.parse_upload_from_uploaded_file_with_summaries(
            uploaded_file, session, semester
        )
        return rows

    @staticmethod
    def parse_upload_from_uploaded_file_with_summaries(
        uploaded_file, session: str, semester: str, timer=None
    ) -> Tuple[List[Dict], List[Dict]]:
        """Parse uploaded file; return result rows and per-student summary rows from the sheet."""
        import os
        import tempfile

        ext = os.path.splitext(getattr(uploaded_file, 'name', '') or '')[1].lower()
        if ext not in ('.csv', '.xlsx', '.xls'):
            raise ValueError('Only .csv, .xlsx, .xls are allowed.')

        tmp_path = None
        try:
            if timer is not None:
                timer.stage('save_temp_file')
            with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                for chunk in uploaded_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            rows, summaries, _, _ = ResultUploadService.parse_upload_path(
                tmp_path, session, semester, timer=timer
            )
            return rows, summaries
        finally:
            if tmp_path and os.path.isfile(tmp_path):
                os.unlink(tmp_path)
            try:
                uploaded_file.seek(0)
            except Exception:
                pass

    @staticmethod
    def _result_row_checksum(student, course, score, session: str, semester: str) -> str:
        data = {
            'student_id': student.student_id,
            'course_id': course.id,
            'score': str(score),
            'session': session,
            'semester': semester,
        }
        return hashlib.sha256(json.dumps(data, sort_keys=True).encode()).hexdigest()

    @staticmethod
    def save_semester_summaries_from_file(
        summaries: List[Dict],
        session: str,
        semester: str,
        upload_batch: Optional[ResultUploadBatch] = None,
        student_cache: Optional[Dict[str, User]] = None,
    ) -> int:
        """Persist GPA/RCU/CGPA etc. exactly as in the uploaded sheet (bulk writes)."""
        if not summaries:
            return 0

        matric_set = {
            (s.get('student_id') or '').strip().upper()
            for s in summaries
            if (s.get('student_id') or '').strip()
        }
        if student_cache is None:
            student_cache = {}
        missing = matric_set - set(student_cache.keys())
        if missing:
            for user in User.objects.filter(student_id__in=missing, role=UserRole.STUDENT):
                student_cache[user.student_id.upper()] = user

        student_ids = [u.id for u in student_cache.values() if u.student_id.upper() in matric_set]
        existing: Dict[Tuple[int, str, str], SemesterSummary] = {}
        if student_ids:
            for summary in SemesterSummary.objects.filter(student_id__in=student_ids):
                existing[(summary.student_id, summary.session, summary.semester)] = summary

        to_create: List[SemesterSummary] = []
        to_update: List[SemesterSummary] = []
        update_fields = [
            'le', 'nss', 'rcu', 'ecu', 'cp', 'gpa', 'trcu', 'tecu', 'tcp',
            'pcgpa', 'cgpa', 'outstanding_courses', 'remarks', 'standing', 'raw_summary',
            'upload_batch',
        ]

        for s in summaries:
            student_id = (s.get('student_id') or '').strip().upper()
            if not student_id:
                continue
            student = student_cache.get(student_id)
            if not student:
                continue
            sess = s.get('session') or session
            sem = s.get('semester') or semester
            field_values = {
                'le': str(s.get('le', '')),
                'nss': str(s.get('nss', '')),
                'rcu': str(s.get('rcu', '')),
                'ecu': str(s.get('ecu', '')),
                'cp': str(s.get('cp', '')),
                'gpa': str(s.get('gpa', '')),
                'trcu': str(s.get('trcu', '')),
                'tecu': str(s.get('tecu', '')),
                'tcp': str(s.get('tcp', '')),
                'pcgpa': str(s.get('pcgpa', '')),
                'cgpa': str(s.get('cgpa', '')),
                'outstanding_courses': str(s.get('outstanding_courses', '')),
                'remarks': str(s.get('remarks', '')),
                'standing': str(s.get('standing', '')),
                'raw_summary': '',
            }
            key = (student.id, sess, sem)
            if key in existing:
                obj = existing[key]
                for name, value in field_values.items():
                    setattr(obj, name, value)
                if upload_batch is not None:
                    obj.upload_batch = upload_batch
                to_update.append(obj)
            else:
                create_kw = dict(field_values, student=student, session=sess, semester=sem)
                if upload_batch is not None:
                    create_kw['upload_batch'] = upload_batch
                to_create.append(SemesterSummary(**create_kw))
                existing[key] = to_create[-1]

        if to_create:
            SemesterSummary.objects.bulk_create(to_create, batch_size=200)
        if to_update:
            SemesterSummary.objects.bulk_update(to_update, fields=update_fields, batch_size=200)
        return len(to_create) + len(to_update)

    @staticmethod
    @transaction.atomic
    def submit_hod_validated_rows(
        *,
        batch: ResultUploadBatch,
        valid_rows: List[Dict],
        invalid_rows: List[Dict],
        uploaded_by: User,
        department: Optional[Department],
        session: str,
        semester: str,
        summaries: Optional[List[Dict]] = None,
        timer=None,
    ) -> Dict[str, Any]:
        """
        Persist a validated HOD upload using bulk DB operations (no per-row queries).
        """
        session_val = (session or '').strip()
        semester_val = _normalize_semester_for_upload(semester)
        dept_id = department.id if department else None
        submit_errors: List[str] = []

        if timer is not None:
            timer.stage('student_cache')

        matric_set = {
            (row.get('matric_no') or '').strip().upper()
            for row in valid_rows
            if (row.get('matric_no') or '').strip()
        }
        student_cache = {
            u.student_id.upper(): u
            for u in User.objects.filter(
                student_id__in=matric_set,
                role=UserRole.STUDENT,
            ).select_related('department_fk')
        }

        if timer is not None:
            timer.stage('course_cache')

        course_cache = _build_course_upload_cache(dept_id)

        claim_ids: set = set()
        for row in valid_rows:
            course_code = (row.get('course_code') or '').strip().replace(' ', '').upper()
            course = _resolve_course_from_cache(course_code, course_cache)
            if course and dept_id and not course.department_id:
                claim_ids.add(course.id)
        if claim_ids:
            Course.objects.filter(id__in=claim_ids, department_id__isnull=True).update(
                department_id=dept_id
            )

        if timer is not None:
            timer.stage('existing_results_cache')

        student_ids = [s.id for s in student_cache.values()]
        existing_by_key: Dict[Tuple[int, int], Result] = {}
        if student_ids:
            for result in Result.objects.filter(
                student_id__in=student_ids,
                session=session_val,
                semester=semester_val,
            ):
                existing_by_key[(result.student_id, result.course_id)] = result

        if timer is not None:
            timer.stage('persist_results')

        to_create: List[Result] = []
        to_update: List[Result] = []
        pending_new_keys: set = set()

        for row_data in valid_rows:
            parsed = row_data.get('_parsed') or {}
            matric_no = (row_data.get('matric_no') or '').strip().upper()
            course_code = (row_data.get('course_code') or '').strip().replace(' ', '').upper()

            try:
                score = Decimal(str(row_data['score']))
            except Exception:
                submit_errors.append(f'{matric_no}/{course_code}: invalid score')
                continue

            student = student_cache.get(matric_no)
            if not student:
                submit_errors.append(f'{matric_no}/{course_code}: student not found')
                continue

            course = _resolve_course_from_cache(course_code, course_cache)
            if not course:
                submit_errors.append(f'{matric_no}/{course_code}: course not found')
                continue

            grade = (parsed.get('grade') or row_data.get('grade') or '').strip().upper()
            if grade not in ('A', 'B', 'C', 'D', 'E', 'F'):
                grade = ''

            checksum = ResultUploadService._result_row_checksum(
                student, course, score, session_val, semester_val
            )
            key = (student.id, course.id)

            if key in existing_by_key:
                result = existing_by_key[key]
                if not result.is_deleted:
                    submit_errors.append(f'{matric_no}/{course_code}: duplicate result')
                    continue
                result.is_deleted = False
                result.deleted_at = None
                result.deleted_by = None
                result.score = score
                result.status = 'HOD_REVIEW'
                result.uploaded_by = uploaded_by
                result.upload_batch = batch
                result.department = department
                result.checksum = checksum
                if grade:
                    result.grade = grade
                to_update.append(result)
                continue

            if key in pending_new_keys:
                submit_errors.append(f'{matric_no}/{course_code}: duplicate row in file')
                continue

            create_kw = dict(
                student=student,
                course=course,
                score=score,
                session=session_val,
                semester=semester_val,
                status='HOD_REVIEW',
                uploaded_by=uploaded_by,
                department=department,
                upload_batch=batch,
                checksum=checksum,
            )
            if grade:
                create_kw['grade'] = grade
            to_create.append(Result(**create_kw))
            pending_new_keys.add(key)

        if to_create:
            Result.objects.bulk_create(to_create, batch_size=500)
        if to_update:
            Result.objects.bulk_update(
                to_update,
                fields=[
                    'is_deleted', 'deleted_at', 'deleted_by', 'score', 'status',
                    'uploaded_by', 'upload_batch', 'department', 'checksum', 'grade',
                ],
                batch_size=500,
            )

        created_count = len(to_create) + len(to_update)

        if timer is not None:
            timer.stage('persist_error_rows')

        report_failed: List[Dict] = []
        error_row_models: List[ResultRow] = []
        seen_line_nos: set = set()

        for row in invalid_rows:
            parsed = row.get('_parsed') or {}
            errs = row.get('errors') or []
            line_no = int(row.get('row_number') or 0) or 0
            if line_no in seen_line_nos:
                line_no = max(seen_line_nos, default=0) + 1
            seen_line_nos.add(line_no)
            err_msg = '; '.join(errs)[:500] if errs else 'Validation failed'
            report_failed.append({
                'line_no': row.get('row_number') or '',
                'reg_number': row.get('matric_no') or parsed.get('student_id') or '',
                'course_code': row.get('course_code') or parsed.get('course_code') or '',
                'error_message': err_msg,
            })
            error_row_models.append(
                ResultRow(
                    batch=batch,
                    line_no=line_no,
                    reg_number=row.get('matric_no') or '',
                    course_code=row.get('course_code') or '',
                    status=ResultRow.RowStatus.ERROR,
                    error_message=err_msg,
                    session=session,
                    semester=semester,
                )
            )

        for err_msg in submit_errors:
            parts = err_msg.split(':', 1)
            key = parts[0] if parts else err_msg
            matric, course = (key.split('/', 1) + [''])[:2] if '/' in key else (key, '')
            report_failed.append({
                'line_no': '',
                'reg_number': matric,
                'course_code': course,
                'error_message': parts[1].strip() if len(parts) > 1 else err_msg,
            })

        if error_row_models:
            ResultRow.objects.bulk_create(error_row_models, batch_size=500)

        if timer is not None:
            timer.stage('persist_summaries')

        summaries_saved = ResultUploadService.save_semester_summaries_from_file(
            summaries or [],
            session,
            semester,
            upload_batch=batch,
            student_cache=student_cache,
        )

        batch.success_count = created_count
        batch.error_count = len(invalid_rows) + len(submit_errors)
        batch.save(update_fields=['success_count', 'error_count'])

        if report_failed:
            ResultUploadService._write_error_report_csv(batch, report_failed)

        if timer is not None:
            timer.stage('response_build')

        return {
            'created_count': created_count,
            'submit_errors': submit_errors,
            'report_failed': report_failed,
            'summaries_saved': summaries_saved,
        }

    @staticmethod
    def validate_parsed_rows(
        rows_data: List[Dict],
        session: str,
        semester: str,
        department_id: Optional[int] = None,
        timer=None,
    ) -> List[Dict]:
        """Validate canonical parsed rows. Uses bulk caches — no per-row DB queries."""
        from common.validators.student_id_validator import validate_student_id_format

        if timer is not None:
            timer.stage('student_cache')

        session_val = (session or '').strip()
        semester_val = _normalize_semester_for_upload(semester)

        matric_set = set()
        for data in rows_data:
            sid = (data.get('student_id') or data.get('matric_number') or '').strip().upper()
            if sid:
                matric_set.add(sid)

        dept = Department.objects.filter(pk=department_id).first() if department_id else None

        student_cache = {
            u.student_id.upper(): u
            for u in User.objects.filter(
                student_id__in=matric_set,
                role=UserRole.STUDENT,
            ).select_related('department_fk')
        }

        if timer is not None:
            timer.stage('course_cache')

        course_cache = _build_course_upload_cache(department_id)

        if timer is not None:
            timer.stage('duplicate_cache')

        student_ids = [s.id for s in student_cache.values()]
        existing_pairs: set = set()
        if student_ids and session_val and semester_val:
            existing_pairs = set(
                Result.objects.filter(
                    student_id__in=student_ids,
                    session=session_val,
                    semester=semester_val,
                    is_deleted=False,
                ).values_list('student_id', 'course_id')
            )

        if timer is not None:
            timer.stage('row_validation')

        report: List[Dict] = []
        for line_no, data in enumerate(rows_data, start=1):
            errors: List[str] = []
            warnings: List[str] = []

            student_id = (data.get('student_id') or data.get('matric_number') or '').strip().upper()
            course_code = (data.get('course_code') or '').strip().replace(' ', '').upper()
            row_session = (session or '').strip() or (data.get('session') or '').strip()
            row_semester = (
                _normalize_semester_for_upload(semester)
                if (semester or '').strip()
                else _normalize_semester_for_upload(data.get('semester') or '')
            )
            score_raw = data.get('score')

            if not student_id:
                errors.append('matric_no/student_id is required')
            if not course_code:
                errors.append('course_code is required')
            if score_raw is None or str(score_raw).strip() == '':
                errors.append('score is required')

            student = None
            course = None
            score = None

            if student_id and not errors:
                student = student_cache.get(student_id)
                if student is None:
                    try:
                        validate_student_id_format(student_id)
                        errors.append(
                            f'Student {student_id} is not in the system. '
                            'Add them via Invite or bulk CSV first — results can be saved before they activate.'
                        )
                    except ValueError as exc:
                        errors.append(str(exc))
                elif dept is not None:
                    student_dept_id = getattr(student, 'department_fk_id', None)
                    if student_dept_id is not None and student_dept_id != dept.id:
                        errors.append(
                            f'Student {student_id} is not in department {dept.code}. '
                            'Assign them to your department in User management.'
                        )

            if course_code and not errors:
                course = _resolve_course_from_cache(course_code, course_cache)
                if not course:
                    errors.append(
                        f'Course {course_code} is not in the catalogue. Add the course in Courses first.'
                    )

            if score_raw is not None and str(score_raw).strip() != '' and not errors:
                try:
                    score = Decimal(str(score_raw))
                    if score < 0 or score > 100:
                        errors.append('Score must be between 0 and 100')
                except (InvalidOperation, ValueError):
                    errors.append(f'Invalid score: {score_raw}')

            grade = (data.get('grade') or '').strip().upper()
            if grade and grade not in ('A', 'B', 'C', 'D', 'E', 'F'):
                warnings.append(f'Invalid grade "{grade}"; use A–F or leave blank.')

            if student and course and score is not None and not errors:
                if (student.id, course.id) in existing_pairs:
                    errors.append(
                        f'Result already saved for {student_id} · {course_code} · '
                        f'{row_session} {row_semester}. Open All Results and clear filters to view it.'
                    )

            report.append({
                'row_number': line_no,
                'matric_no': student_id,
                'course_code': course_code,
                'score': str(score) if score is not None else str(score_raw or ''),
                'grade': grade,
                'errors': errors,
                'warnings': warnings,
                'valid': len(errors) == 0,
                '_parsed': data,
            })

        return report

    @staticmethod
    def _read_excel_raw_rows(file_path: str, ext: str, timer=None) -> List[List]:
        """Read first usable Excel sheet as raw cell grid (openpyxl read_only when possible)."""
        if ext == '.xls':
            try:
                import xlrd
                if timer is not None:
                    timer.stage('open_workbook')
                wb = xlrd.open_workbook(file_path)
                sh = wb.sheet_by_index(0)
                if timer is not None:
                    timer.stage('read_worksheet')
                return [
                    [xlrd.sheet.cell_displaytext(sh, r, c) for c in range(sh.ncols)]
                    for r in range(sh.nrows)
                ]
            except Exception:
                pass  # fall through — file may be xlsx saved with .xls extension

        try:
            from openpyxl import load_workbook
            if timer is not None:
                timer.stage('open_workbook')
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if timer is not None:
                timer.stage('read_worksheet')
            rows: List[List] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([_excel_cell_str(c) for c in row])
            wb.close()
            return rows
        except Exception:
            import pandas as pd
            if timer is not None:
                timer.stage('open_workbook')
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
            if timer is not None:
                timer.stage('read_worksheet')
            return df.fillna('').values.tolist()

    @staticmethod
    def _write_error_report_csv(batch: ResultUploadBatch, report_failed: List[Dict]) -> None:
        """Write failed rows CSV and attach one-time download token to batch."""
        import os
        import csv
        import secrets
        from datetime import timedelta
        from django.conf import settings

        if not report_failed:
            return
        media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
        report_dir = os.path.join(media_root, 'reports')
        os.makedirs(report_dir, exist_ok=True)
        report_path = os.path.join(report_dir, f'{batch.id}_errors.csv')
        with open(report_path, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(
                fh, fieldnames=['line_no', 'reg_number', 'course_code', 'error_message']
            )
            writer.writeheader()
            for row in report_failed:
                writer.writerow({
                    'line_no': row.get('line_no', ''),
                    'reg_number': row.get('reg_number', ''),
                    'course_code': row.get('course_code', ''),
                    'error_message': row.get('error_message', ''),
                })
        batch.report_download_token = secrets.token_urlsafe(32)
        batch.report_download_expires_at = timezone.now() + timedelta(days=7)
        batch.save(update_fields=['report_download_token', 'report_download_expires_at'])

    @staticmethod
    def _apply_rows_to_batch(
        batch: ResultUploadBatch,
        rows_data: List[Dict],
        uploaded_by: User,
        department_id: Optional[int],
        update_progress: bool = False,
    ) -> List[Dict]:
        """Apply parsed rows to an existing batch. Returns report_failed list."""
        from apps.accounts.models import UserRole
        from common.validators.student_id_validator import validate_student_id_format

        report_failed: List[Dict] = []
        success_count = 0
        error_count = 0
        dept = Department.objects.filter(pk=department_id).first() if department_id else None
        total_rows = len(rows_data)

        for line_no, data in enumerate(rows_data, start=1):
            student_id = (data.get('student_id') or data.get('matric_number') or '').strip().upper()
            course_code = (data.get('course_code') or '').strip().replace(' ', '').upper()
            session_val = data.get('session') or batch.session
            semester_val = data.get('semester') or batch.semester
            if not semester_val or str(semester_val).upper() not in ('FIRST', 'SECOND'):
                sem_upper = str(semester_val).upper()
                if '1' in sem_upper or 'FIRST' in sem_upper:
                    semester_val = 'FIRST'
                elif '2' in sem_upper or 'SECOND' in sem_upper:
                    semester_val = 'SECOND'

            if not all([student_id, course_code, session_val, semester_val]):
                report_failed.append({
                    'line_no': line_no,
                    'reg_number': student_id or '',
                    'course_code': course_code or '',
                    'error_message': 'Missing required fields: student ID, course code, session, and semester.',
                })
                ResultRow.objects.create(
                    batch=batch,
                    line_no=line_no,
                    reg_number=student_id or '',
                    course_code=course_code or '',
                    status=ResultRow.RowStatus.ERROR,
                    error_message='Missing required fields: student ID, course code, session, and semester.',
                    session=session_val or '',
                    semester=semester_val or '',
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            try:
                score = Decimal(str(data.get('score', 0)))
            except (TypeError, ValueError, Exception):
                report_failed.append({
                    'line_no': line_no, 'reg_number': student_id, 'course_code': course_code,
                    'error_message': 'Invalid score. Enter a number between 0 and 100.',
                })
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code,
                    status=ResultRow.RowStatus.ERROR,
                    error_message='Invalid score. Enter a number between 0 and 100.',
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            if score < 0 or score > 100:
                report_failed.append({
                    'line_no': line_no, 'reg_number': student_id, 'course_code': course_code,
                    'error_message': f'Score must be between 0 and 100 (got {score}).',
                })
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR,
                    error_message=f'Score must be between 0 and 100 (got {score}).',
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            try:
                validate_student_id_format(student_id)
            except Exception as e:
                err = _friendly_report_error(e)
                if len(err) > 500:
                    err = err[:497] + '...'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err,
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            student = User.objects.filter(student_id=student_id).first()
            if not student:
                err = 'Student is not in the system. Add the student in User management first.'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err[:500],
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            if student.role != UserRole.STUDENT:
                err = 'This ID is not registered as a student. Correct in User management.'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err[:500],
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            student_dept_id = getattr(student, 'department_fk_id', None)
            if department_id is not None and student_dept_id is not None and student_dept_id != department_id:
                err = 'Student is in another department. Assign them to your department in User management or leave department blank.'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err[:500],
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            course = get_course_for_upload(course_code, department_id=department_id)
            if not course:
                err = 'Course is not in the catalogue. Add the course in Courses first.'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err[:500],
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            if department_id is not None and not course.department_id:
                course.department_id = department_id
                course.save(update_fields=['department'])

            course_dept_id = getattr(course, 'department_id', None)
            if department_id is not None and course_dept_id is not None and course_dept_id != department_id:
                register_borrowed_course(department_id, course)

            if ResultUploadService.check_duplicate_result(student, course, session_val, semester_val):
                err = 'A result for this student and course (same session and semester) already exists.'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err[:500],
                    session=session_val, semester=semester_val,
                )
                error_count += 1
                if update_progress:
                    batch.progress = line_no
                    if line_no % 10 == 0 or line_no == total_rows:
                        batch.save(update_fields=['progress'])
                continue

            try:
                grade_from_file = (data.get('grade') or '').strip().upper()
                if grade_from_file not in ('A', 'B', 'C', 'D', 'E', 'F'):
                    grade_from_file = ''
                grade_point_from_file = data.get('grade_point')
                if grade_point_from_file is not None and grade_point_from_file != '':
                    try:
                        grade_point_from_file = Decimal(str(grade_point_from_file))
                    except Exception:
                        grade_point_from_file = None
                else:
                    grade_point_from_file = None
                result = ResultUploadService._save_result(
                    student=student,
                    course=course,
                    score=score,
                    session=session_val,
                    semester=semester_val,
                    uploaded_by=uploaded_by,
                    department=dept,
                    upload_batch=batch,
                    status='PENDING',
                    remark=data.get('remark', ''),
                    grade=grade_from_file or None,
                    grade_point=grade_point_from_file,
                )
                ResultRow.objects.create(
                    batch=batch,
                    line_no=line_no,
                    reg_number=student_id,
                    course_code=course_code,
                    score=score,
                    grade=result.grade or '',
                    session=session_val,
                    semester=semester_val,
                    status=ResultRow.RowStatus.ATTACHED,
                    result=result,
                )
                success_count += 1
            except Exception as e:
                err = _friendly_report_error(e)
                if len(err) > 500:
                    err = err[:497] + '...'
                report_failed.append({'line_no': line_no, 'reg_number': student_id, 'course_code': course_code, 'error_message': err})
                ResultRow.objects.create(
                    batch=batch, line_no=line_no, reg_number=student_id, course_code=course_code, score=score,
                    status=ResultRow.RowStatus.ERROR, error_message=err,
                    session=session_val, semester=semester_val,
                )
                error_count += 1

            if update_progress:
                batch.progress = line_no
                if line_no % 10 == 0 or line_no == total_rows:
                    batch.save(update_fields=['progress'])

        batch.success_count = success_count
        batch.error_count = error_count
        batch.status = ResultUploadBatch.Status.COMPLETED
        batch.completed_at = timezone.now()
        if update_progress:
            batch.progress = total_rows
        batch.save()
        return report_failed


class BatchApprovalService:
    """
    Enterprise batch-level approval: HOD approves or rejects an entire upload batch.
    All results in the batch are updated to APPROVED or REJECTED; no per-course or per-result approval.
    """
    @staticmethod
    @transaction.atomic
    def approve_batch(batch: ResultUploadBatch, approved_by: User) -> int:
        """
        Approve the entire batch. All results in this batch are set to APPROVED.
        Returns number of results updated.
        """
        if batch.approval_status != ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            raise ValueError(f'Batch is not pending approval (current: {batch.get_approval_status_display()}).')
        if batch.status != ResultUploadBatch.Status.COMPLETED:
            raise ValueError('Only completed batches can be approved.')
        now = timezone.now()
        batch.approval_status = ResultUploadBatch.ApprovalStatus.APPROVED
        batch.approved_by = approved_by
        batch.approved_at = now
        batch.rejection_reason = ''
        batch.save()
        updated = Result.objects.filter(
            upload_batch=batch,
        ).update(
            status='APPROVED',
            approved_by=approved_by,
            approved_at=now,
            is_editable=False,
        )
        return updated

    @staticmethod
    def sync_batch_if_all_results_approved(batch, approved_by: User) -> bool:
        """
        If this batch is PENDING_APPROVAL and every result in it is already approved
        (e.g. via individual approval in Results Hub), mark the batch as APPROVED
        so it shows under "Approved" and the Pending count decreases.
        Does not change any Result rows. Returns True if batch was synced.
        """
        if batch is None or batch.status != ResultUploadBatch.Status.COMPLETED:
            return False
        if batch.approval_status != ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            return False
        total = batch.success_count or 0
        if total == 0:
            return False
        approved_count = Result.objects.filter(
            upload_batch=batch,
            status__in=['APPROVED', 'LOCKED_PUBLISHED'],
        ).count()
        if approved_count < total:
            return False
        now = timezone.now()
        batch.approval_status = ResultUploadBatch.ApprovalStatus.APPROVED
        batch.approved_by = approved_by
        batch.approved_at = now
        batch.rejection_reason = ''
        batch.save()
        return True

    @staticmethod
    @transaction.atomic
    def reject_batch(batch: ResultUploadBatch, rejected_by: User, reason: str = '') -> int:
        """
        Reject the entire batch. All results in this batch are set to REJECTED.
        Returns number of results updated.
        """
        if batch.approval_status != ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            raise ValueError(f'Batch is not pending approval (current: {batch.get_approval_status_display()}).')
        if batch.status != ResultUploadBatch.Status.COMPLETED:
            raise ValueError('Only completed batches can be rejected.')
        now = timezone.now()
        batch.approval_status = ResultUploadBatch.ApprovalStatus.REJECTED
        batch.approved_by = rejected_by
        batch.approved_at = now
        batch.rejection_reason = (reason or '')[:2000]
        batch.save()
        updated = Result.objects.filter(
            upload_batch=batch,
        ).update(
            status='REJECTED',
            approved_by=rejected_by,
            approved_at=now,
            rejection_reason=batch.rejection_reason,
        )
        return updated

    @staticmethod
    @transaction.atomic
    def unapprove_batch(batch: ResultUploadBatch, user: User) -> int:
        """
        Revert an approved batch to pending (draft). All results in the batch are set to DRAFT.
        Returns number of results updated.
        """
        if batch.approval_status != ResultUploadBatch.ApprovalStatus.APPROVED:
            raise ValueError(f'Only approved batches can be unapproved (current: {batch.get_approval_status_display()}).')
        if batch.status != ResultUploadBatch.Status.COMPLETED:
            raise ValueError('Only completed batches can be unapproved.')
        batch.approval_status = ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL
        batch.approved_by = None
        batch.approved_at = None
        batch.rejection_reason = ''
        batch.save()
        updated = Result.objects.filter(upload_batch=batch).update(
            status='DRAFT',
            approved_by=None,
            approved_at=None,
            rejection_reason='',
            is_editable=True,
        )
        return updated

    @staticmethod
    @transaction.atomic
    def reopen_batch(batch: ResultUploadBatch, user: User) -> int:
        """
        Re-open a rejected batch for approval. All results in the batch are set to DRAFT.
        Returns number of results updated.
        """
        if batch.approval_status != ResultUploadBatch.ApprovalStatus.REJECTED:
            raise ValueError(f'Only rejected batches can be re-opened (current: {batch.get_approval_status_display()}).')
        if batch.status != ResultUploadBatch.Status.COMPLETED:
            raise ValueError('Only completed batches can be re-opened.')
        batch.approval_status = ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL
        batch.approved_by = None
        batch.approved_at = None
        batch.rejection_reason = ''
        batch.save()
        updated = Result.objects.filter(upload_batch=batch).update(
            status='DRAFT',
            approved_by=None,
            approved_at=None,
            rejection_reason='',
            is_editable=True,
        )
        return updated


class ResultSummaryService:
    """Service for generating result summaries"""
    
    @staticmethod
    def get_semester_summary(student: User, session: str, semester: str) -> Dict:
        """
        Get comprehensive semester summary
        
        Returns:
            Dict with TCU, TCP, GPA, CGPA, Standing, and other metrics
        """
        # Get semester results
        results = Result.objects.filter(
            student=student,
            session=session,
            semester=semester,
            status='APPROVED'
        )
        
        # Calculate semester metrics
        semester_data = GPACalculationService.calculate_semester_gpa(
            student, session, semester
        )
        
        # Calculate CGPA
        cgpa_data = GPACalculationService.calculate_cgpa(student)
        
        # Get academic standing
        standing = GPACalculationService.get_academic_standing(cgpa_data['cgpa'])
        
        # Get all-time totals
        all_results = Result.objects.filter(
            student=student,
            status='APPROVED'
        )
        
        total_registered_credits = sum(r.course.credit_units for r in all_results)
        total_earned_credits = sum(
            r.course.credit_units for r in all_results if r.grade != 'F'
        )
        # Format decimals as exact 2-decimal strings (no float) for 100% accuracy
        gpa_dec = semester_data['gpa']
        cp_dec = semester_data['total_credit_points']
        tcp_dec = cgpa_data['total_credit_points']
        cgpa_dec = cgpa_data['cgpa']
        gpa_str = str(gpa_dec.quantize(Decimal('0.01'))) if isinstance(gpa_dec, Decimal) else str(round(float(gpa_dec), 2))
        cp_str = str(cp_dec.quantize(Decimal('0.01'))) if isinstance(cp_dec, Decimal) else str(round(float(cp_dec), 2))
        tcp_str = str(tcp_dec.quantize(Decimal('0.01'))) if isinstance(tcp_dec, Decimal) else str(round(float(tcp_dec), 2))
        cgpa_str = str(cgpa_dec.quantize(Decimal('0.01'))) if isinstance(cgpa_dec, Decimal) else str(round(float(cgpa_dec), 2))

        return {
            # Semester-specific
            'semester': semester,
            'session': session,
            'registered_credit_units': semester_data['total_credit_units'],
            'earned_credit_units': sum(
                r.course.credit_units for r in results if r.grade != 'F'
            ),
            'credit_points': cp_str,
            'gpa': gpa_str,
            'courses_count': semester_data['courses_count'],

            # Cumulative (exact string decimals)
            'total_registered_credit_units': total_registered_credits,
            'total_earned_credit_units': total_earned_credits,
            'total_credit_points': tcp_str,
            'cgpa': cgpa_str,
            'academic_standing': standing,
            'previous_cgpa': cgpa_str,

            # Additional metrics (matching screenshot format)
            'level': student.student_id.split('/')[0] if student.student_id else 'N/A',
        }
