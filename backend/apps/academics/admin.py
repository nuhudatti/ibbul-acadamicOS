"""
Result Management Admin - Production-Ready Academic System
World-Class UI/UX for Result Upload, Update, and Approval
"""
from django.contrib import admin
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import path
from django import forms
from django.http import HttpResponse, JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count, Sum
from django.utils.html import format_html
from django.utils import timezone
from django.urls import reverse
import csv
import io
from decimal import Decimal
import os
import tempfile

from django.utils.text import gettext_lazy as _
from django.contrib.admin.filters import SimpleListFilter

from .models import Course, Result, GPA, SemesterSummary, ResultUploadBatch, ResultRow, CourseAssignment, Faculty, Department
from .ibbul_format import map_to_canonical_columns, BULK_REQUIRED_HEADERS, MANUAL_COURSE_LINE_FORMAT, MANUAL_SUMMARY_FORMAT
from apps.accounts.models import User, UserRole
from apps.accounts.audit import log_audit
from apps.accounts.models import AuditLog
from apps.accounts.scope import ScopeLevel, build_scope, filter_by_scope, is_hod, get_hod_department_id
from .services import ResultUploadService, BatchApprovalService, ResultSummaryService


def _academics_scope_info(request):
    """Scope info for academics changelist banner: scope_label, is_scoped (HOD/Faculty Admin)."""
    if not getattr(request, 'user', None) or not getattr(request.user, 'is_authenticated', False):
        return {'scope_label': None, 'is_scoped': False}
    role = getattr(request.user, 'role', None)
    role_str = str(role).upper() if role else ''
    if role_str in ('DEPARTMENT_ADMIN', 'HOD'):
        dept = getattr(request.user, 'department_fk', None)
        if dept:
            name = getattr(dept, 'name', None) or getattr(dept, 'code', None) or str(dept)
            return {'scope_label': name, 'is_scoped': True, 'scope_type': 'department'}
        return {'scope_label': _('Department'), 'is_scoped': True, 'scope_type': 'department'}
    if role_str == 'FACULTY_ADMIN':
        fac = getattr(request.user, 'faculty', None)
        if fac:
            name = getattr(fac, 'name', None) or getattr(fac, 'code', None) or str(fac)
            return {'scope_label': name, 'is_scoped': True, 'scope_type': 'faculty'}
        return {'scope_label': _('Faculty'), 'is_scoped': True, 'scope_type': 'faculty'}
    return {'scope_label': None, 'is_scoped': False}


class CourseCreditUnitsListFilter(SimpleListFilter):
    """Filter courses by credit units (1–6)."""
    title = _('Credit units')
    parameter_name = 'credit_units'

    def lookups(self, request, model_admin):
        return [(i, f'{i} unit{"s" if i != 1 else ""}') for i in range(1, 7)]

    def queryset(self, request, queryset):
        value = self.value()
        if not value or not value.isdigit():
            return queryset
        return queryset.filter(credit_units=int(value))


class CourseLecturerListFilter(SimpleListFilter):
    """Filter courses by assigned lecturer (examiner)."""
    title = _('Lecturer')
    parameter_name = 'lecturer'

    def lookups(self, request, model_admin):
        qs = filter_by_scope(Course.objects.all(), request.user, request)
        examiner_ids = CourseAssignment.objects.filter(
            course__in=qs
        ).values_list('examiner_id', flat=True).distinct()
        from apps.accounts.models import User
        users = User.objects.filter(pk__in=examiner_ids).order_by('email')
        return [(str(u.pk), u.get_full_name() or u.email or str(u)) for u in users]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        return queryset.filter(examiner_assignments__examiner_id=value).distinct()


class ScopedDepartmentListFilter(SimpleListFilter):
    """For CourseAssignment: HOD sees only their department; Faculty Admin sees their faculty's departments."""
    title = _('Department')
    parameter_name = 'course__department'

    def lookups(self, request, model_admin):
        if not getattr(request, 'user', None):
            return ()
        user = request.user
        role_str = str(getattr(user, 'role', '') or '').upper()
        dept_id = get_hod_department_id(user)
        if is_hod(user) and dept_id:
            dept = Department.objects.filter(pk=dept_id).first()
            if dept:
                return ((str(dept.pk), str(dept)),)
        if role_str == 'FACULTY_ADMIN' and getattr(user, 'faculty_id', None):
            qs = Department.objects.filter(
                faculty_id=user.faculty_id,
                courses__isnull=False,
            ).distinct().order_by('code')
            return [(str(d.pk), str(d)) for d in qs]
        # Super Admin: all departments that have courses
        qs = Department.objects.filter(courses__isnull=False).distinct().order_by('faculty__code', 'code')
        return [(str(d.pk), str(d)) for d in qs]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        return queryset.filter(course__department_id=value)


class BatchDepartmentListFilter(SimpleListFilter):
    """For ResultUploadBatch: HOD sees only their dept; Faculty Admin faculty depts; Super Admin all."""
    title = _('Department')
    parameter_name = 'department'

    def lookups(self, request, model_admin):
        if not getattr(request, 'user', None):
            return ()
        user = request.user
        role_str = str(getattr(user, 'role', '') or '').upper()
        dept_id = get_hod_department_id(user)
        if is_hod(user) and dept_id:
            dept = Department.objects.filter(pk=dept_id).first()
            if dept:
                return ((str(dept.pk), str(dept)),)
        if role_str == 'FACULTY_ADMIN' and getattr(user, 'faculty_id', None):
            qs = Department.objects.filter(faculty_id=user.faculty_id).distinct().order_by('code')
            return [(str(d.pk), str(d)) for d in qs]
        qs = Department.objects.filter(id__in=ResultUploadBatch.objects.values_list('department_id', flat=True).distinct()).order_by('faculty__code', 'code')
        return [(str(d.pk), str(d)) for d in qs]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        return queryset.filter(department_id=value)


class BatchFacultyListFilter(SimpleListFilter):
    """For ResultUploadBatch: Faculty Admin sees only their faculty; Super Admin all."""
    title = _('Faculty')
    parameter_name = 'faculty'

    def lookups(self, request, model_admin):
        if not getattr(request, 'user', None):
            return ()
        user = request.user
        role_str = str(getattr(user, 'role', '') or '').upper()
        if role_str == 'FACULTY_ADMIN' and getattr(user, 'faculty_id', None):
            fac = Faculty.objects.filter(pk=user.faculty_id).first()
            if fac:
                return ((str(fac.pk), str(fac)),)
        qs = Faculty.objects.filter(id__in=ResultUploadBatch.objects.values_list('faculty_id', flat=True).distinct()).order_by('code')
        return [(str(f.pk), str(f)) for f in qs]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        return queryset.filter(faculty_id=value)


def _is_hod(request):
    """True if user is Department Admin (HOD). HOD sees department-scoped sidebar only."""
    return is_hod(request.user) if getattr(request, 'user', None) else False


def _is_examiner(request):
    """True if user is Examiner (Lecturer). Lecturer sees only Results + assigned courses."""
    if not getattr(request, 'user', None):
        return False
    return str(getattr(request.user, 'role', '')).upper() == 'EXAMINER'


def _hod_department(request):
    """Return Department for HOD (or None). Used to scope upload lists and set result.department."""
    dept_id = get_hod_department_id(request.user) if getattr(request, 'user', None) else None
    return Department.objects.filter(pk=dept_id).first() if dept_id else None


def _manual_entry_friendly_message(exc, line_num, course_code=''):
    """Convert exception to a clear, non-technical message for manual result entry."""
    msg = str(exc).strip()
    course_ref = f' ({course_code})' if course_code else ''
    if 'not found' in msg.lower() and 'student' in msg.lower():
        return f'Line {line_num}{course_ref}: This student is not in the system. Add the student in User management first.'
    if 'not in your department' in msg.lower():
        return f'Line {line_num}{course_ref}: This course is not in your department. Add the course in Courses and assign it to your department.'
    if 'not found' in msg.lower() and 'course' in msg.lower():
        return f'Line {line_num}{course_ref}: This course is not in the catalogue. Add the course under your department in Courses first.'
    if 'already exists' in msg.lower() or 'duplicate' in msg.lower():
        return f'Line {line_num}{course_ref}: A result for this student and course (same session and semester) already exists.'
    if 'not a student' in msg.lower() or 'is not a student' in msg.lower():
        return f'Line {line_num}{course_ref}: This ID is not registered as a student. Add or correct the student in User management.'
    if 'required' in msg.lower() and 'course' in msg.lower():
        return f'Line {line_num}{course_ref}: Course code is required. Check the format: course code, credit unit, grade, score, remark.'
    if 'unexpected keyword' in msg.lower() or 'got an unexpected' in msg.lower():
        return f'Line {line_num}{course_ref}: This line could not be saved. Please try again. If it continues, contact support.'
    if msg:
        return f'Line {line_num}{course_ref}: {msg}'
    return f'Line {line_num}{course_ref}: This line could not be saved. Check that the student and course exist and the score is between 0 and 100.'


def _upload_error_summary(report_failed):
    """Build a short, friendly summary of why upload rows failed (for the flash message and UI)."""
    if not report_failed:
        return ''
    categories = {
        'student_not_in_system': 0,
        'course_not_in_catalogue': 0,
        'course_not_in_scope': 0,
        'already_uploaded': 0,
        'invalid_score': 0,
        'missing_required': 0,
        'wrong_department': 0,
        'other': 0,
    }
    for item in report_failed:
        msg = (item.get('error_message') or '').strip().lower()
        if 'student' in msg and ('not in the system' in msg or 'not found' in msg or 'not registered' in msg):
            categories['student_not_in_system'] += 1
        elif 'course' in msg and ('not in the catalogue' in msg or 'not found' in msg):
            categories['course_not_in_catalogue'] += 1
        elif 'course' in msg and 'not in your department' in msg:
            categories['course_not_in_scope'] += 1
        elif 'already exists' in msg or 'already uploaded' in msg:
            categories['already_uploaded'] += 1
        elif 'score' in msg and ('0 and 100' in msg or 'invalid' in msg):
            categories['invalid_score'] += 1
        elif 'missing required' in msg or 'required fields' in msg:
            categories['missing_required'] += 1
        elif 'another department' in msg or 'different department' in msg:
            categories['wrong_department'] += 1
        else:
            categories['other'] += 1
    parts = []
    if categories['student_not_in_system']:
        n = categories['student_not_in_system']
        parts.append(f'student not in the system ({n} row{"s" if n != 1 else ""})')
    if categories['course_not_in_catalogue']:
        n = categories['course_not_in_catalogue']
        parts.append(f'course not in the catalogue ({n} row{"s" if n != 1 else ""})')
    if categories['course_not_in_scope']:
        n = categories['course_not_in_scope']
        parts.append(f'course not in your department ({n} row{"s" if n != 1 else ""})')
    if categories['already_uploaded']:
        n = categories['already_uploaded']
        parts.append(f'result already uploaded ({n} row{"s" if n != 1 else ""})')
    if categories['invalid_score']:
        n = categories['invalid_score']
        parts.append(f'invalid score ({n} row{"s" if n != 1 else ""})')
    if categories['missing_required']:
        n = categories['missing_required']
        parts.append(f'missing student ID, course, session or semester ({n} row{"s" if n != 1 else ""})')
    if categories['wrong_department']:
        n = categories['wrong_department']
        parts.append(f'student or course in another department ({n} row{"s" if n != 1 else ""})')
    if categories['other']:
        n = categories['other']
        parts.append(f'other ({n} row{"s" if n != 1 else ""})')
    if not parts:
        return 'Download the error report below to see each row and how to fix it.'
    if len(parts) == 1:
        return f'Usually because: {parts[0]}. Download the error report below to see each row and how to fix it.'
    return 'Usually because: ' + '; '.join(parts) + '. Download the error report below to see each row and how to fix it.'


class ScopeFilteredAdminMixin:
    """Module 2: Filter admin queryset by request.scope (faculty/department/examiner)."""
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        scope = getattr(request, 'scope', None)
        if scope is None:
            scope = build_scope(request.user)
        if scope is not None and scope.level < ScopeLevel.GLOBAL:
            return filter_by_scope(qs, request.user, request)
        return qs


class FileUploadForm(forms.Form):
    """Form for uploading Excel or CSV files"""
    file = forms.FileField(
        label='Excel or CSV File',
        widget=forms.FileInput(attrs={
            'accept': '.csv,.xlsx,.xls',
            'class': 'form-control'
        })
    )
    session = forms.CharField(
        max_length=20,
        initial='2023/2024',
        widget=forms.TextInput(attrs={'placeholder': 'e.g., 2023/2024', 'class': 'form-control'})
    )
    semester = forms.ChoiceField(
        choices=[('', 'Select semester'), ('FIRST', 'First semester'), ('SECOND', 'Second semester')],
        widget=forms.Select(attrs={'class': 'ur-select ur-select-native', 'required': True}),
        required=True
    )


@admin.register(Faculty)
class FacultyAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    list_display = ('code', 'name', 'is_active')
    search_fields = ('code', 'name')

    def has_module_permission(self, request):
        """Super Admin / Faculty Admin only. HOD and Examiner do not see Faculties."""
        return request.user.is_staff and not _is_hod(request) and not _is_examiner(request)

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff and not _is_hod(request)


@admin.register(Department)
class DepartmentAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    list_display = ('code', 'name', 'faculty', 'is_active')
    list_filter = ('faculty',)
    search_fields = ('code', 'name')

    def has_module_permission(self, request):
        """Super Admin / Faculty Admin only. HOD and Examiner do not see Departments."""
        return request.user.is_staff and not _is_hod(request) and not _is_examiner(request)

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff and not _is_hod(request)


# Course catalogue: must be populated before result upload (upload never creates courses)
@admin.register(Course)
class CourseAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    list_display = ('code', 'title', 'credit_units', 'semester', 'level', 'department', 'enrollment_display', 'lecturer_display', 'status_display', 'archive_action')
    list_filter = ('semester', 'level', 'is_active', 'department', CourseCreditUnitsListFilter, CourseLecturerListFilter)
    search_fields = ('code', 'title', 'level', 'department__code', 'department__name')
    ordering = ['level', 'code']
    change_list_template = 'admin/academics/course/change_list.html'
    list_per_page = 25
    actions = ['archive_courses_action', 'restore_courses_action']

    def has_module_permission(self, request):
        return request.user.is_staff

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('department').prefetch_related(
            'examiner_assignments__examiner'
        ).annotate(
            student_count=Count('results__student', distinct=True)
        )

    def enrollment_display(self, obj):
        """Number of distinct students who have a result for this course."""
        n = getattr(obj, 'student_count', 0) or 0
        return format_html('<span class="course-enrollment-badge">{}</span>', n)
    enrollment_display.short_description = _('Students enrolled')
    enrollment_display.admin_order_field = 'student_count'

    def lecturer_display(self, obj):
        """Assigned lecturer(s) or Unassigned."""
        rel = getattr(obj, 'examiner_assignments', None)
        assignments = list(rel.all()) if rel else []
        if not assignments:
            return format_html('<span class="course-lecturer-unassigned">— Unassigned</span>')
        names = []
        for a in assignments:
            ex = getattr(a, 'examiner', None)
            if ex:
                n = ex.get_full_name() or ex.email or str(ex)
                if n and n not in names:
                    names.append(n)
        if not names:
            return format_html(
                '<span class="course-lecturer-unassigned">'
                '<span class="course-lecturer-warning-icon" aria-hidden="true">⚠</span> Unassigned</span>'
            )
        return format_html('<span class="course-lecturer-names">{}</span>', ', '.join(names[:3]) + (' …' if len(names) > 3 else ''))
    lecturer_display.short_description = _('Lecturer')

    def status_display(self, obj):
        """Active or Inactive (archived) badge."""
        if obj.is_active:
            return format_html('<span class="course-status-badge course-status-active">Active</span>')
        return format_html('<span class="course-status-badge course-status-inactive">Archived</span>')
    status_display.short_description = _('Status')
    status_display.admin_order_field = 'is_active'

    def archive_action(self, obj):
        """Per-row Archive or Restore link."""
        if not obj.pk:
            return ''
        if obj.is_active:
            url = reverse('admin:academics_course_archive', args=[obj.pk])
            return format_html('<a href="{}" class="course-archive-link">Archive</a>', url)
        url = reverse('admin:academics_course_restore', args=[obj.pk])
        return format_html('<a href="{}" class="course-restore-link">Restore</a>', url)
    archive_action.short_description = _('Action')

    @admin.action(description=_('Archive selected courses'))
    def archive_courses_action(self, request, queryset):
        n = queryset.filter(is_active=True).update(is_active=False)
        if n:
            self.message_user(request, _('%s course(s) archived.') % n, messages.SUCCESS)
        else:
            self.message_user(request, _('No active courses selected.'), messages.WARNING)

    @admin.action(description=_('Restore selected courses'))
    def restore_courses_action(self, request, queryset):
        n = queryset.filter(is_active=False).update(is_active=True)
        if n:
            self.message_user(request, _('%s course(s) restored.') % n, messages.SUCCESS)
        else:
            self.message_user(request, _('No archived courses selected.'), messages.WARNING)

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff

    def has_add_permission(self, request):
        """Lecturer is view-only; HOD/Faculty/Super Admin can add courses."""
        if _is_examiner(request):
            return False
        return request.user.is_staff

    def has_change_permission(self, request, obj=None):
        """Lecturer is view-only."""
        if _is_examiner(request):
            return False
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        """Lecturer is view-only."""
        if _is_examiner(request):
            return False
        return request.user.is_staff

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['is_examiner'] = _is_examiner(request)
        extra_context['scope_info'] = _academics_scope_info(request)
        base_qs = filter_by_scope(Course.objects.all(), request.user, request)
        base_qs = base_qs.select_related('department')
        # Single aggregate: overview + level + semester (exact counts from course table)
        agg = base_qs.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True)),
            inactive=Count('id', filter=Q(is_active=False)),
            level_100=Count('id', filter=Q(level='100')),
            level_200=Count('id', filter=Q(level='200')),
            level_300=Count('id', filter=Q(level='300')),
            level_400=Count('id', filter=Q(level='400')),
            level_other=Count('id', filter=~Q(level__in=['100', '200', '300', '400'])),
            semester_first=Count('id', filter=Q(semester='FIRST')),
            semester_second=Count('id', filter=Q(semester='SECOND')),
        )
        # Course usage: courses with/without students (annotate then filter count)
        qs_with_students = base_qs.annotate(_sc=Count('results__student', distinct=True))
        courses_with_students = qs_with_students.filter(_sc__gte=1).count()
        courses_without_students = qs_with_students.filter(_sc=0).count()
        total_registrations = qs_with_students.aggregate(s=Sum('_sc'))['s'] or 0
        # Assigned / Unassigned: courses with or without lecturer
        assigned_courses = base_qs.filter(examiner_assignments__isnull=False).distinct().count()
        unassigned_courses = base_qs.annotate(_ac=Count('examiner_assignments')).filter(_ac=0).count()
        extra_context['course_stats'] = {
            'total': agg['total'] or 0,
            'active': agg['active'] or 0,
            'inactive': agg['inactive'] or 0,
            'level_100': agg['level_100'] or 0,
            'level_200': agg['level_200'] or 0,
            'level_300': agg['level_300'] or 0,
            'level_400': agg['level_400'] or 0,
            'level_other': agg['level_other'] or 0,
            'semester_first': agg['semester_first'] or 0,
            'semester_second': agg['semester_second'] or 0,
            'total_registrations': total_registrations,
            'courses_with_students': courses_with_students,
            'courses_without_students': courses_without_students,
            'assigned_courses': assigned_courses,
            'unassigned_courses': unassigned_courses,
        }
        return super().changelist_view(request, extra_context)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'title':
            kwargs['help_text'] = 'Use the full course title (e.g. "Internet", ".Net Programming", "Database Systems"). This is shown in Results; if empty or same as code, a hint to edit is shown.'
        return super().formfield_for_dbfield(db_field, request, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """HOD: restrict department to their department only."""
        if db_field.name == 'department':
            hod_dept = _hod_department(request)
            if hod_dept:
                kwargs['queryset'] = Department.objects.filter(pk=hod_dept.pk)
            else:
                kwargs['queryset'] = Department.objects.all().order_by('name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_changeform_initial_data(self, request):
        """HOD adding a course: pre-select their department so the dropdown is not blank."""
        initial = super().get_changeform_initial_data(request)
        hod_dept = _hod_department(request)
        if hod_dept and 'department' not in initial:
            initial['department'] = hod_dept.pk
        return initial

    def save_model(self, request, obj, form, change):
        """HOD adding a course: ensure department is set to their department so course appears in their list. Audit log."""
        if not change and _hod_department(request) and (not obj.department_id or obj.department_id != _hod_department(request).pk):
            obj.department = _hod_department(request)
        super().save_model(request, obj, form, change)
        action = AuditLog.Action.COURSE_UPDATED if change else AuditLog.Action.COURSE_CREATED
        identifier = f'{obj.code} - {obj.title or ""}'
        extra = {'course_id': obj.pk, 'code': obj.code, 'department_id': obj.department_id}
        log_audit(action, request=request, user=request.user, identifier=identifier, extra=extra)

    def delete_model(self, request, obj):
        """Audit course deletion."""
        identifier = f'{obj.code} - {obj.title or ""}'
        extra = {'course_id': obj.pk, 'code': obj.code, 'department_id': obj.department_id}
        log_audit(AuditLog.Action.COURSE_DELETED, request=request, user=request.user, identifier=identifier, extra=extra)
        super().delete_model(request, obj)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<path:object_id>/archive/', self.admin_site.admin_view(self.course_archive_view), name='academics_course_archive'),
            path('<path:object_id>/restore/', self.admin_site.admin_view(self.course_restore_view), name='academics_course_restore'),
        ]
        return custom + urls

    def course_archive_view(self, request, object_id):
        """Set course is_active=False and redirect to changelist."""
        obj = get_object_or_404(Course, pk=object_id)
        if not self.has_change_permission(request, obj):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        obj.is_active = False
        obj.save(update_fields=['is_active'])
        self.message_user(request, _('Course "%s" archived.') % obj.code, messages.SUCCESS)
        return redirect(reverse('admin:academics_course_changelist'))

    def course_restore_view(self, request, object_id):
        """Set course is_active=True and redirect to changelist."""
        obj = get_object_or_404(Course, pk=object_id)
        if not self.has_change_permission(request, obj):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        obj.is_active = True
        obj.save(update_fields=['is_active'])
        self.message_user(request, _('Course "%s" restored.') % obj.code, messages.SUCCESS)
        return redirect(reverse('admin:academics_course_changelist'))


# GPA model is NOT registered in admin; managed through summary uploads

@admin.register(Result)
class ResultAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    """
    Result Admin - Production-Ready Academic Result Management.
    Lecturer: view-only, scoped to assigned courses.
    """
    list_display = ('student_display', 'course_display', 'score_display', 'grade', 'session', 'semester', 'batch_status_column', 'status', 'actions_column')
    list_filter = ('status', 'semester', 'session', 'grade', 'created_at')
    search_fields = ('student__student_id', 'student__email', 'student__first_name', 'student__last_name', 'course__code', 'course__title')
    readonly_fields = ('student', 'course', 'score', 'grade', 'grade_point', 'remark', 'session', 'semester', 'status', 'uploaded_by', 'approved_by', 'approved_at', 'created_at', 'updated_at')
    list_per_page = 50

    def has_add_permission(self, request):
        """Results are created only via upload — no add form."""
        return False

    def has_change_permission(self, request, obj=None):
        """Allow opening change view for read-only display only."""
        return super().has_change_permission(request, obj)

    def get_readonly_fields(self, request, obj=None):
        """When viewing/editing a result, all fields are read-only — use upload to change data."""
        if obj:
            return list(self.readonly_fields) + [
                f.name for f in self.model._meta.fields
                if f.name not in self.readonly_fields
            ]
        return self.readonly_fields
    date_hierarchy = 'created_at'
    change_list_template = 'admin/academics/result/change_list.html'
    actions_on_top = True
    actions_on_bottom = True
    
    def has_module_permission(self, request):
        """Allow staff users to see Result module."""
        return request.user.is_staff

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff

    def changelist_view(self, request, extra_context=None):
        """Redirect to Results Hub (handled by results_hub_view via get_urls)."""
        return redirect('admin:academics_result_changelist')

    def results_hub_view(self, request):
        """Results Hub: one row per student, click to expand courses. Served at /admin/academics/result/."""
        is_exam = _is_examiner(request)
        scope_info = _academics_scope_info(request)
        try:
            hub_ctx = self._get_results_hub_context(request)
        except Exception:
            hub_ctx = {
                'hub_student_groups': [],
                'hub_sessions': [],
                'hub_status_choices': getattr(Result, 'STATUS_CHOICES', []),
                'hub_filter_session': '',
                'hub_filter_semester': '',
                'hub_filter_status': '',
                'hub_search_query': '',
                'hub_pending_batches_count': 0,
            }
        if isinstance(hub_ctx, HttpResponse):
            return hub_ctx
        context = {
            **self.admin_site.each_context(request),
            'title': _('Results'),
            'opts': self.model._meta,
            'is_examiner': is_exam,
            'show_lecturer_ui': is_exam,
            'scope_info': scope_info,
            'show_results_hub': True,
            **hub_ctx,
        }
        return render(request, 'admin/academics/result/results_hub_page.html', context)


    # Lecturer (Examiner): view-only — no add, change, or delete.
    def has_add_permission(self, request):
        if _is_examiner(request):
            return False
        return True

    def has_change_permission(self, request, obj=None):
        if _is_examiner(request):
            return False
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if _is_examiner(request):
            return False
        return super().has_delete_permission(request, obj)

    def get_actions(self, request):
        """Lecturer: no bulk actions — view-only."""
        if _is_examiner(request):
            return {}
        return super().get_actions(request)

    def add_view(self, request, form_url='', extra_context=None):
        """Redirect Add Result to upload page"""
        return redirect('admin:academics_result_upload_results')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('student', 'course', 'uploaded_by', 'approved_by', 'upload_batch', 'department')

    def _get_results_hub_context(self, request):
        """Build context for Results Hub: one row per student, expandable to courses with Approve/Edit/Comment. Handles batch approve POST."""
        base_qs = self.get_queryset(request)
        qs = base_qs
        # Apply same filters as list_filter (session, semester, status)
        session_val = request.GET.get('session', '').strip()
        semester_val = request.GET.get('semester', '').strip()
        status_val = request.GET.get('status', '').strip()
        if session_val:
            qs = qs.filter(session=session_val)
        if semester_val:
            qs = qs.filter(semester=semester_val)
        if status_val:
            qs = qs.filter(status=status_val)
        upload_batch_id = request.GET.get('upload_batch', '').strip()
        if upload_batch_id and upload_batch_id.isdigit():
            qs = qs.filter(upload_batch_id=int(upload_batch_id))
        # Search (Django changelist uses 'q' for search)
        search_query = request.GET.get('q', '').strip()
        if search_query:
            qs = qs.filter(
                Q(student__student_id__icontains=search_query)
                | Q(student__email__icontains=search_query)
                | Q(student__first_name__icontains=search_query)
                | Q(student__last_name__icontains=search_query)
                | Q(course__code__icontains=search_query)
                | Q(course__title__icontains=search_query)
            )
        # Hub batch/summary actions (single POST can have both result and summary selection)
        if request.method == 'POST' and not _is_examiner(request):
            result_ids_str = request.POST.get('result_ids', '').strip()
            summary_ids_str = request.POST.get('summary_ids', '').strip()
            result_ids = [x.strip() for x in result_ids_str.split(',') if x.strip() and x.strip().isdigit()] if result_ids_str else []
            summary_ids = [x.strip() for x in summary_ids_str.split(',') if x.strip() and x.strip().isdigit()] if summary_ids_str else []
            did_something = False
            ajax_response = {'success': True, 'message': '', 'result_ids_approved': [], 'result_ids_unapproved': [], 'summary_ids_approved': [], 'summary_ids_unapproved': [], 'batches_synced': []}
            if request.POST.get('hub_batch_approve') == '1' and result_ids:
                for pk in result_ids:
                    result = base_qs.filter(pk=pk).first()
                    if result and result.status != 'APPROVED':
                        result.status = 'APPROVED'
                        result.approved_by = request.user
                        result.approved_at = timezone.now()
                        result.save()
                        did_something = True
                        ajax_response['result_ids_approved'].append(str(result.pk))
                        log_audit(
                            AuditLog.Action.RESULT_SINGLE_APPROVED,
                            request=request, user=request.user,
                            identifier=f'{result.student.student_id or result.student.email} {result.course.code} {result.session} {result.semester}',
                            extra={'result_id': result.pk, 'student_id': result.student_id, 'course_code': result.course.code},
                        )
                if did_something:
                    messages.success(request, f'Result(s) approved.')
                    # If any approved result belongs to a batch, check if that batch is now fully approved → mark batch Approved
                    batch_ids = set()
                    for pk in result_ids:
                        r = base_qs.filter(pk=pk).first()
                        if r and getattr(r, 'upload_batch_id', None):
                            batch_ids.add(r.upload_batch_id)
                    for bid in batch_ids:
                        batch = ResultUploadBatch.objects.filter(pk=bid).first()
                        if batch and BatchApprovalService.sync_batch_if_all_results_approved(batch, request.user):
                            log_audit(
                                AuditLog.Action.RESULT_BATCH_APPROVED,
                                request=request, user=request.user,
                                identifier=batch.filename,
                                extra={'batch_id': batch.id, 'synced_from_individual': True},
                            )
                            ajax_response['batches_synced'].append(batch.id)
                            messages.success(request, f'Batch "{batch.filename}" marked approved (all results approved).')
            if request.POST.get('hub_batch_unapprove') == '1' and result_ids:
                unapproved = 0
                for pk in result_ids:
                    result = base_qs.filter(pk=pk).first()
                    if result and result.status in ('APPROVED', 'LOCKED_PUBLISHED'):
                        result.status = 'DRAFT'
                        result.approved_by = None
                        result.approved_at = None
                        result.save()
                        unapproved += 1
                        ajax_response['result_ids_unapproved'].append(str(result.pk))
                        log_audit(
                            AuditLog.Action.RESULT_UPDATED,
                            request=request, user=request.user,
                            identifier=f'{result.student.student_id or result.student.email} {result.course.code} {result.session} {result.semester}',
                            extra={'result_id': result.pk, 'action': 'unapprove', 'student_id': result.student_id, 'course_code': result.course.code},
                        )
                if unapproved:
                    messages.success(request, f'{unapproved} result(s) unapproved.')
            if request.POST.get('hub_batch_delete') == '1' and result_ids and self.has_delete_permission(request):
                deleted_count = 0
                for pk in result_ids:
                    result = base_qs.filter(pk=pk).first()
                    if result:
                        identifier = f'{result.student.student_id or result.student.email} {result.course.code} {result.session} {result.semester}'
                        log_audit(AuditLog.Action.RESULT_DELETED, request=request, user=request.user, identifier=identifier, extra={'result_id': result.pk, 'student_id': result.student_id, 'course_code': result.course.code})
                        result.delete()
                        deleted_count += 1
                if deleted_count:
                    messages.success(request, f'{deleted_count} result(s) deleted.')
            summary_qs = SemesterSummary.objects.all()
            hod_dept = _hod_department(request)
            if hod_dept:
                summary_qs = summary_qs.filter(student__department_fk_id=hod_dept.pk)
            if request.POST.get('hub_delete_summary') == '1' and summary_ids and self.has_delete_permission(request):
                deleted_summary_count = 0
                for pk in summary_ids:
                    summary = summary_qs.filter(pk=pk).first()
                    if summary:
                        identifier = f'{summary.student.student_id or summary.student.email} summary {summary.session} {summary.semester}'
                        log_audit(AuditLog.Action.RESULT_DELETED, request=request, user=request.user, identifier=identifier, extra={'semester_summary_id': summary.pk, 'student_id': summary.student_id})
                        summary.delete()
                        deleted_summary_count += 1
                if deleted_summary_count:
                    messages.success(request, f'{deleted_summary_count} summary/summaries deleted.')
            if request.POST.get('hub_approve_summary') == '1' and summary_ids:
                approved = 0
                for pk in summary_ids:
                    summary = summary_qs.filter(pk=pk).first()
                    if summary and not summary.approved:
                        summary.approved = True
                        summary.approved_by = request.user
                        summary.approved_at = timezone.now()
                        summary.save()
                        approved += 1
                        ajax_response['summary_ids_approved'].append(str(summary.pk))
                if approved:
                    messages.success(request, f'{approved} summary/summaries approved.')
            if request.POST.get('hub_unapprove_summary') == '1' and summary_ids:
                unapproved = 0
                for pk in summary_ids:
                    summary = summary_qs.filter(pk=pk).first()
                    if summary and summary.approved:
                        summary.approved = False
                        summary.approved_by = None
                        summary.approved_at = None
                        summary.save()
                        unapproved += 1
                        ajax_response['summary_ids_unapproved'].append(str(summary.pk))
                if unapproved:
                    messages.success(request, f'{unapproved} summary/summaries unapproved.')
            if request.method == 'POST' and not _is_examiner(request) and (
                request.POST.get('hub_batch_approve') == '1' or request.POST.get('hub_batch_unapprove') == '1'
                or request.POST.get('hub_batch_delete') == '1' or request.POST.get('hub_delete_summary') == '1'
                or request.POST.get('hub_approve_summary') == '1' or request.POST.get('hub_unapprove_summary') == '1'
            ):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    msgs = list(messages.get_messages(request))
                    ajax_response['message'] = msgs[0].message if msgs else 'Done.'
                    return JsonResponse(ajax_response)
                return redirect('admin:academics_result_changelist' + ('?' + request.GET.urlencode() if request.GET else ''))
        # Group by (student, session, semester) so different terms show as separate rows
        qs_ordered = qs.select_related('student', 'course', 'upload_batch').order_by('student__student_id', 'session', 'semester', 'course__code')
        student_groups_dict = {}
        for r in qs_ordered:
            key = (r.student_id, r.session, r.semester)
            if key not in student_groups_dict:
                student_groups_dict[key] = {
                    'student': r.student,
                    'results': [],
                    'status_counts': {},
                    'session': r.session,
                    'semester': r.semester,
                    'group_key': str(r.student_id) + '_' + (r.session or '').replace('/', '-') + '_' + (r.semester or ''),
                }
            student_groups_dict[key]['results'].append(r)
            status_key = 'APPROVED' if r.status in ('APPROVED', 'LOCKED_PUBLISHED') else r.status
            if status_key in ('HOD_REVIEW', 'FACULTY_REVIEW'):
                status_key = 'REVIEW'
            student_groups_dict[key]['status_counts'][status_key] = student_groups_dict[key]['status_counts'].get(status_key, 0) + 1
        # Attach SemesterSummary for each group (GPA, CGPA, RCU, ECU, remarks)
        def _fmt_two(val):
            """Format to exactly 2 decimal places (for GPA, PCGPA, CGPA only)."""
            if val is None or (isinstance(val, str) and not str(val).strip()):
                return None
            try:
                return f'{float(str(val).strip()):.2f}'
            except (TypeError, ValueError):
                return str(val).strip() or None

        def _fmt_no_trailing(val):
            """Format as integer when whole, else minimal decimals (no .00) — for RCU, ECU, CP, etc."""
            if val is None or (isinstance(val, str) and not str(val).strip()):
                return None
            try:
                f = float(str(val).strip())
                if f == int(f):
                    return str(int(f))
                return f'{f:.2f}'.rstrip('0').rstrip('.')
            except (TypeError, ValueError):
                return str(val).strip() or None

        for key, group in student_groups_dict.items():
            summary = SemesterSummary.objects.filter(
                student=group['student'],
                session=group['session'],
                semester=group['semester'],
            ).select_related('upload_batch').first()
            group['summary'] = summary
            if summary:
                group['summary_display'] = {
                    'gpa': _fmt_two(summary.gpa),
                    'pcgpa': _fmt_two(summary.pcgpa),
                    'cgpa': _fmt_two(summary.cgpa),
                    'rcu': _fmt_no_trailing(summary.rcu) or summary.rcu or '—',
                    'ecu': _fmt_no_trailing(summary.ecu) or summary.ecu or '—',
                    'cp': _fmt_no_trailing(summary.cp) or summary.cp or '—',
                    'trcu': _fmt_no_trailing(summary.trcu) or summary.trcu or '—',
                    'tecu': _fmt_no_trailing(summary.tecu) or summary.tecu or '—',
                    'tcp': _fmt_no_trailing(summary.tcp) or summary.tcp or '—',
                }
            else:
                group['summary_display'] = None
            # All courses in this group are approved (use result.status only so unapprove updates the row)
            def _result_approved(r):
                return r.status in ('APPROVED', 'LOCKED_PUBLISHED')
            n = len(group['results'])
            group['all_approved'] = n > 0 and all(_result_approved(r) for r in group['results'])
        student_groups = list(student_groups_dict.values())
        # Pending batches count for badge
        batch_qs = filter_by_scope(ResultUploadBatch.objects.all(), request.user, request)
        pending_batches_count = batch_qs.filter(
            status=ResultUploadBatch.Status.COMPLETED,
            approval_status=ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL,
        ).count()
        # Session options from scoped queryset
        hub_sessions = list(base_qs.values_list('session', flat=True).distinct().order_by('-session')[:50])
        ctx = {
            'hub_student_groups': student_groups,
            'hub_filter_session': session_val,
            'hub_filter_semester': semester_val,
            'hub_filter_status': status_val,
            'hub_search_query': search_query,
            'hub_pending_batches_count': pending_batches_count,
            'hub_sessions': hub_sessions,
            'hub_status_choices': Result.STATUS_CHOICES,
        }
        if upload_batch_id and upload_batch_id.isdigit():
            ctx['hub_upload_batch_id'] = int(upload_batch_id)
        return ctx

    def student_display(self, obj):
        """Display student with ID"""
        if obj.student.student_id:
            return f"{obj.student.student_id} - {obj.student.get_full_name()}"
        return obj.student.get_full_name()
    student_display.short_description = 'Student'
    student_display.admin_order_field = 'student__student_id'
    
    def course_display(self, obj):
        """Display course code and title (from catalogue). Hint if title missing or same as code."""
        c = obj.course
        title = (c.title or '').strip()
        code = (c.code or '').strip()
        if not title or title.upper() == code.upper():
            return format_html(
                '{} — <span style="color:#6b7280;">(edit in <a href="/admin/academics/course/{}/change/">Courses</a> to set title)</span>',
                code, c.pk
            )
        return f"{code} - {title}"
    course_display.short_description = 'Course'
    course_display.admin_order_field = 'course__code'

    def score_display(self, obj):
        """Score as uploaded (no trailing .00)."""
        return obj.get_score_display() if obj else ''
    score_display.short_description = 'Score'
    score_display.admin_order_field = 'score'

    def batch_status_column(self, obj):
        """Enterprise: show upload batch approval status; per-result approve only when no batch."""
        batch = getattr(obj, 'upload_batch', None)
        if not batch:
            return format_html('<span style="color:#6b7280;">Manual</span>')
        url = reverse('admin:academics_resultuploadbatch_change', args=[batch.pk])
        if batch.approval_status == ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            return format_html('<a href="{}">Batch #{} — Pending approval</a>', url, batch.pk)
        if batch.approval_status == ResultUploadBatch.ApprovalStatus.APPROVED:
            return format_html('<a href="{}">Batch #{} — ✓ Approved</a>', url, batch.pk)
        if batch.approval_status == ResultUploadBatch.ApprovalStatus.REJECTED:
            return format_html('<a href="{}">Batch #{} — ✗ Rejected</a>', url, batch.pk)
        return format_html('<a href="{}">Batch #{}</a>', url, batch.pk)
    batch_status_column.short_description = 'Batch'
    batch_status_column.admin_order_field = 'upload_batch__approval_status'

    def actions_column(self, obj):
        """Per-result approve only for manual entries (no batch). Batch results are approved via Upload Batches."""
        batch = getattr(obj, 'upload_batch', None)
        if batch and batch.approval_status == ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            url = reverse('admin:academics_resultuploadbatch_changelist')
            return format_html(
                '<span style="color:#6b7280;">Approve in <a href="{}">Upload Batches</a></span>',
                url,
            )
        if batch and batch.approval_status == ResultUploadBatch.ApprovalStatus.APPROVED:
            return format_html(
                '<span style="color:#059669;">✓ Approved by batch</span><br><small>{}</small>',
                obj.approved_by.get_full_name() if obj.approved_by else '—',
            )
        if batch and batch.approval_status == ResultUploadBatch.ApprovalStatus.REJECTED:
            return format_html('<span style="color:#dc2626;">✗ Rejected by batch</span>')
        if obj.status == 'PENDING':
            return format_html(
                '<div style="display: flex; gap: 8px;">'
                '<a href="{}" class="button" style="background: #10b981; color: white; padding: 6px 12px; border-radius: 6px; text-decoration: none; font-size: 12px; font-weight: 500;">✓ Approve</a>'
                '<a href="{}" class="button" style="background: #ef4444; color: white; padding: 6px 12px; border-radius: 6px; text-decoration: none; font-size: 12px; font-weight: 500;">✗ Reject</a>'
                '</div>',
                f'/admin/academics/result/{obj.id}/approve/?status=APPROVED',
                f'/admin/academics/result/{obj.id}/approve/?status=REJECTED'
            )
        elif obj.status == 'APPROVED':
            return format_html(
                '<div style="color: #10b981; font-weight: 600;">✓ Approved</div>'
                '<small style="color: #6b7280;">by {}</small>',
                obj.approved_by.get_full_name() if obj.approved_by else 'System'
            )
        else:
            return format_html('<span style="color: #ef4444; font-weight: 600;">✗ Rejected</span>')
    actions_column.short_description = 'Actions'
    actions_column.allow_tags = True
    
    fieldsets = (
        ('Student & Course', {
            'fields': ('student', 'course')
        }),
        ('Score & Grade', {
            'fields': ('score', 'grade', 'grade_point', 'remark'),
            'description': 'Display only. Values are exactly as uploaded — no calculation or editing. Add or change results only via Upload results.'
        }),
        ('Session Info', {
            'fields': ('session', 'semester', 'status')
        }),
        ('Audit Trail', {
            'fields': ('uploaded_by', 'approved_by', 'approved_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filter dropdowns by scope: HOD sees only their department's students and courses."""
        hod_dept = _hod_department(request)
        if db_field.name == "student":
            qs = User.objects.filter(
                role=UserRole.STUDENT,
                student_id__isnull=False
            ).exclude(student_id='').order_by('student_id')
            if hod_dept:
                qs = qs.filter(department_fk_id=hod_dept.pk)
            kwargs["queryset"] = qs
        elif db_field.name == "course":
            qs = kwargs.get("queryset") or Course.objects.all()
            if hod_dept:
                qs = qs.filter(department_id=hod_dept.pk)
            kwargs["queryset"] = qs
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def get_form(self, request, obj=None, **kwargs):
        """Customize form to show student_id in dropdown with better styling"""
        form = super().get_form(request, obj, **kwargs)
        
        # Customize student field to display student_id first
        if 'student' in form.base_fields:
            student_field = form.base_fields['student']
            student_field.label_from_instance = lambda obj: f"{obj.student_id} - {obj.first_name} {obj.last_name}" if obj.student_id else f"{obj.email} - {obj.first_name} {obj.last_name}"
            # Make it searchable and styled
            student_field.widget.attrs.update({
                'class': 'select2',
                'style': 'width: 100%; min-width: 300px;'
            })
        
        # Customize course field for better styling
        if 'course' in form.base_fields:
            course_field = form.base_fields['course']
            course_field.label_from_instance = lambda obj: f"{obj.code} - {obj.title} ({obj.credit_units} units)"
            course_field.widget.attrs.update({
                'class': 'select2',
                'style': 'width: 100%; min-width: 300px;'
            })
        
        return form
    
    def save_model(self, request, obj, form, change):
        """Set uploaded_by and department when saving. HOD: result scoped to their department."""
        if not change:
            obj.uploaded_by = request.user
            if not obj.status:
                obj.status = 'PENDING'
            hod_dept = _hod_department(request)
            if hod_dept and (not obj.department_id or obj.department_id == hod_dept.pk):
                obj.department = hod_dept
        super().save_model(request, obj, form, change)
        identifier = f'{obj.student.student_id or obj.student.email} {obj.course.code} {obj.session} {obj.semester}'
        extra = {'result_id': obj.pk, 'student_id': obj.student_id, 'course_code': obj.course.code, 'session': obj.session, 'semester': obj.semester}
        if change:
            log_audit(AuditLog.Action.RESULT_UPDATED, request=request, user=request.user, identifier=identifier, extra=extra)
        else:
            log_audit(AuditLog.Action.RESULT_CREATED, request=request, user=request.user, identifier=identifier, extra=extra)

    def delete_model(self, request, obj):
        """Audit result deletion (log before delete so we have full object)."""
        identifier = f'{obj.student.student_id or obj.student.email} {obj.course.code} {obj.session} {obj.semester}'
        extra = {'result_id': obj.pk, 'student_id': obj.student_id, 'course_code': obj.course.code, 'session': obj.session, 'semester': obj.semester}
        log_audit(AuditLog.Action.RESULT_DELETED, request=request, user=request.user, identifier=identifier, extra=extra)
        super().delete_model(request, obj)
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            # Serve Results Hub (one row per student, expand = courses) at the main list URL
            path('', self.admin_site.admin_view(self.results_hub_view), name='academics_result_changelist'),
            path('upload-results/', self.admin_site.admin_view(self.upload_results_view), name='academics_result_upload_results'),
            path('upload-results/template/', self.admin_site.admin_view(self.download_template_view), name='academics_result_download_template'),
            path('upload-results/report/<int:batch_id>/', self.admin_site.admin_view(self.import_report_download_view), name='academics_result_import_report'),
            path('upload-results/manual-report/', self.admin_site.admin_view(self.manual_entry_report_download_view), name='academics_result_manual_entry_report'),
            path('student-summary/', self.admin_site.admin_view(self.student_summary_view), name='academics_result_student_summary'),
            path('export-ibbul-sheet/', self.admin_site.admin_view(self.export_ibbul_result_sheet_view), name='academics_result_export_ibbul_sheet'),
            path('<path:object_id>/approve/', self.admin_site.admin_view(self.approve_result_view), name='academics_result_approve'),
            path('<path:object_id>/unapprove/', self.admin_site.admin_view(self.unapprove_result_view), name='academics_result_unapprove'),
        ]
        return custom_urls + urls

    def download_template_view(self, request):
        """Download CSV template for result uploads. Production-accurate IBBUL format."""
        if not request.user.is_staff:
            return HttpResponse('Forbidden', status=403)

        template_type = request.GET.get('type', 'complete')  # 'minimal' or 'complete'

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="IBBUL_result_upload_template_{template_type}.csv"'
        writer = csv.writer(response)

        if template_type == 'minimal':
            # Required columns only — exact headers the system expects
            writer.writerow(['matric_number', 'course_code', 'score', 'session', 'semester'])
            writer.writerow(['U22/FNS/CSC/0001', 'CSC301', '75', '2023/2024', 'FIRST'])
            writer.writerow(['U22/FNS/CSC/0001', 'CSC302', '68', '2023/2024', 'FIRST'])
            writer.writerow(['U22/FNS/CSC/0002', 'CSC301', '82', '2023/2024', 'FIRST'])
            writer.writerow(['U22/FNS/CSC/0002', 'CSC302', '71', '2023/2024', 'FIRST'])
        else:
            # Full template — all columns including Outstanding courses and Remarks (IBBUL official order)
            writer.writerow([
                'matric_number', 'course_code', 'course_title', 'credit_unit', 'score', 'grade',
                'level', 'session', 'semester',
                'le', 'nss', 'rcu', 'ecu', 'cp', 'gpa', 'trcu', 'tecu', 'tcp', 'pcgpa', 'cgpa',
                'outstanding_courses', 'remarks', 'standing'
            ])
            writer.writerow([
                'U22/FNS/CSC/0001', 'CSC301', '.Net Programming', '3', '75', 'A',
                '300', '2023/2024', 'FIRST',
                '300', '5', '15', '12', '48', '4.0', '60', '55', '220', '3.8', '3.9',
                '', 'Good', 'Good'
            ])
            writer.writerow([
                'U22/FNS/CSC/0001', 'CSC302', 'Data Structures', '3', '68', 'B',
                '300', '2023/2024', 'FIRST',
                '300', '5', '15', '12', '48', '4.0', '60', '55', '220', '3.8', '3.9',
                '', 'Good', 'Good'
            ])
            writer.writerow([
                'U22/FNS/CSC/0002', 'CSC301', '.Net Programming', '3', '82', 'A',
                '300', '2023/2024', 'FIRST',
                '300', '5', '15', '15', '60', '4.0', '60', '60', '240', '3.9', '4.0',
                '', 'Excellent', 'Excellent'
            ])
            writer.writerow([
                'U22/FNS/CSC/0002', 'CSC302', 'Data Structures', '3', '79', 'B',
                '300', '2023/2024', 'FIRST',
                '300', '5', '15', '15', '60', '4.0', '60', '60', '240', '3.9', '4.0',
                '', 'Excellent', 'Excellent'
            ])

        log_audit(
            AuditLog.Action.RESULT_UPLOAD_STARTED,
            request=request,
            user=request.user,
            identifier=f'template_{template_type}',
            extra={'action': 'template_download', 'template_type': template_type},
        )
        return response

    def export_ibbul_result_sheet_view(self, request):
        """Export a single student's result for session/semester as Excel in exact IBBUL official format (same layout as Untitled.xls)."""
        if not request.user.is_staff:
            return HttpResponse('Forbidden', status=403)
        student_id = request.GET.get('student_id', '').strip().upper()
        session_val = request.GET.get('session', '').strip()
        semester_val = request.GET.get('semester', '').strip()
        if not student_id or not session_val or not semester_val:
            messages.error(request, 'Required: student_id, session, semester.')
            return redirect('admin:academics_result_changelist')
        if semester_val.upper() not in ('FIRST', 'SECOND'):
            if '1' in semester_val.upper() or 'FIRST' in semester_val.upper():
                semester_val = 'FIRST'
            elif '2' in semester_val.upper() or 'SECOND' in semester_val.upper():
                semester_val = 'SECOND'
        try:
            student = User.objects.get(student_id=student_id, role=UserRole.STUDENT)
        except User.DoesNotExist:
            messages.error(request, f'Student not found: {student_id}')
            return redirect('admin:academics_result_changelist')
        results = Result.objects.filter(
            student=student, session=session_val, semester=semester_val
        ).select_related('course').order_by('course__code')
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl. pip install openpyxl')
            return redirect('admin:academics_result_changelist')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Result'
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        # Row 1: University
        ws.append(['IBBUL - Ibrahim Badamasi Babangida University, Lapai'])
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=12)
        # Row 2: Faculty, Department, Session, Semester, Level
        dept = getattr(student, 'department_fk', None)
        faculty = getattr(dept, 'faculty', None) if dept else None
        ws.append([
            f"Faculty: {getattr(faculty, 'name', '') or ''}",
            f"Department: {getattr(dept, 'name', '') or ''}",
            f"Session: {session_val}",
            f"Semester: {semester_val}",
            f"Level: {getattr(student, 'level', '') or student_id.split('/')[0] if student_id else ''}",
        ])
        ws.merge_cells('A2:E2')
        # Row 3: Student
        ws.append(['Reg. No.', 'Name', 'Level'])
        for c in range(1, 4):
            ws.cell(row=3, column=c).font = header_font
        ws.append([
            student_id,
            student.get_full_name() or '',
            getattr(student, 'level', '') or (student_id.split('/')[0] if student_id else ''),
        ])
        # Row 5: Course table header (IBBUL exact)
        ws.append(['S/N', 'Course Code', 'Course Title', 'Credit Unit', 'Score', 'Grade', 'Grade Point', 'Remark'])
        for c in range(1, 9):
            ws.cell(row=5, column=c).font = header_font
        row_num = 6
        for sn, r in enumerate(results, 1):
            ws.append([
                sn,
                r.course.code,
                r.course.title or '',
                r.course.credit_units or '',
                r.get_score_display() or '',
                r.grade or '',
                float(r.grade_point) if r.grade_point is not None else '',
                r.remark or '',
            ])
            row_num += 1
        # Summary row (IBBUL: LE, NSS, RCU, ECU, CP, GPA, TRCU, TECU, TCP, PCGPA, CGPA, Standing)
        summary = SemesterSummary.objects.filter(
            student=student, session=session_val, semester=semester_val
        ).first()
        if summary:
            ws.append(['Summary:', summary.le, summary.nss, summary.rcu, summary.ecu, summary.cp, summary.gpa,
                       summary.trcu, summary.tecu, summary.tcp, summary.pcgpa, summary.cgpa, summary.standing])
        else:
            from decimal import Decimal
            from .services import GPACalculationService
            gpa_data = GPACalculationService.calculate_semester_gpa(student, session_val, semester_val)
            cgpa_data = GPACalculationService.calculate_cgpa(student)
            standing = GPACalculationService.get_academic_standing(cgpa_data['cgpa'])
            rcu = gpa_data.get('total_credit_units', 0)
            cp = gpa_data.get('total_credit_points', 0)
            gpa = gpa_data.get('gpa', 0)
            nss = gpa_data.get('courses_count', 0)
            # Exact 2 dp for export (no float representation errors)
            def _two_dp(v):
                if v is None: return 0
                d = v if isinstance(v, Decimal) else Decimal(str(v))
                return float(d.quantize(Decimal('0.01')))
            ws.append([
                'Summary:', rcu, nss, rcu, rcu, _two_dp(cp), _two_dp(gpa),
                cgpa_data.get('total_credit_units', 0), cgpa_data.get('total_credit_units', 0),
                _two_dp(cgpa_data.get('total_credit_points', 0)), '', _two_dp(cgpa_data.get('cgpa', 0)), standing,
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="IBBUL_Result_{student_id}_{session_val}_{semester_val}.xlsx"'
        return response

    def import_report_download_view(self, request, batch_id):
        """Download import report CSV (failed rows only). Scoped: HOD only their department batches."""
        batch_qs = filter_by_scope(ResultUploadBatch.objects.all(), request.user, request)
        batch = get_object_or_404(batch_qs, pk=batch_id)
        if not request.user.is_staff:
            return HttpResponse('Forbidden', status=403)
        failed_rows = ResultRow.objects.filter(batch=batch, status=ResultRow.RowStatus.ERROR).order_by('line_no')
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['Row', 'Student ID', 'Course', 'Score', 'What went wrong'])
        for row in failed_rows:
            writer.writerow([row.line_no, row.reg_number, row.course_code, getattr(row, 'score', '') or '', row.error_message or ''])
        log_audit(
            AuditLog.Action.RESULT_IMPORT_REPORT_DOWNLOAD,
            request=request,
            user=request.user,
            identifier=batch.filename,
            extra={'batch_id': batch_id, 'failed_count': failed_rows.count()},
        )
        resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="import_report_batch_{batch_id}.csv"'
        return resp

    def manual_entry_report_download_view(self, request):
        """Download error report CSV for the last manual entry that had failures (same format as bulk report)."""
        if not request.user.is_staff:
            return HttpResponse('Forbidden', status=403)
        report = request.session.get('manual_entry_last_report', [])
        if not report:
            messages.info(request, 'No manual entry error report available. Submit a manual entry with some failed rows first.')
            return redirect('admin:academics_result_upload_results')
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['Row', 'Student ID', 'Course', 'Score', 'What went wrong'])
        student_id = request.session.get('manual_entry_report_student_id', '')
        for row in report:
            writer.writerow([
                row.get('line_no', ''),
                student_id,
                row.get('course_code', ''),
                row.get('score', ''),
                row.get('error_message', ''),
            ])
        resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="manual_entry_error_report.csv"'
        return resp

    def student_summary_view(self, request):
        """View summary table (RCU, ECU, CP, GPA, TRCU, TECU, TCP, CGPA, REMARKS) for a student — each student views all of his."""
        if not request.user.is_staff:
            return HttpResponse('Forbidden', status=403)
        students_qs = User.objects.filter(role=UserRole.STUDENT, student_id__isnull=False).exclude(student_id='').order_by('student_id')
        hod_dept = _hod_department(request)
        if hod_dept:
            students_qs = students_qs.filter(department_fk_id=hod_dept.pk)
        students = list(students_qs)
        student_id = (request.GET.get('student_id') or request.POST.get('student_id') or '').strip().upper()
        session_val = (request.GET.get('session') or request.POST.get('session') or '2023/2024').strip()
        semester_val = (request.GET.get('semester') or request.POST.get('semester') or 'FIRST').strip()
        if str(semester_val).upper() not in ('FIRST', 'SECOND'):
            if '1' in str(semester_val).upper() or 'FIRST' in str(semester_val).upper():
                semester_val = 'FIRST'
            else:
                semester_val = 'SECOND'
        summary_data = None
        results_list = []
        stored_summary = None
        if student_id:
            student = User.objects.filter(student_id=student_id, role=UserRole.STUDENT).first()
            if student and (not hod_dept or getattr(student, 'department_fk_id', None) == hod_dept.pk):
                results_list = list(
                    Result.objects.filter(student=student, session=session_val, semester=semester_val)
                    .select_related('course').order_by('course__code')
                )
                stored_summary = SemesterSummary.objects.filter(
                    student=student, session=session_val, semester=semester_val
                ).first()
                # Use only stored summary from upload (no calculation). OUTSTANDING COURSES and REMARKS from file.
                if stored_summary:
                    def _two(s):
                        if s is None or not str(s).strip(): return ''
                        try: return f'{float(s):.2f}'
                        except (TypeError, ValueError): return str(s).strip()
                    summary_data = {
                        'registered_credit_units': stored_summary.rcu,
                        'earned_credit_units': stored_summary.ecu,
                        'credit_points': stored_summary.cp,
                        'gpa': _two(stored_summary.gpa),
                        'total_registered_credit_units': stored_summary.trcu,
                        'total_earned_credit_units': stored_summary.tecu,
                        'total_credit_points': stored_summary.tcp,
                        'cgpa': _two(stored_summary.cgpa),
                        'outstanding_courses': getattr(stored_summary, 'outstanding_courses', '') or '',
                        'remarks': getattr(stored_summary, 'remarks', '') or '',
                    }
                else:
                    summary_data = {
                        'registered_credit_units': '', 'earned_credit_units': '', 'credit_points': '',
                        'gpa': '', 'total_registered_credit_units': '', 'total_earned_credit_units': '',
                        'total_credit_points': '', 'cgpa': '', 'outstanding_courses': '', 'remarks': '',
                    }
            else:
                summary_data = {}
                summary_data['remarks'] = 'Student not found or not in your scope.'
        context = {
            **self.admin_site.each_context(request),
            'title': 'Student result summary',
            'opts': self.model._meta,
            'students': students,
            'student_id': student_id,
            'session_val': session_val,
            'semester_val': semester_val,
            'summary_data': summary_data,
            'results_list': results_list,
            'stored_summary': stored_summary,
        }
        return render(request, 'admin/academics/student_summary.html', context)

    def approve_result_view(self, request, object_id):
        """Approve or reject a result. Scoped: HOD can only approve results in their department."""
        qs = self.get_queryset(request)
        result = get_object_or_404(qs, pk=object_id)
        status = request.GET.get('status', 'APPROVED')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if status not in ['APPROVED', 'REJECTED']:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
            messages.error(request, 'Invalid status')
            return redirect('admin:academics_result_changelist')
        result.status = status
        result.approved_by = request.user
        result.approved_at = timezone.now()
        result.save()
        action_type = AuditLog.Action.RESULT_SINGLE_APPROVED if status == 'APPROVED' else AuditLog.Action.RESULT_SINGLE_REJECTED
        log_audit(
            action_type,
            request=request,
            user=request.user,
            identifier=f'{result.student.student_id or result.student.email} {result.course.code} {result.session} {result.semester}',
            extra={'result_id': result.pk, 'student_id': result.student_id, 'course_code': result.course.code, 'session': result.session, 'semester': result.semester},
        )
        if status == 'APPROVED' and getattr(result, 'upload_batch_id', None):
            batch = ResultUploadBatch.objects.filter(pk=result.upload_batch_id).first()
            if batch and BatchApprovalService.sync_batch_if_all_results_approved(batch, request.user):
                log_audit(
                    AuditLog.Action.RESULT_BATCH_APPROVED,
                    request=request, user=request.user,
                    identifier=batch.filename,
                    extra={'batch_id': batch.id, 'synced_from_individual': True},
                )
                messages.success(request, 'Result approved. Batch marked approved (all results in batch approved).')
            else:
                messages.success(request, 'Result approved successfully.')
        else:
            messages.success(request, f'Result {status.lower()} successfully.')
        if is_ajax:
            return JsonResponse({'success': True, 'result_id': result.pk, 'status': status})
        return redirect('admin:academics_result_changelist')
    
    def unapprove_result_view(self, request, object_id):
        """Unapprove a result: set to DRAFT and clear approved_by/approved_at. Scoped like approve."""
        qs = self.get_queryset(request)
        result = get_object_or_404(qs, pk=object_id)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if result.status not in ('APPROVED', 'LOCKED_PUBLISHED'):
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'Result was not approved.'}, status=400)
            messages.warning(request, 'Result was not approved.')
            return redirect('admin:academics_result_changelist' + ('?' + request.GET.urlencode() if request.GET else ''))
        result.status = 'DRAFT'
        result.approved_by = None
        result.approved_at = None
        result.save()
        log_audit(
            AuditLog.Action.RESULT_UPDATED,
            request=request,
            user=request.user,
            identifier=f'{result.student.student_id or result.student.email} {result.course.code} {result.session} {result.semester}',
            extra={'result_id': result.pk, 'action': 'unapprove', 'student_id': result.student_id, 'course_code': result.course.code},
        )
        messages.success(request, 'Result unapproved and set to draft.')
        if is_ajax:
            return JsonResponse({'success': True, 'result_id': result.pk, 'status': 'DRAFT'})
        return redirect('admin:academics_result_changelist' + ('?' + request.GET.urlencode() if request.GET else ''))
    
    def upload_results_view(self, request):
        """Upload Results page. HOD sees only their department's students and uploads scoped to their department."""
        file_form = FileUploadForm()
        students_qs = User.objects.filter(role=UserRole.STUDENT, student_id__isnull=False).exclude(student_id='').order_by('student_id')
        hod_dept = _hod_department(request)
        if hod_dept:
            students_qs = students_qs.filter(department_fk_id=hod_dept.pk)
        students = list(students_qs)
        
        if request.method == 'POST':
            if 'file_upload' in request.POST:
                file_form = FileUploadForm(request.POST, request.FILES)
                if not file_form.is_valid():
                    if not request.FILES.get('file'):
                        messages.error(request, 'Please select a file to upload. No file was received.')
                    context = {
                        **self.admin_site.each_context(request),
                        'file_form': file_form,
                        'students': students,
                        'title': 'Upload Results',
                        'opts': self.model._meta,
                        'has_view_permission': True,
                        'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
                        'manual_summary_format': MANUAL_SUMMARY_FORMAT,
                    }
                    return render(request, 'admin/academics/upload_results.html', context)
                uploaded_file = request.FILES['file']
                session = file_form.cleaned_data['session']
                semester = file_form.cleaned_data['semester']
                try:
                    # Read uploaded file into memory so we use ONLY the file the user sent (never disk path)
                    file_content = uploaded_file.read()
                    file_size_bytes = len(file_content)
                    if not file_content:
                        messages.warning(request, 'The uploaded file is empty.')
                        file_form = FileUploadForm()
                        context = {
                            **self.admin_site.each_context(request),
                            'file_form': file_form,
                            'students': students,
                            'title': 'Upload Results',
                            'opts': self.model._meta,
                            'has_view_permission': True,
                            'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
                            'manual_summary_format': MANUAL_SUMMARY_FORMAT,
                        }
                        return render(request, 'admin/academics/upload_results.html', context)
                    file_handle = io.BytesIO(file_content)
                    file_name = uploaded_file.name.lower()
                    file_ext = os.path.splitext(file_name)[1]
                    results_data = []
                    summaries_from_file = []
                    if file_ext in ['.xlsx', '.xls']:
                        from .parsers.ibbul_wide import (
                            parse_ibbul_university_excel,
                            parse_ibbul_wide_excel,
                        )
                        raw_rows = []
                        # CRITICAL: Write uploaded bytes to a UNIQUE temp file and read only from that.
                        # This guarantees no library can open any other file (e.g. Untitled.xls in project).
                        suffix = '.xls' if file_ext == '.xls' else '.xlsx'
                        fd, temp_path = None, None
                        try:
                            fd, temp_path = tempfile.mkstemp(suffix=suffix, prefix='upload_')
                            os.write(fd, file_content)
                            os.close(fd)
                            fd = None
                            try:
                                import pandas as pd
                            except ImportError:
                                messages.error(
                                    request,
                                    'Excel upload requires pandas. Install with: pip install pandas openpyxl xlrd'
                                )
                                return redirect('admin:academics_result_upload_results')
                            def get_sheet_raw_rows(path, sheet_index, is_xls):
                                if is_xls:
                                    try:
                                        import xlrd
                                        wb = xlrd.open_workbook(path)
                                        if sheet_index >= wb.nsheets:
                                            return None
                                        sh = wb.sheet_by_index(sheet_index)
                                        return [[xlrd.sheet.cell_displaytext(sh, r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
                                    except Exception:
                                        return None
                                try:
                                    df = pd.read_excel(path, header=None, sheet_name=sheet_index, engine='openpyxl')
                                    return df.fillna('').values.tolist()
                                except Exception:
                                    return None

                            num_sheets = 1
                            if file_ext == '.xls':
                                try:
                                    import xlrd
                                    wb = xlrd.open_workbook(temp_path)
                                    num_sheets = max(1, wb.nsheets)
                                except Exception:
                                    num_sheets = 1
                            else:
                                try:
                                    xl = pd.ExcelFile(temp_path, engine='openpyxl')
                                    num_sheets = max(1, len(xl.sheet_names))
                                except Exception:
                                    num_sheets = 1

                            # Prefer sheet whose first student ID is U22 > U25 > U10 (so your file's students win)
                            raw_rows = []
                            sheet_used = 0
                            best_priority = -1  # 2=U22, 1=U25, 0=U10, -1=other
                            for idx in range(num_sheets):
                                rows = get_sheet_raw_rows(temp_path, idx, file_ext == '.xls')
                                if not rows:
                                    continue
                                uni_result = parse_ibbul_university_excel(rows, session=session, semester=semester)
                                uni = uni_result[0] if isinstance(uni_result, tuple) else uni_result
                                wide = parse_ibbul_wide_excel(rows, session=session, semester=semester)
                                cand = uni if uni else wide
                                if not cand:
                                    if not raw_rows:
                                        raw_rows = rows
                                        sheet_used = idx
                                    continue
                                first_id = (cand[0].get('student_id') or cand[0].get('matric_number') or '').strip().upper()
                                if first_id.startswith('U22'):
                                    priority = 2
                                elif first_id.startswith('U25'):
                                    priority = 1
                                elif first_id.startswith('U10'):
                                    priority = 0
                                else:
                                    priority = -1
                                if priority > best_priority:
                                    best_priority = priority
                                    raw_rows = rows
                                    sheet_used = idx
                            if not raw_rows and num_sheets:
                                for idx in range(num_sheets):
                                    rows = get_sheet_raw_rows(temp_path, idx, file_ext == '.xls')
                                    if rows:
                                        uni_result = parse_ibbul_university_excel(rows, session=session, semester=semester)
                                        uni = uni_result[0] if isinstance(uni_result, tuple) else uni_result
                                        wide = parse_ibbul_wide_excel(rows, session=session, semester=semester)
                                        if uni or wide:
                                            raw_rows = rows
                                            sheet_used = idx
                                            break
                            if not raw_rows:
                                raw_rows = get_sheet_raw_rows(temp_path, 0, file_ext == '.xls') or []
                            if raw_rows and num_sheets > 1:
                                messages.info(request, f'Using sheet {sheet_used + 1} of {num_sheets} (contains U22/U10/U25 data).')
                        finally:
                            if fd is not None:
                                try:
                                    os.close(fd)
                                except Exception:
                                    pass
                            if temp_path and os.path.exists(temp_path):
                                try:
                                    os.unlink(temp_path)
                                except Exception:
                                    pass
                        university_results, summaries_from_file = parse_ibbul_university_excel(raw_rows, session=session, semester=semester)
                        if university_results:
                            results_data = university_results
                        else:
                            wide_results = parse_ibbul_wide_excel(raw_rows, session=session, semester=semester)
                            if wide_results:
                                results_data = wide_results
                                # Still extract summaries from sheet so summary can be added when deleted
                                _, summaries_from_file = parse_ibbul_university_excel(raw_rows, session=session, semester=semester)
                            else:
                                results_data = []
                                summaries_from_file = []
                                messages.warning(
                                    request,
                                    'No IBBUL format detected. Sheet must have MATRIC.NO and course columns '
                                    '(e.g. CSC 401, CSC 403) with cells like "63 B" or score/grade. '
                                    'Or use the wide layout: course codes every 2 columns with score in the first.'
                                )
                    elif file_ext == '.csv':
                        try:
                            decoded_file = file_content.decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                decoded_file = file_content.decode('utf-8-sig')
                            except Exception:
                                decoded_file = file_content.decode('latin-1')
                        io_string = io.StringIO(decoded_file)
                        reader = csv.DictReader(io_string)
                        from .ibbul_format import normalize_column_name
                        for row in reader:
                            row_normalized = {normalize_column_name(k): (v.strip() if v else '') for k, v in row.items()}
                            data = map_to_canonical_columns(row_normalized)
                            student_id = (data.get('student_id') or '').strip()
                            course_code = (data.get('course_code') or '').strip().replace(' ', '').upper()
                            csv_session = data.get('session') or session
                            csv_semester = data.get('semester') or semester
                            if csv_semester:
                                csv_semester_upper = str(csv_semester).upper()
                                if '1' in csv_semester_upper or 'FIRST' in csv_semester_upper:
                                    csv_semester = 'FIRST'
                                elif '2' in csv_semester_upper or 'SECOND' in csv_semester_upper:
                                    csv_semester = 'SECOND'
                            # Summary row: student_id present but no course_code, and has at least one summary field
                            summary_fields = ('gpa', 'cgpa', 'rcu', 'trcu', 'tecu', 'tcp', 'le', 'nss', 'pcgpa', 'ecu', 'cp')
                            has_summary = any(data.get(f) for f in summary_fields)
                            if student_id and not course_code and has_summary:
                                data['session'] = csv_session
                                data['semester'] = csv_semester
                                data['student_id'] = student_id
                                summaries_from_file.append(data)
                                continue
                            if not student_id or not course_code:
                                continue
                            data['session'] = csv_session
                            data['semester'] = csv_semester
                            score_str = (data.get('score') or '').strip()
                            if score_str == '':
                                continue
                            try:
                                score_val = float(score_str)
                            except (TypeError, ValueError):
                                continue
                            data['student_id'] = student_id
                            data['course_code'] = course_code
                            data['score'] = score_val
                            results_data.append(data)
                    else:
                        messages.error(request, f'Unsupported file format: {file_ext}. Please upload .xlsx, .xls, or .csv')
                        file_form = FileUploadForm()
                        context = {
                            **self.admin_site.each_context(request),
                            'file_form': file_form,
                            'students': students,
                            'title': 'Upload Results',
                            'opts': self.model._meta,
                            'has_view_permission': True,
                            'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
                            'manual_summary_format': MANUAL_SUMMARY_FORMAT,
                        }
                        return render(request, 'admin/academics/upload_results.html', context)
                    rows_for_batch = [
                        d for d in results_data
                        if (d.get('student_id') or d.get('matric_number')) and d.get('course_code') and d.get('score') is not None
                    ]
                    # Show what we actually parsed so user can verify it matches their file
                    first_ids_from_parsed = []
                    seen_ids = set()
                    for d in rows_for_batch:
                        sid = (d.get('student_id') or d.get('matric_number') or '').strip()
                        if sid and sid.upper() not in seen_ids:
                            seen_ids.add(sid.upper())
                            first_ids_from_parsed.append(sid)
                            if len(first_ids_from_parsed) >= 5:
                                break
                    if not rows_for_batch:
                        size_kb = file_size_bytes / 1024.0
                        if summaries_from_file:
                            # File has only summary rows (or summary rows only) — save summaries and show success
                            summaries_saved = 0
                            for s in (summaries_from_file or []):
                                try:
                                    student = User.objects.filter(student_id=s.get('student_id')).first()
                                    if student:
                                        sess = s.get('session', session)
                                        sem = s.get('semester', semester)
                                        if not SemesterSummary.objects.filter(student=student, session=sess, semester=sem).exists():
                                            SemesterSummary.objects.create(
                                                student=student,
                                                session=sess,
                                                semester=sem,
                                                le=str(s.get('le', '')),
                                                nss=str(s.get('nss', '')),
                                                rcu=str(s.get('rcu', '')),
                                                ecu=str(s.get('ecu', '')),
                                                cp=str(s.get('cp', '')),
                                                gpa=str(s.get('gpa', '')),
                                                trcu=str(s.get('trcu', '')),
                                                tecu=str(s.get('tecu', '')),
                                                tcp=str(s.get('tcp', '')),
                                                pcgpa=str(s.get('pcgpa', '')),
                                                cgpa=str(s.get('cgpa', '')),
                                                outstanding_courses=str(s.get('outstanding_courses', '')),
                                                remarks=str(s.get('remarks', '')),
                                                standing=str(s.get('standing', '')),
                                                raw_summary='',
                                            )
                                            summaries_saved += 1
                                except Exception:
                                    pass
                            if summaries_saved > 0:
                                messages.success(
                                    request,
                                    f'No result rows in file. {summaries_saved} summary/summaries saved successfully.',
                                )
                            else:
                                messages.warning(
                                    request,
                                    'No result rows in file. No summaries were added (each student already has a summary for this session/semester, or no matching students in the system).',
                                )
                            context = {
                                **self.admin_site.each_context(request),
                                'file_form': FileUploadForm(),
                                'students': students,
                                'title': 'Upload Results',
                                'opts': self.model._meta,
                                'has_view_permission': True,
                                'upload_batch': None,
                                'summaries_saved': summaries_saved,
                                'report_failed_count': 0,
                                'upload_error_summary': '',
                                'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
                                'manual_summary_format': MANUAL_SUMMARY_FORMAT,
                            }
                            return render(request, 'admin/academics/upload_results.html', context)
                        messages.warning(
                            request,
                            f'No rows with student_id, course_code, and score found. '
                            f'Uploaded file: "{uploaded_file.name}", size {size_kb:.1f} KB. '
                            'If this size does not match your file, the wrong file may have been sent (e.g. browser cache). '
                            'Try: rename your file (e.g. my_results.xlsx), then upload again.'
                        )
                    else:
                        log_audit(
                            AuditLog.Action.RESULT_UPLOAD_STARTED,
                            request=request,
                            user=request.user,
                            identifier=uploaded_file.name,
                            extra={'row_count': len(rows_for_batch), 'session': session, 'semester': semester},
                        )
                        try:
                            dept = _hod_department(request)
                            batch, report_failed = ResultUploadService.process_upload_batch(
                                filename=uploaded_file.name,
                                rows_data=rows_for_batch,
                                uploaded_by=request.user,
                                session=session,
                                semester=semester,
                                department_id=dept.pk if dept else None,
                                faculty_id=dept.faculty_id if dept else None,
                            )
                            # Save exact summary from file (no calculation) — TRCU, TECU, TCP, CGPA etc. as uploaded
                            # Only add summary when missing or deleted for that student/session/semester; never overwrite existing.
                            summaries_saved = 0
                            for s in (summaries_from_file or []):
                                try:
                                    student = User.objects.filter(student_id=s.get('student_id')).first()
                                    if student:
                                        sess = s.get('session', session)
                                        sem = s.get('semester', semester)
                                        if not SemesterSummary.objects.filter(student=student, session=sess, semester=sem).exists():
                                            SemesterSummary.objects.create(
                                                student=student,
                                                session=sess,
                                                semester=sem,
                                                le=str(s.get('le', '')),
                                                nss=str(s.get('nss', '')),
                                                rcu=str(s.get('rcu', '')),
                                                ecu=str(s.get('ecu', '')),
                                                cp=str(s.get('cp', '')),
                                                gpa=str(s.get('gpa', '')),
                                                trcu=str(s.get('trcu', '')),
                                                tecu=str(s.get('tecu', '')),
                                                tcp=str(s.get('tcp', '')),
                                                pcgpa=str(s.get('pcgpa', '')),
                                                cgpa=str(s.get('cgpa', '')),
                                                outstanding_courses=str(s.get('outstanding_courses', '')),
                                                remarks=str(s.get('remarks', '')),
                                                standing=str(s.get('standing', '')),
                                                raw_summary='',
                                                upload_batch=batch,
                                            )
                                            summaries_saved += 1
                                except Exception:
                                    pass
                            log_audit(
                                AuditLog.Action.RESULT_UPLOAD_COMPLETED,
                                request=request,
                                user=request.user,
                                identifier=uploaded_file.name,
                                extra={
                                    'batch_id': batch.id,
                                    'success_count': batch.success_count,
                                    'error_count': batch.error_count,
                                    'filename': batch.filename,
                                },
                            )
                            success_count = batch.success_count
                            error_count = batch.error_count
                            error_reason = _upload_error_summary(report_failed)
                            if error_count == 0:
                                messages.success(
                                    request,
                                    f'Upload complete. All {success_count} result(s) were saved successfully. You can approve this batch from Upload batches when ready.',
                                )
                            elif success_count > 0:
                                messages.success(
                                    request,
                                    f'{success_count} result(s) saved successfully. {error_count} row(s) could not be saved. {error_reason}',
                                )
                            else:
                                if summaries_saved > 0:
                                    messages.warning(
                                        request,
                                        f'No result rows were saved. {summaries_saved} summary/summaries were saved. {error_count} row(s) had errors. {error_reason}',
                                    )
                                else:
                                    messages.warning(
                                        request,
                                        f'No results were saved. {error_count} row(s) had errors. {error_reason}',
                                    )
                            context = {
                                **self.admin_site.each_context(request),
                                'file_form': FileUploadForm(),
                                'students': students,
                                'title': 'Upload Results',
                                'opts': self.model._meta,
                                'has_view_permission': True,
                                'upload_batch': batch,
                                'summaries_saved': summaries_saved,
                                'report_failed_count': len(report_failed),
                                'upload_error_summary': error_reason if report_failed else '',
                                'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
                                'manual_summary_format': MANUAL_SUMMARY_FORMAT,
                            }
                            return render(request, 'admin/academics/upload_results.html', context)
                        except Exception as e:
                            import traceback
                            traceback.print_exc()
                            messages.error(request, 'The file could not be processed. Please check that it matches the template (correct columns and format). If the problem continues, contact support.')
                    file_form = FileUploadForm()
                except Exception as e:
                    import traceback
                    error_details = traceback.format_exc()
                    messages.error(request, 'The file could not be read. Please use a valid Excel or CSV file that matches the template.')
                    print(f"Upload error: {error_details}")
            
            elif 'manual_entry' in request.POST:
                student_id = request.POST.get('student_id')
                level = request.POST.get('level')
                session = request.POST.get('session', '2023/2024')  # Get session from form
                semester = request.POST.get('semester')
                course_entries = request.POST.get('course_entry', '').strip()
                summary = request.POST.get('summary', '').strip()
                
                if not student_id:
                    messages.error(request, 'Please select a student')
                elif not session:
                    messages.error(request, 'Please enter academic session (e.g., 2023/2024)')
                elif not semester:
                    messages.error(request, 'Please select semester')
                elif not course_entries:
                    messages.error(request, 'Please enter course entries (one per line)')
                else:
                    try:
                        # Parse multiple course entries (one per line)
                        course_lines = [line.strip() for line in course_entries.split('\n') if line.strip()]
                        
                        if not course_lines:
                            messages.error(request, 'No course entries found. Enter at least one course per line.')
                        else:
                            created_count = 0
                            error_count = 0
                            manual_entry_failed = []

                            for line_num, course_line in enumerate(course_lines, 1):
                                course_code = ''
                                score_str = ''
                                try:
                                    # Parse course entry (format: course_code, credit_unit, grade, score, remark — no course_title; title comes from catalogue)
                                    parts = [p.strip() for p in course_line.split(',')]
                                    
                                    if len(parts) < 4:
                                        msg = 'Invalid format. Use: course code, credit unit, grade, score (e.g. CSC301, 3, A, 75).'
                                        messages.warning(request, f'Line {line_num}: {msg} Skipping.')
                                        manual_entry_failed.append({'line_no': line_num, 'course_code': '', 'score': '', 'error_message': msg})
                                        error_count += 1
                                        continue
                                    
                                    # course_code, credit_unit, grade, score, remark (optional)
                                    course_code = parts[0].strip().replace(' ', '').upper()
                                    credit_unit = parts[1].strip() if len(parts) > 1 else ''
                                    grade = parts[2].strip() if len(parts) > 2 else ''
                                    score_str = parts[3].strip() if len(parts) > 3 else ''
                                    remark = parts[4].strip() if len(parts) > 4 else ''
                                    
                                    # Try to parse score
                                    if not score_str or score_str.strip() == '':
                                        msg = 'Score is required.'
                                        messages.warning(request, f'Line {line_num}: {msg} Skipping: {course_code}')
                                        manual_entry_failed.append({'line_no': line_num, 'course_code': course_code, 'score': '', 'error_message': msg})
                                        error_count += 1
                                        continue
                                    
                                    # Remove any non-numeric characters except digits and decimal point
                                    cleaned_score = ''.join(c for c in score_str if c.isdigit() or c == '.')
                                    
                                    if not cleaned_score:
                                        msg = f'Invalid score value "{score_str}".'
                                        messages.warning(request, f'Line {line_num}: {msg} Skipping: {course_code}')
                                        manual_entry_failed.append({'line_no': line_num, 'course_code': course_code, 'score': score_str, 'error_message': msg})
                                        error_count += 1
                                        continue
                                    
                                    score = Decimal(cleaned_score)
                                    
                                    # Validate score range
                                    if score < 0 or score > 100:
                                        msg = f'Score must be between 0 and 100. Got: {score}.'
                                        messages.warning(request, f'Line {line_num}: {msg} Skipping: {course_code}')
                                        manual_entry_failed.append({'line_no': line_num, 'course_code': course_code, 'score': str(score), 'error_message': msg})
                                        error_count += 1
                                        continue
                                    
                                    # Create result (course title comes from catalogue; student and course must exist)
                                    result = ResultUploadService.create_result(
                                        student_id=student_id,
                                        course_code=course_code,
                                        score=score,
                                        session=session,
                                        semester=semester,
                                        uploaded_by=request.user,
                                        course_title=None,
                                        credit_unit=int(credit_unit) if credit_unit and credit_unit.isdigit() else None,
                                        remark=remark if remark else None,
                                        department=hod_dept if hod_dept else None,
                                    )
                                    created_count += 1
                                    
                                except ValueError as e:
                                    friendly = _manual_entry_friendly_message(e, line_num, course_code)
                                    messages.warning(request, friendly)
                                    manual_entry_failed.append({
                                        'line_no': line_num,
                                        'course_code': course_code,
                                        'score': score_str,
                                        'error_message': friendly.split(': ', 1)[-1] if ': ' in friendly else str(e),
                                    })
                                    error_count += 1
                                except Exception as e:
                                    friendly = _manual_entry_friendly_message(e, line_num, course_code)
                                    messages.warning(request, friendly)
                                    manual_entry_failed.append({
                                        'line_no': line_num,
                                        'course_code': course_code,
                                        'score': score_str,
                                        'error_message': friendly.split(': ', 1)[-1] if ': ' in friendly else str(e),
                                    })
                                    error_count += 1
                            
                            # Save summary if provided (format: LE, NSS, RCU, ECU, CP, GPA, TRCU, TECU, TCP, PCGPA, CGPA, Outstanding courses, Remarks — 13 values; no Standing)
                            if summary and summary.strip():
                                try:
                                    summary_parts = [p.strip() for p in summary.split(',')]
                                    if len(summary_parts) >= 13:
                                        student = User.objects.get(student_id=student_id)
                                        SemesterSummary.objects.update_or_create(
                                            student=student,
                                            session=session,
                                            semester=semester,
                                            defaults={
                                                'le': summary_parts[0] if len(summary_parts) > 0 else '',
                                                'nss': summary_parts[1] if len(summary_parts) > 1 else '',
                                                'rcu': summary_parts[2] if len(summary_parts) > 2 else '',
                                                'ecu': summary_parts[3] if len(summary_parts) > 3 else '',
                                                'cp': summary_parts[4] if len(summary_parts) > 4 else '',
                                                'gpa': summary_parts[5] if len(summary_parts) > 5 else '',
                                                'trcu': summary_parts[6] if len(summary_parts) > 6 else '',
                                                'tecu': summary_parts[7] if len(summary_parts) > 7 else '',
                                                'tcp': summary_parts[8] if len(summary_parts) > 8 else '',
                                                'pcgpa': summary_parts[9] if len(summary_parts) > 9 else '',
                                                'cgpa': summary_parts[10] if len(summary_parts) > 10 else '',
                                                'outstanding_courses': summary_parts[11] if len(summary_parts) > 11 else '',
                                                'remarks': summary_parts[12] if len(summary_parts) > 12 else '',
                                                'standing': '',
                                                'raw_summary': summary.strip()
                                            }
                                        )
                                except Exception as e:
                                    messages.warning(request, 'Summary could not be saved. Make sure the summary line has 13 values separated by commas (e.g. LE, NSS, RCU, ECU, CP, GPA, TRCU, TECU, TCP, PCGPA, CGPA, Outstanding courses, Remarks).')
                            
                            # Summary message and audit
                            if created_count > 0:
                                log_audit(
                                    AuditLog.Action.RESULT_MANUAL_ENTRY,
                                    request=request,
                                    user=request.user,
                                    identifier=student_id,
                                    extra={'created_count': created_count, 'error_count': error_count, 'session': session, 'semester': semester},
                                )
                                messages.success(request, f'✓ Successfully created {created_count} result(s) for {student_id}')
                            if error_count > 0:
                                messages.warning(request, f'{error_count} course(s) could not be saved. See the messages above for each line.')
                                request.session['manual_entry_last_report'] = manual_entry_failed[:1000]
                                request.session['manual_entry_report_student_id'] = student_id
                            else:
                                request.session.pop('manual_entry_last_report', None)
                                request.session.pop('manual_entry_report_student_id', None)
                            
                    except Exception as e:
                        messages.error(request, 'We couldn\'t process this entry. Please check the format (student ID, then one line per course). If the problem continues, contact support.')
        
        context = {
            **self.admin_site.each_context(request),
            'file_form': file_form,
            'students': students,
            'title': 'Upload Results',
            'opts': self.model._meta,
            'has_view_permission': True,
            'manual_course_line_format': MANUAL_COURSE_LINE_FORMAT,
            'manual_summary_format': MANUAL_SUMMARY_FORMAT,
            'manual_entry_failed_count': len(request.session.get('manual_entry_last_report', [])),
            'manual_entry_has_report': bool(request.session.get('manual_entry_last_report')),
        }
        return render(request, 'admin/academics/upload_results.html', context)
    
    def changelist_view(self, request, extra_context=None):
        """Add upload CSV button to changelist"""
        extra_context = extra_context or {}
        extra_context['show_upload_button'] = True
        return super().changelist_view(request, extra_context)


# GPA model is NOT registered in admin
# GPA records are managed through summary uploads only


@admin.register(ResultUploadBatch)
class ResultUploadBatchAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    """
    Enterprise: approve or reject an entire upload batch (not per course or per result).
    HOD sees batches in their scope; one action approves/rejects all results in the batch.
    """
    list_display = (
        'id', 'filename', 'approval_status_display', 'batch_progress_column', 'batch_actions_column',
        'session', 'semester', 'status', 'success_count', 'error_count', 'created_at',
    )
    list_filter = ('approval_status', 'status', 'session', 'semester', BatchDepartmentListFilter, BatchFacultyListFilter, 'created_at')
    search_fields = ('filename', 'session', 'semester')
    readonly_fields = (
        'filename', 'uploaded_by', 'department', 'faculty', 'status', 'approval_status',
        'approved_by', 'approved_at', 'rejection_reason', 'session', 'semester',
        'success_count', 'error_count', 'created_at', 'completed_at',
    )
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    list_per_page = 25
    change_list_template = 'admin/academics/resultuploadbatch/change_list.html'
    change_form_template = 'admin/academics/resultuploadbatch/change_form.html'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['scope_info'] = _academics_scope_info(request)
        qs = self.get_queryset(request)
        extra_context['batch_pending_count'] = qs.filter(
            status=ResultUploadBatch.Status.COMPLETED,
            approval_status=ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL,
        ).count()
        extra_context['batch_approved_count'] = qs.filter(approval_status=ResultUploadBatch.ApprovalStatus.APPROVED).count()
        extra_context['batch_rejected_count'] = qs.filter(approval_status=ResultUploadBatch.ApprovalStatus.REJECTED).count()
        return super().changelist_view(request, extra_context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        batch = ResultUploadBatch.objects.filter(pk=object_id).annotate(
            _approved_results_count=Count(
                'results',
                filter=Q(results__status__in=['APPROVED', 'LOCKED_PUBLISHED']),
                distinct=True,
            )
        ).first()
        if batch:
            extra_context['batch_approved_count'] = getattr(batch, '_approved_results_count', 0) or 0
            extra_context['batch_total_results'] = batch.success_count or 0
            extra_context['batch_hub_url'] = reverse('admin:academics_result_changelist') + '?upload_batch=' + str(batch.pk)
        return super().change_view(request, object_id, form_url, extra_context)

    class ResultBatchInline(admin.TabularInline):
        model = Result
        fk_name = 'upload_batch'
        extra = 0
        max_num = 0
        can_delete = False
        fields = ('student', 'course', 'score', 'grade', 'status', 'session', 'semester')
        readonly_fields = ('student', 'course', 'score', 'grade', 'status', 'session', 'semester')
        ordering = ['student__student_id', 'course__code']
        show_change_link = True

        def get_queryset(self, request):
            return super().get_queryset(request).select_related('student', 'course')

    inlines = [ResultBatchInline]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = filter_by_scope(qs, request.user, request).select_related('uploaded_by', 'department', 'approved_by')
        # Annotate how many results in this batch are already approved (individually or via batch)
        qs = qs.annotate(
            _approved_results_count=Count(
                'results',
                filter=Q(results__status__in=['APPROVED', 'LOCKED_PUBLISHED']),
                distinct=True,
            )
        )
        return qs

    def has_module_permission(self, request):
        """Allow any staff to see Upload Batches (same as Results); no separate permission required."""
        return getattr(request.user, 'is_staff', False)

    def has_view_permission(self, request, obj=None):
        """Allow any staff to view the batch list and detail."""
        return getattr(request.user, 'is_staff', False)

    def has_change_permission(self, request, obj=None):
        """Approve/Reject are custom views; allow staff so they can use batch actions."""
        return getattr(request.user, 'is_staff', False)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def approval_status_display(self, obj):
        total = obj.success_count or 0
        approved = getattr(obj, '_approved_results_count', 0) or 0
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL and obj.status == ResultUploadBatch.Status.COMPLETED:
            if total == 0:
                return format_html('<span class="rub-pill rub-pill-pending">Pending approval</span>')
            if approved == 0:
                return format_html('<span class="rub-pill rub-pill-pending">Pending approval</span>')
            if approved >= total:
                return format_html('<span class="rub-pill rub-pill-approved">✓ Approved</span>')
            return format_html('<span class="rub-pill rub-pill-partial">Partially Approved</span>')
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.APPROVED:
            return format_html('<span class="rub-pill rub-pill-approved">✓ Approved</span>')
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.REJECTED:
            return format_html('<span class="rub-pill rub-pill-rejected">✗ Rejected</span>')
        return obj.get_approval_status_display()
    approval_status_display.short_description = 'Approval'
    approval_status_display.admin_order_field = 'approval_status'

    def batch_progress_column(self, obj):
        """Show Results: X / Y approved; Status: Approved when done."""
        total = obj.success_count or 0
        approved = getattr(obj, '_approved_results_count', 0) or 0
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            if total == 0:
                return format_html('<span class="rub-meta-muted">—</span>')
            if approved >= total:
                return format_html(
                    '<span class="rub-progress-full">Results: {} / {} approved</span>',
                    approved, total,
                )
            if approved == 0:
                return format_html(
                    '<span class="rub-progress-pending">Results: 0 / {} approved</span>',
                    total,
                )
            return format_html(
                '<span class="rub-progress-partial">Results: {} / {} approved</span>',
                approved, total,
            )
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.APPROVED:
            return format_html(
                '<span class="rub-meta-approved">Results: {} / {} approved<br><small>Status: Approved</small></span>',
                total, total,
            )
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.REJECTED:
            return format_html('<span class="rub-meta-rejected">{} rejected</span>', total)
        return format_html('<span class="rub-meta-muted">—</span>')
    batch_progress_column.short_description = 'Results'

    def batch_actions_column(self, obj):
        report_url = reverse('admin:academics_result_import_report', args=[obj.pk])
        # Always show Report for completed batches (failed-rows CSV)
        report_link = format_html('<a href="{}" class="rub-btn rub-btn-report">↓ Report</a>', report_url)
        if obj.status != ResultUploadBatch.Status.COMPLETED:
            if obj.error_count and obj.error_count > 0:
                return format_html('<div class="rub-actions-cell">{}</div>', report_link)
            return format_html('<div class="rub-actions-cell"><span class="rub-meta-muted">—</span></div>')
        # COMPLETED: show actions by approval status
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.PENDING_APPROVAL:
            approve_url = reverse('admin:academics_resultuploadbatch_approve', args=[obj.pk])
            reject_url = reverse('admin:academics_resultuploadbatch_reject', args=[obj.pk])
            return format_html(
                '<div class="rub-actions-cell">'
                '<a href="{}" class="rub-btn rub-btn-approve">✓ Approve</a>'
                '<a href="{}" class="rub-btn rub-btn-reject">✗ Reject</a>'
                '{}'
                '</div>',
                approve_url, reject_url, report_link,
            )
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.APPROVED:
            unapprove_url = reverse('admin:academics_resultuploadbatch_unapprove', args=[obj.pk])
            approved_by = obj.approved_by.get_full_name() if obj.approved_by else '—'
            return format_html(
                '<div class="rub-actions-cell rub-actions-approved">'
                '<span class="rub-meta-approved" title="Approved by {}">✓ {}</span>'
                '{}'
                '<a href="{}" class="rub-btn rub-btn-unapprove" data-confirm="Revert this batch to draft? All {} result(s) will need approval again." title="Revert batch to draft">↩ Unapprove</a>'
                '</div>',
                approved_by, approved_by, report_link, unapprove_url, obj.success_count or 0,
            )
        if obj.approval_status == ResultUploadBatch.ApprovalStatus.REJECTED:
            reopen_url = reverse('admin:academics_resultuploadbatch_reopen', args=[obj.pk])
            return format_html(
                '<div class="rub-actions-cell">'
                '<span class="rub-meta-rejected">✗ Rejected</span>'
                '{}'
                '<a href="{}" class="rub-btn rub-btn-reopen" data-confirm="Re-open this batch for approval? All {} result(s) will be set to draft." title="Re-open for approval">↻ Re-open</a>'
                '</div>',
                report_link, reopen_url, obj.success_count or 0,
            )
        return format_html('<div class="rub-actions-cell">{}</div>', report_link)
    batch_actions_column.short_description = 'Actions'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/approve/', self.admin_site.admin_view(self.approve_batch_view), name='academics_resultuploadbatch_approve'),
            path('<int:pk>/reject/', self.admin_site.admin_view(self.reject_batch_view), name='academics_resultuploadbatch_reject'),
            path('<int:pk>/unapprove/', self.admin_site.admin_view(self.unapprove_batch_view), name='academics_resultuploadbatch_unapprove'),
            path('<int:pk>/reopen/', self.admin_site.admin_view(self.reopen_batch_view), name='academics_resultuploadbatch_reopen'),
        ]
        return custom + urls

    def approve_batch_view(self, request, pk):
        batch = get_object_or_404(self.get_queryset(request), pk=pk)
        try:
            updated = BatchApprovalService.approve_batch(batch, request.user)
            log_audit(
                AuditLog.Action.RESULT_BATCH_APPROVED,
                request=request,
                user=request.user,
                identifier=batch.filename,
                extra={'batch_id': batch.id, 'results_updated': updated},
            )
            messages.success(request, f'Batch approved. {updated} result(s) marked approved.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('admin:academics_resultuploadbatch_changelist')

    def reject_batch_view(self, request, pk):
        batch = get_object_or_404(self.get_queryset(request), pk=pk)
        reason = (request.GET.get('reason') or request.POST.get('reason') or '')[:2000]
        try:
            updated = BatchApprovalService.reject_batch(batch, request.user, reason=reason)
            log_audit(
                AuditLog.Action.RESULT_BATCH_REJECTED,
                request=request,
                user=request.user,
                identifier=batch.filename,
                extra={'batch_id': batch.id, 'results_updated': updated, 'reason': reason[:200]},
            )
            messages.success(request, f'Batch rejected. {updated} result(s) marked rejected.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('admin:academics_resultuploadbatch_changelist')

    def unapprove_batch_view(self, request, pk):
        batch = get_object_or_404(self.get_queryset(request), pk=pk)
        try:
            updated = BatchApprovalService.unapprove_batch(batch, request.user)
            log_audit(
                AuditLog.Action.RESULT_BATCH_UNAPPROVED,
                request=request,
                user=request.user,
                identifier=batch.filename,
                extra={'batch_id': batch.id, 'results_updated': updated},
            )
            messages.success(request, f'Batch reverted to draft. {updated} result(s) set to draft.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('admin:academics_resultuploadbatch_changelist')

    def reopen_batch_view(self, request, pk):
        batch = get_object_or_404(self.get_queryset(request), pk=pk)
        try:
            updated = BatchApprovalService.reopen_batch(batch, request.user)
            log_audit(
                AuditLog.Action.RESULT_BATCH_REOPENED,
                request=request,
                user=request.user,
                identifier=batch.filename,
                extra={'batch_id': batch.id, 'results_updated': updated},
            )
            messages.success(request, f'Batch re-opened for approval. {updated} result(s) set to draft.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('admin:academics_resultuploadbatch_changelist')


@admin.register(CourseAssignment)
class CourseAssignmentAdmin(ScopeFilteredAdminMixin, admin.ModelAdmin):
    """Course Responsibility Dashboard: assign lecturers to courses, workload, quick actions."""
    list_display = (
        'examiner_display', 'course', 'course_semester_display', 'students_display',
        'results_status_display', 'course_department', 'quick_actions_display',
    )
    list_filter = (ScopedDepartmentListFilter,)
    search_fields = ('examiner__email', 'examiner__first_name', 'examiner__last_name', 'course__code', 'course__title')
    autocomplete_fields = ('course',)
    ordering = ['course__code', 'examiner__email']
    change_list_template = 'admin/academics/courseassignment/change_list.html'
    actions_on_top = True
    actions_on_bottom = True
    list_per_page = 25

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('examiner', 'course', 'course__department', 'course__department__faculty').annotate(
            student_count=Count('course__results__student', distinct=True),
            results_count=Count('course__results'),
        )

    def has_module_permission(self, request):
        """Lecturer does not see Course assignments (view-only: Results + Courses only)."""
        if _is_examiner(request):
            return False
        return request.user.is_staff

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff and not _is_examiner(request)

    def has_add_permission(self, request):
        if _is_examiner(request):
            return False
        return request.user.is_staff

    def has_change_permission(self, request, obj=None):
        return request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        return request.user.is_staff

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Scope add/edit: HOD sees only their department; Faculty Admin sees their faculty's departments."""
        if _is_examiner(request):
            return super().formfield_for_foreignkey(db_field, request, **kwargs)
        hod_dept = _hod_department(request)
        role_str = str(getattr(request.user, 'role', '') or '').upper()
        if db_field.name == 'examiner':
            qs = User.objects.filter(role__in=(UserRole.EXAMINER, 'EXAMINER')).order_by('email')
            if hod_dept:
                qs = qs.filter(department_fk_id=hod_dept.pk)
            elif role_str == 'FACULTY_ADMIN' and getattr(request.user, 'faculty_id', None):
                qs = qs.filter(department_fk__faculty_id=request.user.faculty_id)
            kwargs['queryset'] = qs
        elif db_field.name == 'course':
            qs = kwargs.get('queryset') or Course.objects.all().order_by('code')
            if hod_dept:
                qs = qs.filter(department_id=hod_dept.pk)
            elif role_str == 'FACULTY_ADMIN' and getattr(request.user, 'faculty_id', None):
                qs = qs.filter(department__faculty_id=request.user.faculty_id)
            kwargs['queryset'] = qs
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = AuditLog.Action.COURSE_ASSIGNMENT_CREATED if not change else AuditLog.Action.COURSE_ASSIGNMENT_UPDATED
        identifier = f'{obj.examiner.email or obj.examiner.get_full_name()} ← {obj.course.code}'
        extra = {
            'assignment_id': obj.pk,
            'examiner_id': obj.examiner_id,
            'examiner_email': obj.examiner.email,
            'course_id': obj.course_id,
            'course_code': obj.course.code,
        }
        log_audit(action, request=request, user=request.user, identifier=identifier, extra=extra)

    def delete_model(self, request, obj):
        identifier = f'{obj.examiner.email or obj.examiner.get_full_name()} ← {obj.course.code}'
        extra = {
            'assignment_id': obj.pk,
            'examiner_id': obj.examiner_id,
            'course_id': obj.course_id,
            'course_code': obj.course.code,
        }
        log_audit(AuditLog.Action.COURSE_ASSIGNMENT_DELETED, request=request, user=request.user, identifier=identifier, extra=extra)
        super().delete_model(request, obj)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['scope_info'] = _academics_scope_info(request)
        return super().changelist_view(request, extra_context)

    def examiner_display(self, obj):
        if not obj.examiner:
            return '—'
        return obj.examiner.get_full_name() or obj.examiner.email or str(obj.examiner)
    examiner_display.short_description = _('Lecturer')
    examiner_display.admin_order_field = 'examiner__email'

    def course_semester_display(self, obj):
        if not obj.course:
            return '—'
        return obj.course.get_semester_display() if hasattr(obj.course, 'get_semester_display') else (obj.course.semester or '—')
    course_semester_display.short_description = _('Semester')
    course_semester_display.admin_order_field = 'course__semester'

    def students_display(self, obj):
        n = getattr(obj, 'student_count', 0) or 0
        return format_html('<span class="ca-students-badge">{}</span>', n)
    students_display.short_description = _('Students')
    students_display.admin_order_field = 'student_count'

    def results_status_display(self, obj):
        n = getattr(obj, 'results_count', 0) or 0
        if n > 0:
            return format_html('<span class="ca-results-yes">✓ {} entries</span>', n)
        return format_html('<span class="ca-results-no">No results</span>')
    results_status_display.short_description = _('Results')
    results_status_display.admin_order_field = 'results_count'

    def course_department(self, obj):
        return obj.course.department.name if obj.course and obj.course.department else '—'
    course_department.short_description = _('Department')

    def quick_actions_display(self, obj):
        if not obj.pk or not obj.course_id:
            return '—'
        change_url = reverse('admin:academics_courseassignment_change', args=[obj.pk])
        delete_url = reverse('admin:academics_courseassignment_delete', args=[obj.pk])
        result_changelist = reverse('admin:academics_result_changelist') + '?course__id__exact=' + str(obj.course_id)
        upload_url = reverse('admin:academics_result_upload_results')
        links = [
            format_html('<a href="{}" class="ca-action-link">View results</a>', result_changelist),
            format_html('<a href="{}" class="ca-action-link">Upload results</a>', upload_url),
            format_html('<a href="{}" class="ca-action-link">Change lecturer</a>', change_url),
            format_html('<a href="{}" class="ca-action-link ca-action-remove">Remove</a>', delete_url),
        ]
        return format_html('<span class="ca-quick-actions">{}</span>', format_html(' · '.join(links)))
    quick_actions_display.short_description = _('Actions')

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['scope_info'] = _academics_scope_info(request)
        base_courses = filter_by_scope(Course.objects.all(), request.user, request)
        total_assignments = filter_by_scope(CourseAssignment.objects.all(), request.user, request).count()
        assigned_course_ids = set(
            filter_by_scope(CourseAssignment.objects.all(), request.user, request)
            .values_list('course_id', flat=True).distinct()
        )
        courses_with_lecturer = len(assigned_course_ids)
        total_courses = base_courses.count()
        unassigned_count = total_courses - courses_with_lecturer
        top = (
            filter_by_scope(CourseAssignment.objects.all(), request.user, request)
            .values('examiner__first_name', 'examiner__last_name', 'examiner__email')
            .annotate(c=Count('id'))
            .order_by('-c')
            .first()
        )
        top_lecturer_name = '—'
        top_lecturer_count = 0
        if top:
            fn = top.get('examiner__first_name') or ''
            ln = top.get('examiner__last_name') or ''
            top_lecturer_name = (fn + ' ' + ln).strip() or top.get('examiner__email') or '—'
            top_lecturer_count = top['c']
        extra_context['assignment_stats'] = {
            'total_assignments': total_assignments,
            'unassigned_courses': unassigned_count,
            'courses_with_lecturer': courses_with_lecturer,
            'total_courses': total_courses,
            'top_lecturer_name': top_lecturer_name,
            'top_lecturer_count': top_lecturer_count,
        }
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('bulk-assign/', self.admin_site.admin_view(self.bulk_assign_view), name='academics_courseassignment_bulk_assign'),
        ]
        return custom + urls

    def bulk_assign_view(self, request):
        """Bulk assign: select lecturer + courses, create assignments in one go."""
        from django import forms as django_forms
        from django.utils.translation import gettext_lazy as __
        hod_dept = _hod_department(request)
        role_str = str(getattr(request.user, 'role', '') or '').upper()
        examiner_qs = User.objects.filter(role__in=(UserRole.EXAMINER, 'EXAMINER')).order_by('first_name', 'last_name', 'email')
        if hod_dept:
            examiner_qs = examiner_qs.filter(department_fk_id=hod_dept.pk)
        elif role_str == 'FACULTY_ADMIN' and getattr(request.user, 'faculty_id', None):
            examiner_qs = examiner_qs.filter(department_fk__faculty_id=request.user.faculty_id)
        course_qs = filter_by_scope(Course.objects.all(), request.user, request).filter(is_active=True).order_by('level', 'code')

        class BulkAssignForm(django_forms.Form):
            examiner = django_forms.ModelChoiceField(queryset=examiner_qs, required=True, label=__('Lecturer'), empty_label=__('Select lecturer…'))
            courses = django_forms.ModelMultipleChoiceField(queryset=course_qs, widget=django_forms.CheckboxSelectMultiple, required=False, label=__('Select courses'))

            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                eqs = list(self.fields['examiner'].queryset)
                empty = self.fields['examiner'].empty_label or '---------'
                self.fields['examiner'].choices = [('', empty)] + [
                    (u.pk, f"{u.get_full_name()} ({u.email})" if getattr(u, 'email', None) else u.get_full_name())
                    for u in eqs
                ]

        if request.method == 'POST':
            form = BulkAssignForm(request.POST)
            if form.is_valid():
                examiner = form.cleaned_data['examiner']
                courses = form.cleaned_data['courses'] or []
                created = 0
                for course in courses:
                    __unused, created_ = CourseAssignment.objects.get_or_create(examiner=examiner, course=course)
                    if created_:
                        created += 1
                        log_audit(
                            AuditLog.Action.COURSE_ASSIGNMENT_CREATED,
                            request=request, user=request.user,
                            identifier=f'{examiner.email or examiner.get_full_name()} ← {course.code}',
                            extra={'examiner_id': examiner.pk, 'course_id': course.pk, 'bulk': True},
                        )
                if created:
                    self.message_user(request, _('%s assignment(s) created.') % created, messages.SUCCESS)
                else:
                    self.message_user(request, _('No new assignments (all selected were already assigned).'), messages.INFO)
                return redirect(reverse('admin:academics_courseassignment_changelist'))
        else:
            form = BulkAssignForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'title': __('Assign courses to lecturer'),
            'opts': self.model._meta,
        }
        return render(request, 'admin/academics/courseassignment/bulk_assign.html', context)

