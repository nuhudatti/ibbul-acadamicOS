"""
Learning Engine API — live sync, gradebook, grade sheet export (Learning app only).
"""
import io
from django.core.cache import cache
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from apps.accounts.models import UserRole
from .models import (
    LMSOffering, Lesson, Quiz, QuizAttempt, Submission, Enrollment, LessonProgress,
)
from .grade_data import OfferingGradeData, _lesson_assignment_id
from .serializers import SubmissionSerializer

QUIZ_WEIGHT = 40
ASSIGNMENT_WEIGHT = 60

GRADE_BANDS = [
    (70, 'A'),
    (60, 'B'),
    (50, 'C'),
    (45, 'D'),
    (0, 'F'),
]


def letter_grade(score: float) -> str:
    for threshold, letter in GRADE_BANDS:
        if score >= threshold:
            return letter
    return 'F'


def _best_quiz_attempt(quiz, student):
    return (
        QuizAttempt.objects.filter(
            quiz=quiz,
            student=student,
            status__in=('submitted', 'timed_out'),
            score__isnull=False,
        )
        .order_by('-score')
        .first()
    )


def _student_quiz_average(student, offering, grade_data: OfferingGradeData | None = None):
    if grade_data:
        return grade_data.quiz_average(student.id)
    scores = []
    for lesson in Lesson.objects.filter(
        module__offering=offering, content_type='quiz', is_published=True
    ).select_related('quiz'):
        if not hasattr(lesson, 'quiz'):
            continue
        attempt = _best_quiz_attempt(lesson.quiz, student)
        if attempt:
            scores.append(float(attempt.score))
    if not scores:
        return None
    return round(sum(scores) / len(scores), 2)


def _student_assignment_average(student, offering, grade_data: OfferingGradeData | None = None):
    if grade_data:
        return grade_data.assignment_average(student.id)
    scores = []
    for lesson in Lesson.objects.filter(
        module__offering=offering, content_type='assignment', is_published=True
    ).select_related('assignment'):
        if not hasattr(lesson, 'assignment'):
            continue
        sub = Submission.objects.filter(
            assignment=lesson.assignment, student=student, score__isnull=False
        ).first()
        if sub:
            max_s = lesson.assignment.max_score or 100
            scores.append(round(float(sub.score) / max_s * 100, 2))
    if not scores:
        return None
    return round(sum(scores) / len(scores), 2)


def compute_final_grade(quiz_avg, assignment_avg):
    """Weighted final with dynamic re-normalization when one component is missing."""
    components = []
    if quiz_avg is not None:
        components.append((quiz_avg, QUIZ_WEIGHT))
    if assignment_avg is not None:
        components.append((assignment_avg, ASSIGNMENT_WEIGHT))
    if not components:
        return None, None
    total_weight = sum(w for _, w in components)
    final = sum(s * w for s, w in components) / total_weight
    final = round(final, 2)
    return final, letter_grade(final)


def _module_breakdown(student, offering, grade_data: OfferingGradeData | None = None):
    if grade_data:
        return grade_data.module_breakdown(student.id, compute_final_grade)
    modules = []
    for mod in offering.modules.filter(is_published=True).order_by('order'):
        quiz_scores = []
        asg_scores = []
        for lesson in mod.lessons.filter(is_published=True):
            if lesson.content_type == 'quiz' and hasattr(lesson, 'quiz'):
                att = _best_quiz_attempt(lesson.quiz, student)
                if att:
                    quiz_scores.append(float(att.score))
            elif lesson.content_type == 'assignment' and hasattr(lesson, 'assignment'):
                sub = Submission.objects.filter(
                    assignment=lesson.assignment, student=student, score__isnull=False
                ).first()
                if sub:
                    max_s = lesson.assignment.max_score or 100
                    asg_scores.append(float(sub.score) / max_s * 100)
        q_avg = round(sum(quiz_scores) / len(quiz_scores), 2) if quiz_scores else None
        a_avg = round(sum(asg_scores) / len(asg_scores), 2) if asg_scores else None
        final, letter = compute_final_grade(q_avg, a_avg)
        completed = LessonProgress.objects.filter(
            lesson__module=mod, student=student, completed=True
        ).count()
        total = mod.lessons.filter(is_published=True).count()
        modules.append({
            'module_id': mod.id,
            'module_title': mod.title,
            'quiz_average': q_avg,
            'assignment_average': a_avg,
            'final_score': final,
            'letter_grade': letter,
            'steps_completed': completed,
            'steps_total': total,
        })
    return modules


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def lesson_live_position(request, lesson_id):
    """
    Live lecturer position for reading/PDF sync.
    POST: instructor sets scroll_percent (0-100), page (optional)
    GET: students poll current position
    """
    cache_key = f'lms_live_lesson_{lesson_id}'
    try:
        lesson = Lesson.objects.select_related('module__offering').get(pk=lesson_id)
    except Lesson.DoesNotExist:
        return Response({'detail': 'Lesson not found.'}, status=status.HTTP_404_NOT_FOUND)

    user = request.user

    if request.method == 'POST':
        if user.role != UserRole.EXAMINER:
            return Response(status=status.HTTP_403_FORBIDDEN)
        if lesson.module.offering.instructor_id != user.id:
            return Response({'detail': 'Not your offering.'}, status=status.HTTP_403_FORBIDDEN)
        scroll = float(request.data.get('scroll_percent', 0))
        scroll = max(0, min(100, scroll))
        page = int(request.data.get('page', 1))
        payload = {
            'scroll_percent': scroll,
            'page': page,
            'instructor_name': user.get_full_name() or user.email,
            'active': request.data.get('active', True),
        }
        cache.set(cache_key, payload, timeout=7200)
        return Response(payload)

    data = cache.get(cache_key) or {'scroll_percent': 0, 'page': 1, 'active': False}
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def offering_gradebook(request, offering_id):
    """
    Grade breakdown for an offering.
    Students: own grades only. Instructors: all enrolled students.
    """
    try:
        offering = LMSOffering.objects.select_related('instructor').get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    user = request.user
    is_instructor = user.role == UserRole.EXAMINER and offering.instructor_id == user.id
    is_staff = user.role in (
        UserRole.DEPARTMENT_ADMIN, UserRole.HOD, UserRole.FACULTY_ADMIN, UserRole.SUPER_ADMIN
    )

    if user.role == UserRole.STUDENT:
        students = [user]
        student_ids = [user.id]
    elif is_instructor or is_staff:
        enrollments = list(
            Enrollment.objects.filter(offering=offering, is_active=True).select_related('student')
        )
        students = [e.student for e in enrollments]
        student_ids = [e.student_id for e in enrollments]
    else:
        return Response(status=status.HTTP_403_FORBIDDEN)

    grade_data = OfferingGradeData.load(offering, student_ids)

    rows = []
    for st in students:
        q_avg = grade_data.quiz_average(st.id)
        a_avg = grade_data.assignment_average(st.id)
        final, letter = compute_final_grade(q_avg, a_avg)
        rows.append({
            'student_id': st.student_id or str(st.id),
            'full_name': st.get_full_name(),
            'quiz_average': q_avg,
            'assignment_average': a_avg,
            'quiz_weight': QUIZ_WEIGHT,
            'assignment_weight': ASSIGNMENT_WEIGHT,
            'final_score': final,
            'letter_grade': letter,
            'modules': grade_data.module_breakdown(st.id, compute_final_grade),
        })

    return Response({
        'offering_id': offering.id,
        'course_code': offering.course.code,
        'weights': {'quiz': QUIZ_WEIGHT, 'assignment': ASSIGNMENT_WEIGHT},
        'grade_bands': [{'min': t, 'grade': g} for t, g in GRADE_BANDS],
        'students': rows if (is_instructor or is_staff) else rows[:1],
    })


def _quiz_security_stats(student, offering, grade_data: OfferingGradeData | None = None):
    if grade_data:
        return grade_data.security_stats.get(student.id, (0, 0, 0))
    total_violations = 0
    fullscreen_exits = 0
    tab_switches = 0
    for lesson in Lesson.objects.filter(
        module__offering=offering, content_type='quiz', is_published=True
    ).select_related('quiz'):
        if not hasattr(lesson, 'quiz'):
            continue
        attempts = QuizAttempt.objects.filter(quiz=lesson.quiz, student=student).exclude(
            status='in_progress'
        )
        for att in attempts:
            total_violations += att.focus_loss_count or 0
            for ev in att.violation_log or []:
                et = str(ev.get('type', '')).lower()
                if 'fullscreen' in et:
                    fullscreen_exits += 1
                if 'tab' in et or 'blur' in et or 'hidden' in et:
                    tab_switches += 1
    return total_violations, fullscreen_exits, tab_switches


def _assignment_status(sub, assignment):
    if not sub:
        return 'Missing', ''
    if sub.score is not None:
        return 'OK', ''
    report = sub.similarity_report or {}
    if report.get('flagged') or (sub.similarity_score and float(sub.similarity_score) >= 0.85):
        return 'Review', sub.similarity_score
    if sub.ai_graded and sub.score is None:
        return 'AI Pending', sub.similarity_score
    return 'Pending', sub.similarity_score


def _can_manage_offering(user, offering):
    is_instructor = user.role == UserRole.EXAMINER and offering.instructor_id == user.id
    is_staff = user.role in (
        UserRole.DEPARTMENT_ADMIN, UserRole.HOD, UserRole.FACULTY_ADMIN, UserRole.SUPER_ADMIN
    )
    return is_instructor or is_staff


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def offering_grading_summary(request, offering_id):
    """Dashboard stats for lecturer grading workspace."""
    try:
        offering = LMSOffering.objects.select_related('course__department__faculty', 'instructor').get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    if not _can_manage_offering(request.user, offering):
        return Response(status=status.HTTP_403_FORBIDDEN)

    cache_key = f'lms_grading_summary_{offering_id}'
    cached = cache.get(cache_key)
    if cached is not None:
        return Response(cached)

    enrollments = Enrollment.objects.filter(offering=offering, is_active=True).select_related('student')
    student_ids = list(enrollments.values_list('student_id', flat=True))
    grade_data = OfferingGradeData.load(offering, student_ids)
    payload = grade_data.grading_summary_stats()
    cache.set(cache_key, payload, timeout=60)
    return Response(payload)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_grade_sheet(request, offering_id):
    """
    Export grade sheet as Excel for an offering.
    GET /api/learning/offerings/{id}/grade-sheet/
    """
    try:
        offering = LMSOffering.objects.select_related('course__department__faculty', 'instructor').get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    user = request.user
    is_instructor = user.role == UserRole.EXAMINER and offering.instructor_id == user.id
    is_staff = user.role in (
        UserRole.DEPARTMENT_ADMIN, UserRole.HOD, UserRole.FACULTY_ADMIN, UserRole.SUPER_ADMIN
    )
    if not _can_manage_offering(user, offering):
        return Response(status=status.HTTP_403_FORBIDDEN)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.utils import timezone as dj_tz

    quiz_lessons = list(
        Lesson.objects.filter(
            module__offering=offering, content_type='quiz', is_published=True
        ).select_related('quiz', 'module').order_by('module__order', 'order')
    )
    assignment_lessons = list(
        Lesson.objects.filter(
            module__offering=offering, content_type='assignment', is_published=True
        ).select_related('assignment', 'module').order_by('module__order', 'order')
    )

    enrollments = list(
        Enrollment.objects.filter(offering=offering, is_active=True)
        .select_related('student')
        .order_by('student__student_id')
    )
    student_ids = [e.student_id for e in enrollments]
    grade_data = OfferingGradeData.load(offering, student_ids)

    course = offering.course
    dept = course.department
    faculty_name = dept.faculty.name if dept and dept.faculty else 'IBBUL'
    dept_name = dept.name if dept else ''
    instructor_name = offering.instructor.get_full_name() if offering.instructor else ''
    semester_label = 'First Semester' if offering.semester == 'FIRST' else 'Second Semester'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Grade Sheet'

    meta_font = Font(bold=True, size=11, color='0F6B3E')
    ws['A1'] = 'IBBUL Academic OS — Official Grade Sheet'
    ws['A1'].font = Font(bold=True, size=14, color='0F6B3E')
    ws['A2'] = f'Faculty: {faculty_name}'
    ws['A3'] = f'Department: {dept_name}'
    ws['A4'] = f'Course: {course.code} — {course.title}'
    ws['A5'] = f'Session: {offering.session} · {semester_label} · Lecturer: {instructor_name}'
    ws['A6'] = f'Exported: {dj_tz.now().strftime("%d %B %Y %H:%M")}'
    for r in range(2, 7):
        ws[f'A{r}'].font = meta_font

    header_row = 8
    headers = [
        'Matric No', 'Student Name', 'Program', 'Level',
        'Quiz Score (%)', 'Violations', 'Fullscreen Exits', 'Tab Switches',
    ]
    for i, les in enumerate(assignment_lessons, 1):
        short = les.assignment.title[:28] if hasattr(les, 'assignment') else les.title[:28]
        headers.extend([f'{short} Score', f'{short} Similarity', f'{short} Status'])

    header_fill = PatternFill(start_color='0F6B3E', end_color='0F6B3E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row_idx = header_row + 1
    for enr in enrollments:
        st = enr.student
        q_avg = grade_data.quiz_average(st.id)
        violations, fs_exits, tab_sw = grade_data.security_stats.get(st.id, (0, 0, 0))
        program = getattr(st, 'department_name', None) or st.department or dept_name

        row = [
            st.student_id or str(st.id),
            st.get_full_name(),
            program,
            st.level or '',
            round(q_avg, 1) if q_avg is not None else '',
            violations,
            fs_exits,
            tab_sw,
        ]

        for les in assignment_lessons:
            sub = None
            aid = _lesson_assignment_id(les)
            if aid is not None:
                sub = grade_data.submissions.get((st.id, aid))
            if sub and sub.score is not None:
                row.append(float(sub.score))
            elif sub:
                row.append('')
            else:
                row.append('')

            if sub and sub.similarity_score is not None:
                row.append(f'{float(sub.similarity_score) * 100:.0f}%')
            else:
                row.append('')

            status_label, _ = _assignment_status(sub, les.assignment if hasattr(les, 'assignment') else None)
            row.append(status_label)

        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for r in range(header_row + 1, row_idx):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 3, 42)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{course.code}_{offering.session}_grade_sheet.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def offering_grading_workspace(request, offering_id):
    """
    Combined payload for lecturer grading UI — one round trip instead of 3+N.
    GET /api/learning/offerings/{id}/grading-workspace/
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        offering = LMSOffering.objects.select_related(
            'course', 'course__department', 'instructor'
        ).prefetch_related(
            'modules__lessons__assignment',
        ).get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    if not _can_manage_offering(request.user, offering):
        return Response(status=status.HTTP_403_FORBIDDEN)

    try:
        enrollments = list(
            Enrollment.objects.filter(offering=offering, is_active=True).select_related('student')
        )
        student_ids = [e.student_id for e in enrollments]
        grade_data = OfferingGradeData.load(offering, student_ids)

        rows = []
        for enr in enrollments:
            st = enr.student
            q_avg = grade_data.quiz_average(st.id)
            a_avg = grade_data.assignment_average(st.id)
            final, letter = compute_final_grade(q_avg, a_avg)
            rows.append({
                'student_id': st.student_id or str(st.id),
                'full_name': st.get_full_name(),
                'quiz_average': q_avg,
                'assignment_average': a_avg,
                'quiz_weight': QUIZ_WEIGHT,
                'assignment_weight': ASSIGNMENT_WEIGHT,
                'final_score': final,
                'letter_grade': letter,
            })

        subs_by_assignment = {}
        for assignment_id, subs in grade_data.submissions_by_assignment().items():
            subs_by_assignment[str(assignment_id)] = [
                _submission_grading_dict(sub) for sub in subs
            ]

        assignments_meta = []
        for les in grade_data.assignment_lessons:
            if hasattr(les, 'assignment'):
                a = les.assignment
                assignments_meta.append({
                    'id': a.id,
                    'title': a.title,
                    'max_score': a.max_score or 100,
                    'module_title': les.module.title if les.module_id else '',
                    'enable_ai_grading': bool(getattr(a, 'enable_ai_grading', False)),
                })

        return Response({
            'summary': grade_data.grading_summary_stats(),
            'gradebook': {
                'offering_id': offering.id,
                'course_code': offering.course.code,
                'weights': {'quiz': QUIZ_WEIGHT, 'assignment': ASSIGNMENT_WEIGHT},
                'grade_bands': [{'min': t, 'grade': g} for t, g in GRADE_BANDS],
                'students': rows,
            },
            'assignments': assignments_meta,
            'submissions_by_assignment': subs_by_assignment,
        })
    except Exception as exc:
        logger.exception('grading-workspace failed for offering %s', offering_id)
        return Response(
            {'detail': f'Grading workspace error: {str(exc)[:200]}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def _submission_grading_dict(sub):
    """Lightweight submission payload for grading workspace (no heavy serializer)."""
    return {
        'id': sub.id,
        'assignment': sub.assignment_id,
        'assignment_title': sub.assignment.title if sub.assignment_id else '',
        'student_user_id': sub.student_id,
        'student_matric': sub.student.student_id if sub.student_id else '',
        'student_name': sub.student.get_full_name() if sub.student_id else '',
        'content': sub.content or '',
        'file_key': sub.file_key or '',
        'submitted_at': sub.submitted_at.isoformat() if sub.submitted_at else None,
        'is_late': sub.is_late,
        'score': str(sub.score) if sub.score is not None else None,
        'graded_at': sub.graded_at.isoformat() if sub.graded_at else None,
        'feedback': sub.feedback or '',
        'similarity_score': float(sub.similarity_score) if sub.similarity_score is not None else None,
        'similarity_report': sub.similarity_report or {},
        'ai_suggested_score': float(sub.ai_suggested_score) if sub.ai_suggested_score is not None else None,
        'ai_feedback': getattr(sub, 'ai_feedback', '') or '',
        'ai_graded': bool(getattr(sub, 'ai_graded', False)),
        'ai_confidence_score': float(sub.ai_confidence_score) if getattr(sub, 'ai_confidence_score', None) is not None else None,
        'ai_strengths': getattr(sub, 'ai_strengths', None) or [],
        'ai_weaknesses': getattr(sub, 'ai_weaknesses', None) or [],
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_export_grade_sheet(request, offering_id):
    """Queue Excel grade sheet generation — returns job_id for polling."""
    try:
        offering = LMSOffering.objects.select_related('course').get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    if not _can_manage_offering(request.user, offering):
        return Response(status=status.HTTP_403_FORBIDDEN)

    from .tasks import create_job, enqueue_export
    job_id = create_job('export', total=1)
    enqueue_export(job_id, offering_id, request.user.id)
    return Response({'job_id': job_id, 'status': 'queued'}, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_grade_sheet_job(request, offering_id, job_id):
    """Poll export job — when complete, returns base64 workbook."""
    try:
        offering = LMSOffering.objects.get(pk=offering_id)
    except LMSOffering.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    if not _can_manage_offering(request.user, offering):
        return Response(status=status.HTTP_403_FORBIDDEN)

    from .tasks import get_job
    job = get_job(job_id)
    if not job:
        return Response({'detail': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)

    payload = {
        'job_id': job_id,
        'status': job.get('status'),
        'processed': job.get('processed', 0),
        'total': job.get('total', 1),
        'error': job.get('error'),
    }
    if job.get('status') == 'complete' and job.get('result'):
        payload['download'] = job['result']
    return Response(payload)
