"""
Admin-only views: CSV/Excel user import with auto-generated temporary passwords.
One-time secure export (TTL); all actions audited.
"""
import csv
import io
import logging
import secrets
import uuid
from typing import List, Dict, Any, Tuple

from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Q

from .models import User, UserRole, AuditLog
from .audit import log_audit
from .scope import filter_by_scope, is_hod, get_hod_department_id, is_super_admin
from common.permissions.enterprise_permissions import user_can_bulk_import_users
from common.validators.student_id_validator import (
    validate_student_id_format,
    department_code_from_student_id,
)

TEMP_PASSWORD_EXPORT_TTL = 600  # 10 minutes
USER_MANAGEMENT_PAGE_SIZE = 25
USER_MANAGEMENT_PAGE_SIZES = [10, 25, 50, 100]
AUDIT_LOG_PAGE_SIZE = 25
AUDIT_LOG_PAGE_SIZES = [25, 50, 100, 200]

logger = logging.getLogger(__name__)


def _can_deactivate(actor: User, target: User) -> bool:
    """Only superuser can deactivate staff; no one can deactivate themselves."""
    if actor.pk == target.pk:
        return False
    if target.is_staff and not actor.is_superuser:
        return False
    return True


def _build_user_queryset(request: HttpRequest):
    """
    Build filtered, searchable, sorted User queryset from GET params.
    HOD sees only users in their department (scope applied). Super Admin sees all.
    Returns (queryset, filters_dict for template).
    """
    qs = User.objects.all()
    qs = filter_by_scope(qs, request.user, request)
    filters_used = {}

    # Search: reg_number (case-insensitive), full name (partial), email (partial)
    search = (request.GET.get('search') or '').strip()
    if search:
        term = search
        qs = qs.filter(
            Q(student_id__icontains=term)
            | Q(email__icontains=term)
            | Q(first_name__icontains=term)
            | Q(last_name__icontains=term)
        )
        filters_used['search'] = search

    # Role
    role = request.GET.get('role', '').strip()
    if role and role in dict(UserRole.choices):
        qs = qs.filter(role=role)
        filters_used['role'] = role

    # Account status: active, inactive
    status = request.GET.get('status', '').strip().lower()
    if status == 'active':
        qs = qs.filter(is_active=True)
        filters_used['status'] = 'active'
    elif status == 'inactive':
        qs = qs.filter(is_active=False)
        filters_used['status'] = 'inactive'

    # First login: must_change, set
    first_login = request.GET.get('first_login', '').strip().lower()
    if first_login == 'must_change':
        qs = qs.filter(is_first_login=True)
        filters_used['first_login'] = 'must_change'
    elif first_login == 'set':
        qs = qs.filter(is_first_login=False)
        filters_used['first_login'] = 'set'

    # Department (partial)
    department = request.GET.get('department', '').strip()
    if department:
        qs = qs.filter(department__icontains=department)
        filters_used['department'] = department

    # Last login: never, 7d, 30d
    last_login_filter = request.GET.get('last_login', '').strip().lower()
    if last_login_filter == 'never':
        qs = qs.filter(last_login__isnull=True)
        filters_used['last_login'] = 'never'
    elif last_login_filter == '7d':
        since = timezone.now() - timezone.timedelta(days=7)
        qs = qs.filter(last_login__gte=since)
        filters_used['last_login'] = '7d'
    elif last_login_filter == '30d':
        since = timezone.now() - timezone.timedelta(days=30)
        qs = qs.filter(last_login__gte=since)
        filters_used['last_login'] = '30d'

    # Sort
    sort = request.GET.get('sort', 'identifier').strip()
    if sort == 'name':
        qs = qs.order_by('first_name', 'last_name')
        filters_used['sort'] = 'name'
    elif sort == 'last_login':
        qs = qs.order_by('-last_login')
        filters_used['sort'] = 'last_login'
    else:
        # identifier (student_id / email)
        qs = qs.order_by('student_id', 'email')
        filters_used['sort'] = 'identifier'

    return qs, filters_used

ROLE_MAP = {
    'student': UserRole.STUDENT,
    'students': UserRole.STUDENT,
    'STUDENT': UserRole.STUDENT,
    'lecturer': UserRole.EXAMINER,
    'lecturers': UserRole.EXAMINER,
    'EXAMINER': UserRole.EXAMINER,
    'admin': UserRole.DEPARTMENT_ADMIN,
    'admins': UserRole.DEPARTMENT_ADMIN,
    'HOD': UserRole.DEPARTMENT_ADMIN,
    'hod': UserRole.DEPARTMENT_ADMIN,
    'DEPARTMENT_ADMIN': UserRole.DEPARTMENT_ADMIN,
    'department_admin': UserRole.DEPARTMENT_ADMIN,
    'FACULTY_ADMIN': UserRole.FACULTY_ADMIN,
    'faculty_admin': UserRole.FACULTY_ADMIN,
    'dean': UserRole.FACULTY_ADMIN,
    'SUPER_ADMIN': UserRole.SUPER_ADMIN,
    'super_admin': UserRole.SUPER_ADMIN,
}


def _normalize_role(value: str) -> str:
    v = (value or '').strip()
    return ROLE_MAP.get(v, UserRole.STUDENT)


def _generate_temp_password() -> str:
    """Secure random password: 12 chars, URL-safe, meets basic strength."""
    return secrets.token_urlsafe(12)


def _parse_csv(content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV; expected columns: reg_number/student_id, first_name, last_name, role, email, faculty, department."""
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        row = {k.strip().lower().replace(' ', '_'): v.strip() if isinstance(v, str) else str(v or '').strip() for k, v in row.items()}
        reg = row.get('reg_number') or row.get('student_id') or ''
        email = (row.get('email') or '').strip()
        role = _normalize_role(row.get('role') or 'student')
        if reg or (email and role != UserRole.STUDENT):
            if 'student_id' not in row and row.get('reg_number'):
                row['student_id'] = row['reg_number']
            rows.append(row)
    return rows


def _parse_excel(content: bytes) -> List[Dict[str, Any]]:
    """Parse first sheet of Excel; same column semantics as CSV (incl. faculty, department for staff)."""
    import pandas as pd
    df = pd.read_excel(io.BytesIO(content), engine='openpyxl', sheet_name=0)
    df = df.astype(str).fillna('')
    col_map = {c.strip().lower().replace(' ', '_'): c for c in df.columns}
    rows = []
    for _, r in df.iterrows():
        row = {}
        for norm, orig in col_map.items():
            row[norm] = (r.get(orig) or '').strip()
        reg = row.get('reg_number') or row.get('student_id') or ''
        email = (row.get('email') or '').strip()
        role = _normalize_role(row.get('role') or 'student')
        if reg or (email and role != UserRole.STUDENT):
            if 'student_id' not in row and row.get('reg_number'):
                row['student_id'] = row['reg_number']
            rows.append(row)
    return rows


def _resolve_faculty_department_from_row(row: Dict[str, Any]):
    """Resolve faculty and department_fk from row keys: faculty, faculty_code, department, department_code."""
    from apps.academics.models import Faculty, Department
    faculty = None
    department_fk = None
    fac_code = (row.get('faculty_code') or row.get('faculty') or '').strip().upper()
    fac_name = (row.get('faculty_name') or '').strip()
    dept_code = (row.get('department_code') or row.get('department') or '').strip().upper()
    dept_name = (row.get('department_name') or '').strip()
    if fac_code:
        faculty = Faculty.objects.filter(code__iexact=fac_code).first()
    if not faculty and fac_name:
        faculty = Faculty.objects.filter(name__icontains=fac_name).first()
    if faculty and (dept_code or dept_name):
        if dept_code:
            department_fk = Department.objects.filter(faculty=faculty, code__iexact=dept_code).first()
        if not department_fk and dept_name:
            department_fk = Department.objects.filter(faculty=faculty, name__icontains=dept_name).first()
    return faculty, department_fk


def _create_user_from_row(
    row: Dict[str, Any],
    temp_password: str,
    *,
    scope_department_fk=None,
    scope_faculty=None,
    scope_is_hod: bool = False,
    scope_is_super_admin: bool = False,
) -> Tuple[User, str]:
    """
    Create one user from a row. Returns (user, temp_password).
    Row keys: reg_number/student_id, first_name, last_name, role, email, faculty, faculty_code, department, department_code.
    Enterprise: scope_department_fk/scope_faculty override when set (HOD/Faculty Admin bulk import).
    HOD can only create STUDENT and EXAMINER; all go to HOD's department.
    """
    reg = (row.get('reg_number') or row.get('student_id') or '').strip()
    first_name = (row.get('first_name') or '').strip() or 'N/A'
    last_name = (row.get('last_name') or '').strip() or 'N/A'
    role = _normalize_role(row.get('role') or 'student')
    email = (row.get('email') or '').strip().lower()

    if scope_is_hod and not scope_is_super_admin:
        if role in (UserRole.SUPER_ADMIN, UserRole.FACULTY_ADMIN, UserRole.DEPARTMENT_ADMIN, UserRole.HOD):
            raise ValueError(f'HOD cannot create role {role}. Only STUDENT and EXAMINER allowed.')
        department_fk = scope_department_fk
        faculty = getattr(scope_department_fk, 'faculty', None) if scope_department_fk else None
    elif scope_faculty and not scope_is_super_admin:
        if role == UserRole.SUPER_ADMIN:
            raise ValueError('Faculty Admin cannot create Super Admin.')
        faculty, _resolved_dept = _resolve_faculty_department_from_row(row)
        if not faculty or faculty.pk != scope_faculty.pk:
            faculty = scope_faculty
        department_fk = _resolved_dept if _resolved_dept and getattr(_resolved_dept, 'faculty_id', None) == getattr(faculty, 'pk', None) else None
    else:
        faculty, department_fk = _resolve_faculty_department_from_row(row)

    if role == UserRole.STUDENT:
        reg_upper = reg.upper()
        if not reg_upper:
            raise ValueError('Student ID (reg_number) is required for STUDENT role.')
        # Enterprise: HOD can only import students whose reg number department code matches their department
        if scope_is_hod and not scope_is_super_admin and scope_department_fk:
            reg_dept_code = department_code_from_student_id(reg_upper)
            scope_code = getattr(scope_department_fk, 'code', '')
            if reg_dept_code and reg_dept_code != scope_code:
                scope_name = getattr(scope_department_fk, 'name', '') or scope_code
                raise ValueError(
                    f'Student ID {reg_upper} indicates department {reg_dept_code}. '
                    f'You can only add students for your department ({scope_name} - {scope_code}). '
                    f'Use reg numbers like U22/FNS/{scope_code}/XXXX.'
                )
        if User.objects.filter(student_id__iexact=reg_upper).exists():
            raise ValueError(f'Student ID {reg_upper} already exists')
        user = User.objects.create_user(
            email=email,
            student_id=reg_upper,
            password=temp_password,
            role=UserRole.STUDENT,
            first_name=first_name,
            last_name=last_name,
        )
        # Set department/faculty for scoped import so student is correctly linked
        if scope_department_fk:
            user.department_fk = scope_department_fk
            user.faculty = getattr(scope_department_fk, 'faculty', None)
            user.department = getattr(scope_department_fk, 'name', '') or user.department or ''
            user.save(update_fields=['department_fk', 'faculty', 'department'])
    elif role in (UserRole.HOD, UserRole.DEPARTMENT_ADMIN, UserRole.FACULTY_ADMIN, UserRole.SUPER_ADMIN):
        if not email:
            email = f'{(reg or "staff").replace(" ", "").lower()}@staff.ibbul.edu.ng'
        if User.objects.filter(email__iexact=email).exists():
            raise ValueError(f'Email {email} already exists')
        if role == UserRole.SUPER_ADMIN:
            user = User.objects.create_superuser(email=email, password=temp_password)
        else:
            user = User.objects.create_user(email=email, password=temp_password, role=role)
            user.is_staff = True
            user.save(update_fields=['is_staff'])
        user.first_name = first_name
        user.last_name = last_name
        user.faculty = faculty
        user.department_fk = department_fk
        user.save(update_fields=['first_name', 'last_name', 'faculty', 'department_fk'])
    else:
        if not email:
            email = f'{(reg or "staff").replace(" ", "").lower()}@staff.ibbul.edu.ng'
        if User.objects.filter(email__iexact=email).exists():
            raise ValueError(f'Email {email} already exists')
        user = User.objects.create_user(
            email=email,
            password=temp_password,
            role=UserRole.EXAMINER,
            first_name=first_name,
            last_name=last_name,
        )
        user.faculty = faculty
        user.department_fk = department_fk
        user.save(update_fields=['faculty', 'department_fk'])
    user.is_first_login = True
    user.save(update_fields=['is_first_login'])
    return user, temp_password


def _hub_scope_info(request: HttpRequest):
    """
    Return scope info for hub/templates: scope_label, scope_note, is_scoped.
    HOD = department; Faculty Admin = faculty; Super Admin = no scope.
    """
    user = request.user
    if not getattr(user, 'is_authenticated', False):
        return {'scope_label': None, 'scope_note': None, 'is_scoped': False}
    role = getattr(user, 'role', None)
    role_str = str(role).upper() if role else ''
    if role_str in ('DEPARTMENT_ADMIN', 'HOD'):
        dept = getattr(user, 'department_fk', None)
        if dept:
            name = getattr(dept, 'name', None) or getattr(dept, 'code', None) or str(dept)
            return {
                'scope_label': name,
                'scope_note': 'Stats and data below are limited to your department.',
                'is_scoped': True,
                'scope_type': 'department',
            }
        return {'scope_label': 'Department', 'scope_note': 'Assign a department to your profile to see department-scoped data.', 'is_scoped': True, 'scope_type': 'department'}
    if role_str == 'FACULTY_ADMIN':
        fac = getattr(user, 'faculty', None)
        if fac:
            name = getattr(fac, 'name', None) or getattr(fac, 'code', None) or str(fac)
            return {
                'scope_label': name,
                'scope_note': 'Stats and data below are limited to your faculty.',
                'is_scoped': True,
                'scope_type': 'faculty',
            }
        return {'scope_label': 'Faculty', 'scope_note': 'Assign a faculty to your profile to see faculty-scoped data.', 'is_scoped': True, 'scope_type': 'faculty'}
    return {'scope_label': None, 'scope_note': None, 'is_scoped': False}


@staff_member_required
def users_accounts_hub_view(request: HttpRequest) -> HttpResponse:
    """
    Users / Accounts hub: scoped stats and menu.
    HOD/Faculty Admin see only their department/faculty stats; Super Admin sees all.
    """
    from django.utils import timezone as tz
    # Enterprise: all stats scoped by role (filter_by_scope)
    scoped_users = filter_by_scope(User.objects.all(), request.user, request)
    total_users = scoped_users.count()
    students = scoped_users.filter(role=UserRole.STUDENT).count()
    staff = total_users - students
    active = scoped_users.filter(is_active=True).count()
    never_logged_in = scoped_users.filter(last_login__isnull=True).count()
    must_change_password = scoped_users.filter(is_first_login=True).count()
    since_7d = tz.now() - timezone.timedelta(days=7)
    scoped_user_ids = list(scoped_users.values_list('pk', flat=True))
    logins_7d = (
        AuditLog.objects.filter(
            action=AuditLog.Action.LOGIN_SUCCESS,
            created_at__gte=since_7d,
            user_id__in=scoped_user_ids,
        ).count()
        if scoped_user_ids
        else 0
    )
    scope_info = _hub_scope_info(request)
    context = {
        **admin.site.each_context(request),
        'title': 'Users / Accounts',
        'stats': {
            'total_users': total_users,
            'students': students,
            'staff': staff,
            'active': active,
            'never_logged_in': never_logged_in,
            'must_change_password': must_change_password,
            'logins_7d': logins_7d,
        },
        'scope_info': scope_info,
    }
    return render(request, 'admin/accounts/users_accounts_hub.html', context)


@staff_member_required
def add_student_view(request: HttpRequest) -> HttpResponse:
    """
    Enterprise: Add one student (manual). Scoped: HOD adds to their department.
    Permission: add_user_scoped or add_user (HOD, Faculty Admin, Super Admin).
    Full form: reg_number (validated), first_name, last_name, department display, level.
    Temp password generated; one-time CSV download (TTL). All actions audited.
    """
    if not user_can_bulk_import_users(request.user) and not request.user.has_perm('accounts.add_user_scoped') and not request.user.has_perm('accounts.add_user'):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You do not have permission to add users.')
    if request.method == 'GET':
        scope_info = _hub_scope_info(request)
        # For HOD: show department name so they see scope
        dept_display = None
        if scope_info.get('is_scoped') and scope_info.get('scope_type') == 'department':
            dept_display = scope_info.get('scope_label')
        context = {
            **admin.site.each_context(request),
            'title': 'Add student',
            'scope_info': scope_info,
            'department_display': dept_display,
        }
        return render(request, 'admin/accounts/add_student.html', context)

    reg_number = (request.POST.get('reg_number') or '').strip().upper()
    first_name = (request.POST.get('first_name') or '').strip()
    last_name = (request.POST.get('last_name') or '').strip()
    department = (request.POST.get('department') or '').strip()
    level = (request.POST.get('level') or '').strip()

    # Validation: required fields
    if not reg_number:
        messages.error(request, 'Registration number is required.')
        return redirect('admin_add_student')
    if not first_name:
        messages.error(request, 'First name is required.')
        return redirect('admin_add_student')
    if not last_name:
        messages.error(request, 'Last name is required.')
        return redirect('admin_add_student')

    # Validate student ID format (U22/FNS/CSC/XXXX)
    try:
        from django.core.exceptions import ValidationError
        validate_student_id_format(reg_number)
    except ValidationError as e:
        err_msg = (list(e.messages)[0] if getattr(e, 'messages', None) else str(e)) or 'Invalid registration number format.'
        messages.error(request, err_msg)
        return redirect('admin_add_student')
    except Exception as e:
        messages.error(request, 'Registration number must follow format: U22/FNS/CSC/0001 (e.g. U22/FNS/CSC/0001).')
        return redirect('admin_add_student')

    # Enterprise: HOD can only add students whose reg number matches their department (e.g. GLG for Geology)
    if is_hod(request.user) and not is_super_admin(request.user):
        from apps.academics.models import Department
        hod_dept = getattr(request.user, 'department_fk', None)
        if hod_dept:
            reg_dept_code = department_code_from_student_id(reg_number)
            if reg_dept_code and reg_dept_code != getattr(hod_dept, 'code', ''):
                dept_from_reg = Department.objects.filter(code=reg_dept_code).first()
                dept_label = f'{dept_from_reg.name} ({reg_dept_code})' if dept_from_reg else reg_dept_code
                hod_label = getattr(hod_dept, 'name', '') or getattr(hod_dept, 'code', '')
                messages.error(
                    request,
                    f'Registration number {reg_number} indicates department {dept_label}. '
                    f'You can only add students for your department ({hod_label}). Use a reg number like U22/FNS/{getattr(hod_dept, "code", "")}/XXXX.'
                )
                return redirect('admin_add_student')

    if User.objects.filter(student_id__iexact=reg_number).exists():
        messages.error(request, f'A student with registration number {reg_number} already exists.')
        return redirect('admin_add_student')

    try:
        temp_password = _generate_temp_password()
        extra = {'department': department or '', 'level': level or ''}
        if getattr(request.user, 'department_fk_id', None):
            extra['department_fk_id'] = request.user.department_fk_id
            if getattr(request.user, 'department_fk', None):
                dept = request.user.department_fk
                extra['department'] = getattr(dept, 'name', '') or extra['department']
                extra['faculty_id'] = getattr(dept, 'faculty_id', None)
        user = User.objects.create_user(
            email=None,
            student_id=reg_number,
            password=temp_password,
            role=UserRole.STUDENT,
            first_name=first_name or 'N/A',
            last_name=last_name or 'N/A',
            **extra,
        )
        user.is_first_login = True
        user.save(update_fields=['is_first_login'])

        log_audit(
            AuditLog.Action.ADMIN_ACTION,
            request=request,
            user=request.user,
            extra={'action': 'manual_add_student', 'reg_number': reg_number, 'user_id': user.pk},
        )

        export_id = str(uuid.uuid4())
        cache.set(
            f'temp_passwords_export_{export_id}',
            {
                'user_id': request.user.id,
                'rows': [{'reg_number': reg_number, 'temp_password': temp_password}],
                'after_download_redirect': 'admin_add_student',
            },
            timeout=TEMP_PASSWORD_EXPORT_TTL,
        )

        scope_info = _hub_scope_info(request)
        context = {
            **admin.site.each_context(request),
            'title': 'Student added',
            'reg_number': reg_number,
            'full_name': user.get_full_name(),
            'export_id': export_id,
            'export_ttl_minutes': TEMP_PASSWORD_EXPORT_TTL // 60,
            'scope_info': scope_info,
        }
        return render(request, 'admin/accounts/add_student_result.html', context)
    except Exception as e:
        logger.exception('Add student failed: %s', e)
        messages.error(request, str(e))
        return redirect('admin_add_student')


@staff_member_required
def import_users_view(request: HttpRequest) -> HttpResponse:
    """
    Enterprise bulk import: upload CSV or Excel to create users with temporary passwords.
    Permission: accounts.bulk_import_users (HOD, Faculty Admin, Super Admin).
    HOD: only STUDENT and EXAMINER; all created users are scoped to HOD's department.
    GET: show form. POST: process file, create users, show result.
    """
    if not user_can_bulk_import_users(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You do not have permission to bulk import users.')

    if request.method == 'GET':
        context = {
            **admin.site.each_context(request),
            'title': 'Import users (CSV/Excel)',
        }
        return render(request, 'admin/accounts/import_users.html', context)

    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Please select a CSV or Excel file.')
        return redirect('admin_import_users')

    name = (file.name or '').lower()
    try:
        content = file.read()
    except Exception as e:
        messages.error(request, f'Could not read file: {e}')
        return redirect('admin_import_users')

    if name.endswith('.csv'):
        rows = _parse_csv(content)
    elif name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            rows = _parse_excel(content)
        except Exception as e:
            messages.error(request, f'Invalid Excel file: {e}')
            return redirect('admin_import_users')
    else:
        messages.error(request, 'File must be .csv or .xlsx')
        return redirect('admin_import_users')

    if not rows:
        messages.warning(request, 'No rows with reg_number (students) or email (staff) found in file.')
        return redirect('admin_import_users')

    # Enterprise: HOD/Faculty Admin scope for bulk import
    scope_department_fk = None
    scope_faculty = None
    if is_hod(request.user) and not is_super_admin(request.user):
        dept_id = get_hod_department_id(request.user)
        if dept_id:
            from apps.academics.models import Department
            scope_department_fk = Department.objects.filter(pk=dept_id).select_related('faculty').first()
    elif getattr(request.user, 'role', None) in (UserRole.FACULTY_ADMIN, 'FACULTY_ADMIN') and getattr(request.user, 'faculty_id', None):
        from apps.academics.models import Faculty
        scope_faculty = Faculty.objects.filter(pk=request.user.faculty_id).first()

    created = []
    errors = []
    for i, row in enumerate(rows):
        reg = row.get('reg_number') or row.get('student_id') or ''
        email = (row.get('email') or '').strip()
        role = _normalize_role(row.get('role') or 'student')
        identifier = reg or (email if role != UserRole.STUDENT else '') or ''
        if not identifier:
            continue
        try:
            temp_password = _generate_temp_password()
            user, pwd = _create_user_from_row(
                row,
                temp_password,
                scope_department_fk=scope_department_fk,
                scope_faculty=scope_faculty,
                scope_is_hod=is_hod(request.user),
                scope_is_super_admin=is_super_admin(request.user),
            )
            created.append({
                'reg_number': user.student_id or user.email,
                'full_name': user.get_full_name(),
                'role': user.get_role_display(),
                'temp_password': pwd,
            })
        except Exception as e:
            errors.append(f'Row {i + 2} ({identifier}): {e}')
            logger.exception('Import user failed: %s', row)

    log_audit(
        AuditLog.Action.ADMIN_USER_IMPORT,
        request=request,
        user=request.user,
        extra={'created': len(created), 'errors': len(errors), 'file': name},
    )

    export_id = None
    if created:
        export_id = str(uuid.uuid4())
        cache.set(
            f'temp_passwords_export_{export_id}',
            {
                'user_id': request.user.id,
                'rows': [{'reg_number': c['reg_number'], 'temp_password': c['temp_password']} for c in created],
                'after_download_redirect': 'admin_import_users',
            },
            timeout=TEMP_PASSWORD_EXPORT_TTL,
        )

    context = {
        **admin.site.each_context(request),
        'title': 'Import result',
        'created': created,
        'errors': errors,
        'export_id': export_id,
        'export_ttl_minutes': TEMP_PASSWORD_EXPORT_TTL // 60,
    }
    return render(request, 'admin/accounts/import_users_result.html', context)


@staff_member_required
def import_users_template_download(request: HttpRequest) -> HttpResponse:
    """
    Enterprise: download CSV template for bulk user import.
    Permission: accounts.bulk_import_users. Headers match _parse_csv expectations.
    """
    if not user_can_bulk_import_users(request.user):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('You do not have permission to download the import template.')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'reg_number', 'first_name', 'last_name', 'role', 'email',
        'faculty_code', 'department_code',
    ])
    writer.writerow([
        'U22/FNS/CSC/0001', 'John', 'Doe', 'student', '',
        '', '',
    ])
    writer.writerow([
        '', 'Jane', 'Smith', 'lecturer', 'jane.smith@ibbul.edu.ng',
        'FNS', 'CSC',
    ])
    return response


def _user_management_redirect(request: HttpRequest, query_string: str = '') -> HttpResponse:
    """Redirect to user management, preserving query string (from GET or POST hidden params)."""
    if not query_string and request.method == 'POST':
        from urllib.parse import urlencode
        params = {}
        for key in ('search', 'role', 'status', 'first_login', 'department', 'last_login', 'sort', 'page_size'):
            val = request.POST.get(key, '').strip()
            if val:
                params[key] = val
        if params:
            query_string = urlencode(params)
    url = reverse('admin_user_management')
    if query_string:
        url = f'{url}?{query_string}'
    return redirect(url)


@staff_member_required
def user_management_view(request: HttpRequest) -> HttpResponse:
    """
    Enterprise user management: search, filters, pagination, sort, single and bulk actions.
    GET: list with search, role, status, first_login, department, last_login, sort, page, page_size.
    POST: single (user_id) or bulk (user_ids[]) actions with permission checks; all audited.
    """
    query_string = request.GET.urlencode()
    redirect_with_query = lambda: _user_management_redirect(request, query_string)

    # ----- POST: single or bulk actions -----
    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        user_id = request.POST.get('user_id')
        user_ids = request.POST.getlist('user_ids[]') or request.POST.getlist('user_ids')

        # Single-user action
        if user_id and not user_ids:
            user = get_object_or_404(User, pk=user_id)
            if action == 'reset_to_temp_password':
                temp_password = _generate_temp_password()
                user.set_password(temp_password)
                user.is_first_login = True
                user.last_password_change = None
                user.save(update_fields=['password', 'is_first_login', 'last_password_change'])
                log_audit(
                    AuditLog.Action.ADMIN_PASSWORD_RESET,
                    request=request,
                    user=request.user,
                    extra={'target_user_id': user.pk, 'action': 'reset_to_temp_password'},
                )
                reg = user.student_id or user.email or str(user.pk)
                export_id = str(uuid.uuid4())
                cache.set(
                    f'temp_passwords_export_{export_id}',
                    {
                        'user_id': request.user.id,
                        'rows': [{'reg_number': reg, 'temp_password': temp_password}],
                        'after_download_redirect': 'admin_user_management',
                    },
                    timeout=TEMP_PASSWORD_EXPORT_TTL,
                )
                context = {
                    **admin.site.each_context(request),
                    'title': 'Password reset to temporary',
                    'user_identifier': reg,
                    'export_id': export_id,
                    'export_ttl_minutes': TEMP_PASSWORD_EXPORT_TTL // 60,
                }
                return render(request, 'admin/accounts/reset_temp_result.html', context)
            if action == 'reset_password':
                new_password = request.POST.get('new_password', '').strip()
                if len(new_password) < 8:
                    messages.error(request, 'New password must be at least 8 characters.')
                    return redirect_with_query()
                user.set_password(new_password)
                user.is_first_login = True
                user.last_password_change = None
                user.save(update_fields=['password', 'is_first_login', 'last_password_change'])
                log_audit(
                    AuditLog.Action.ADMIN_PASSWORD_RESET,
                    request=request,
                    user=request.user,
                    extra={'target_user_id': user.pk, 'target_identifier': user.student_id or user.email},
                )
                messages.success(request, f'Password reset for {user.get_full_name()}. They must change it on next login.')
                return redirect_with_query()
            if action == 'activate':
                user.is_active = True
                user.save(update_fields=['is_active'])
                log_audit(
                    AuditLog.Action.ADMIN_ACTIVATE_DEACTIVATE,
                    request=request,
                    user=request.user,
                    extra={'target_user_id': user.pk, 'action': 'activate'},
                )
                messages.success(request, f'{user.get_full_name()} is now active.')
                return redirect_with_query()
            if action == 'deactivate':
                if not _can_deactivate(request.user, user):
                    messages.error(request, 'You cannot deactivate this user (yourself or staff unless Super Admin).')
                    return redirect_with_query()
                user.is_active = False
                user.save(update_fields=['is_active'])
                log_audit(
                    AuditLog.Action.ADMIN_ACTIVATE_DEACTIVATE,
                    request=request,
                    user=request.user,
                    extra={'target_user_id': user.pk, 'action': 'deactivate'},
                )
                messages.success(request, f'{user.get_full_name()} is now deactivated.')
                return redirect_with_query()
            messages.error(request, 'Invalid action.')
            return redirect_with_query()

        # Bulk actions
        if user_ids:
            users_to_act = list(User.objects.filter(pk__in=user_ids))
            if action == 'bulk_activate':
                for u in users_to_act:
                    u.is_active = True
                    u.save(update_fields=['is_active'])
                    log_audit(
                        AuditLog.Action.ADMIN_ACTIVATE_DEACTIVATE,
                        request=request,
                        user=request.user,
                        extra={'target_user_id': u.pk, 'action': 'activate'},
                    )
                messages.success(request, f'{len(users_to_act)} user(s) activated.')
                return redirect_with_query()
            if action == 'bulk_deactivate':
                allowed = [u for u in users_to_act if _can_deactivate(request.user, u)]
                skipped = len(users_to_act) - len(allowed)
                for u in allowed:
                    u.is_active = False
                    u.save(update_fields=['is_active'])
                    log_audit(
                        AuditLog.Action.ADMIN_ACTIVATE_DEACTIVATE,
                        request=request,
                        user=request.user,
                        extra={'target_user_id': u.pk, 'action': 'deactivate'},
                    )
                if allowed:
                    messages.success(request, f'{len(allowed)} user(s) deactivated.')
                if skipped:
                    messages.warning(request, f'{skipped} user(s) skipped (yourself or staff).')
                return redirect_with_query()
            if action == 'bulk_reset_temp':
                rows = []
                for u in users_to_act:
                    temp_password = _generate_temp_password()
                    u.set_password(temp_password)
                    u.is_first_login = True
                    u.last_password_change = None
                    u.save(update_fields=['password', 'is_first_login', 'last_password_change'])
                    reg = u.student_id or u.email or str(u.pk)
                    rows.append({'reg_number': reg, 'temp_password': temp_password})
                    log_audit(
                        AuditLog.Action.ADMIN_PASSWORD_RESET,
                        request=request,
                        user=request.user,
                        extra={'target_user_id': u.pk, 'action': 'reset_to_temp_password'},
                    )
                export_id = str(uuid.uuid4())
                cache.set(
                    f'temp_passwords_export_{export_id}',
                    {
                        'user_id': request.user.id,
                        'rows': rows,
                        'after_download_redirect': 'admin_user_management',
                    },
                    timeout=TEMP_PASSWORD_EXPORT_TTL,
                )
                context = {
                    **admin.site.each_context(request),
                    'title': 'Bulk password reset – download',
                    'export_id': export_id,
                    'export_ttl_minutes': TEMP_PASSWORD_EXPORT_TTL // 60,
                    'row_count': len(rows),
                }
                return render(request, 'admin/accounts/reset_temp_result.html', context)
        messages.error(request, 'Select at least one user.')
        return redirect_with_query()

    # ----- GET: filtered, sorted, paginated list -----
    qs, filters_used = _build_user_queryset(request)
    try:
        page_size = int(request.GET.get('page_size') or USER_MANAGEMENT_PAGE_SIZE)
    except (TypeError, ValueError):
        page_size = USER_MANAGEMENT_PAGE_SIZE
    page_size = max(1, min(page_size, max(USER_MANAGEMENT_PAGE_SIZES)))
    paginator = Paginator(qs, page_size)
    page_number = max(1, int(request.GET.get('page', 1) or 1))
    page = paginator.get_page(page_number)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('partial') == '1'

    get_copy = request.GET.copy()
    _ = get_copy.pop('page', None)
    query_no_page = get_copy.urlencode()
    get_copy.pop('sort', None)
    sort_base = get_copy.urlencode()

    context = {
        **admin.site.each_context(request),
        'title': 'User management',
        'users': page,
        'filters': filters_used,
        'page_sizes': USER_MANAGEMENT_PAGE_SIZES,
        'page_size': page_size,
        'role_choices': UserRole.choices,
        'query_no_page': query_no_page,
        'sort_base': sort_base,
    }
    if is_ajax:
        return render(request, 'admin/accounts/user_management_table_partial.html', context)
    return render(request, 'admin/accounts/user_management.html', context)


def _is_hod(user) -> bool:
    """True if user is Department Admin (HOD)."""
    role = getattr(user, 'role', None)
    return str(role).upper() in ('DEPARTMENT_ADMIN', 'HOD') if role else False


def _audit_base_queryset(request: HttpRequest):
    """
    Base AuditLog queryset: scoped by role.
    HOD: department only. Faculty Admin: faculty only. Super Admin: all.
    """
    qs = AuditLog.objects.all().select_related('user', 'scope_department', 'scope_faculty').order_by('-created_at')
    user = request.user
    if not getattr(user, 'is_authenticated', False):
        return qs
    user_role = getattr(user, 'role', None)
    role_str = (getattr(user_role, 'value', None) or str(user_role).upper() if user_role else '') or ''
    if role_str and '.' in role_str:
        role_str = getattr(user_role, 'value', role_str) or role_str
    # HOD: department-scoped
    if role_str in ('DEPARTMENT_ADMIN', 'HOD') and getattr(user, 'department_fk_id', None):
        dept_id = user.department_fk_id
        qs = qs.filter(
            Q(scope_department_id=dept_id)
            | Q(scope_department_id__isnull=True, user__department_fk_id=dept_id)
        )
        return qs
    # Faculty Admin: faculty-scoped (logs where scope_faculty = their faculty or user in their faculty)
    if role_str == 'FACULTY_ADMIN' and getattr(user, 'faculty_id', None):
        fac_id = user.faculty_id
        qs = qs.filter(
            Q(scope_faculty_id=fac_id)
            | Q(scope_faculty_id__isnull=True, user__faculty_id=fac_id)
            | Q(scope_faculty_id__isnull=True, user__department_fk__faculty_id=fac_id)
        )
    return qs


def _build_audit_queryset(request: HttpRequest):
    """
    Build filtered, searchable AuditLog queryset from GET params.
    HOD sees only logs scoped to their department (scope_department = user.department_fk).
    Returns (queryset, filters_dict for template). Newest first.
    """
    qs = _audit_base_queryset(request)
    filters_used = {}

    # Global search: reg number, email, admin username, action keyword
    search = (request.GET.get('search') or '').strip()
    if search:
        term = search
        qs = qs.filter(
            Q(identifier__icontains=term)
            | Q(action__icontains=term)
            | Q(user__email__icontains=term)
            | Q(user__student_id__icontains=term)
        )
        filters_used['search'] = search

    # Action type
    action = request.GET.get('action', '').strip()
    if action and action in dict(AuditLog.Action.choices):
        qs = qs.filter(action=action)
        filters_used['action'] = action

    # Actor role: Admin (HOD), Examiner, Student, System (no user)
    role = request.GET.get('role', '').strip()
    if role == 'system':
        qs = qs.filter(user__isnull=True)
        filters_used['role'] = 'system'
    elif role and role in dict(UserRole.choices):
        qs = qs.filter(user__role=role)
        filters_used['role'] = role

    # User (actor) filter: by user id for links from other pages
    user_id = request.GET.get('user_id', '').strip()
    if user_id:
        try:
            uid = int(user_id)
            qs = qs.filter(user_id=uid)
            filters_used['user_id'] = user_id
        except ValueError:
            pass

    # Date range: today, 7d, 30d, or custom from/to
    date_range = request.GET.get('date_range', '').strip().lower()
    now = timezone.now()
    if date_range == 'today':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start)
        filters_used['date_range'] = 'today'
    elif date_range == '7d':
        start = now - timezone.timedelta(days=7)
        qs = qs.filter(created_at__gte=start)
        filters_used['date_range'] = '7d'
    elif date_range == '30d':
        start = now - timezone.timedelta(days=30)
        qs = qs.filter(created_at__gte=start)
        filters_used['date_range'] = '30d'
    else:
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        if date_from:
            try:
                from django.utils.dateparse import parse_datetime
                start = parse_datetime(date_from + ' 00:00:00') if len(date_from) <= 10 else parse_datetime(date_from)
                if start:
                    start = timezone.make_aware(start) if timezone.is_naive(start) else start
                    qs = qs.filter(created_at__gte=start)
                    filters_used['date_from'] = date_from
            except Exception:
                pass
        if date_to:
            try:
                from django.utils.dateparse import parse_datetime
                end = parse_datetime(date_to + ' 23:59:59') if len(date_to) <= 10 else parse_datetime(date_to)
                if end:
                    end = timezone.make_aware(end) if timezone.is_naive(end) else end
                    qs = qs.filter(created_at__lte=end)
                    filters_used['date_to'] = date_to
            except Exception:
                pass

    # IP address filter
    ip = request.GET.get('ip', '').strip()
    if ip:
        qs = qs.filter(ip_address=ip)
        filters_used['ip'] = ip

    return qs, filters_used


def _audit_log_stats(request: HttpRequest):
    """
    Return summary counters for audit log dashboard.
    HOD: counts are department-scoped (scope_department = user.department_fk).
    """
    base = _audit_base_queryset(request)
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    result_actions = (
        AuditLog.Action.RESULT_UPLOAD_STARTED,
        AuditLog.Action.RESULT_UPLOAD_COMPLETED,
        AuditLog.Action.RESULT_BATCH_APPROVED,
        AuditLog.Action.RESULT_BATCH_REJECTED,
        AuditLog.Action.RESULT_SINGLE_APPROVED,
        AuditLog.Action.RESULT_SINGLE_REJECTED,
        AuditLog.Action.RESULT_CREATED,
        AuditLog.Action.RESULT_UPDATED,
        AuditLog.Action.RESULT_DELETED,
        AuditLog.Action.RESULT_MANUAL_ENTRY,
        AuditLog.Action.RESULT_IMPORT_REPORT_DOWNLOAD,
    )
    return {
        'total': base.count(),
        'logins_today': base.filter(
            action=AuditLog.Action.LOGIN_SUCCESS,
            created_at__gte=today_start,
        ).count(),
        'failed_logins': base.filter(action=AuditLog.Action.LOGIN_FAILED).count(),
        'admin_actions': base.filter(
            action__in=(
                AuditLog.Action.ADMIN_ACTION,
                AuditLog.Action.ADMIN_USER_IMPORT,
                AuditLog.Action.ADMIN_PASSWORD_RESET,
                AuditLog.Action.ADMIN_ACTIVATE_DEACTIVATE,
                AuditLog.Action.USER_CREATED,
                AuditLog.Action.USER_UPDATED,
            ),
        ).count(),
        'course_actions': base.filter(
            action__in=(
                AuditLog.Action.COURSE_CREATED,
                AuditLog.Action.COURSE_UPDATED,
                AuditLog.Action.COURSE_DELETED,
                AuditLog.Action.COURSE_ASSIGNMENT_CREATED,
                AuditLog.Action.COURSE_ASSIGNMENT_UPDATED,
                AuditLog.Action.COURSE_ASSIGNMENT_DELETED,
            ),
        ).count(),
        'result_actions': base.filter(action__in=result_actions).count(),
    }


@staff_member_required
def audit_log_view(request: HttpRequest) -> HttpResponse:
    """
    Professional audit log UI: read-only table, summary stats, filters, search, export.
    No delete for normal admins; Super Admin can delete from default admin if needed.
    """
    qs, filters_used = _build_audit_queryset(request)
    try:
        page_size = int(request.GET.get('page_size') or AUDIT_LOG_PAGE_SIZE)
    except (TypeError, ValueError):
        page_size = AUDIT_LOG_PAGE_SIZE
    page_size = max(1, min(page_size, max(AUDIT_LOG_PAGE_SIZES)))
    paginator = Paginator(qs, page_size)
    page_number = max(1, int(request.GET.get('page', 1) or 1))
    page = paginator.get_page(page_number)

    stats = _audit_log_stats(request)
    action_choices = list(AuditLog.Action.choices)
    role_choices_with_system = [('', 'All'), ('system', 'System')] + list(UserRole.choices)

    get_copy = request.GET.copy()
    get_copy.pop('page', None)
    query_no_page = get_copy.urlencode()
    export_base = query_no_page  # export respects current filters

    # Department-scoped for HOD: show scope label in template
    audit_scope_label = None
    if _is_hod(request.user):
        dept = getattr(request.user, 'department_fk', None) or getattr(request.user, 'department', None)
        if dept:
            audit_scope_label = getattr(dept, 'name', None) if hasattr(dept, 'name') else str(dept)
        else:
            audit_scope_label = 'Department'

    scope_info = _hub_scope_info(request)

    context = {
        **admin.site.each_context(request),
        'title': 'Audit logs',
        'audit_page': page,
        'filters': filters_used,
        'page_sizes': AUDIT_LOG_PAGE_SIZES,
        'page_size': page_size,
        'action_choices': action_choices,
        'role_choices': role_choices_with_system,
        'query_no_page': query_no_page,
        'export_base': export_base,
        'stats': stats,
        'is_superuser': getattr(request.user, 'is_superuser', False),
        'audit_scope_department': _is_hod(request.user),
        'audit_scope_label': audit_scope_label,
        'scope_info': scope_info,
    }
    return render(request, 'admin/accounts/audit_log.html', context)


@staff_member_required
def audit_log_export_view(request: HttpRequest) -> HttpResponse:
    """
    Export audit logs as CSV, Excel, or PDF. Respects current filters (same GET params).
    """
    fmt = (request.GET.get('format') or 'csv').strip().lower()
    qs, _ = _build_audit_queryset(request)
    # Limit export size for safety
    qs = qs[:10000]
    rows = []
    for log in qs:
        actor = 'System'
        if log.user:
            actor = log.user.get_role_display() if hasattr(log.user, 'get_role_display') else (log.user.student_id or log.user.email or '—')
        created = timezone.localtime(log.created_at).strftime('%Y-%m-%d %H:%M:%S %Z') if log.created_at else ''
        rows.append({
            'action': log.get_action_display() if hasattr(log, 'get_action_display') else log.action,
            'actor': actor,
            'target_identifier': log.identifier or '',
            'ip_address': str(log.ip_address) if log.ip_address else '',
            'device_browser': (log.user_agent or '')[:200],
            'timestamp': created,
        })
    if fmt == 'excel' or fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Audit Logs'[:31]  # Excel sheet name limit 31 chars
            headers = ['Action', 'Actor', 'Target identifier', 'IP address', 'Device / Browser', 'Timestamp']
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, r in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=(r.get('action') or '')[:32767])
                ws.cell(row=row_idx, column=2, value=(r.get('actor') or '')[:32767])
                ws.cell(row=row_idx, column=3, value=(r.get('target_identifier') or '')[:32767])
                ws.cell(row=row_idx, column=4, value=(r.get('ip_address') or '')[:32767])
                ws.cell(row=row_idx, column=5, value=(r.get('device_browser') or '')[:32767])
                ws.cell(row=row_idx, column=6, value=(r.get('timestamp') or '')[:32767])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            return resp
        except Exception as e:
            logger.exception('Excel export failed: %s', e)
            messages.error(request, 'Excel export failed. Try CSV.')
            return redirect(request.META.get('HTTP_REFERER') or reverse('admin_audit_logs'))
    if fmt == 'pdf' or fmt == 'html':
        # Optional PDF: export as HTML; user can open and use browser "Print to PDF"
        from django.template.loader import render_to_string
        html = render_to_string('admin/accounts/audit_log_export_pdf.html', {'rows': rows[:500], 'generated': timezone.now()})
        resp = HttpResponse(html, content_type='text/html; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M")}.html"'
        return resp
    # CSV (default)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=['action', 'actor', 'target_identifier', 'ip_address', 'device_browser', 'timestamp'])
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    return resp


@staff_member_required
def temp_passwords_export_download(request: HttpRequest, export_id: str) -> HttpResponse:
    """
    One-time download of reg_number → temporary_password CSV.
    Accessible only to the admin who performed the upload; downloadable once or until TTL.
    After download: delete from cache and log export event.
    """
    cache_key = f'temp_passwords_export_{export_id}'
    data = cache.get(cache_key)
    if not data:
        messages.error(request, 'This export has expired or was already downloaded.')
        return redirect('admin_import_users')
    if data.get('user_id') != request.user.id:
        messages.error(request, 'You are not authorized to download this export.')
        try:
            return redirect(data.get('after_download_redirect') or 'admin_import_users')
        except Exception:
            return redirect('admin_import_users')
    rows = data.get('rows') or []
    after_redirect = data.get('after_download_redirect') or 'admin_import_users'
    cache.delete(cache_key)
    log_audit(
        AuditLog.Action.ADMIN_ACTION,
        request=request,
        user=request.user,
        extra={'action': 'temp_password_export_download', 'export_id': export_id, 'row_count': len(rows)},
    )
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=['reg_number', 'temporary_password'])
    writer.writeheader()
    for r in rows:
        writer.writerow({'reg_number': r.get('reg_number', ''), 'temporary_password': r.get('temp_password', '')})
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="temp_passwords_{export_id[:8]}.csv"'
    return resp
